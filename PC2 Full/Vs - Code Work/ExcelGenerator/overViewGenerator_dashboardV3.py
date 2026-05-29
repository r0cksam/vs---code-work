"""
generate_dashboard.py  —  Enterprise Edition (Averages Restored)
──────────────────────────────────────────────────────────────────────────────────
Key Updates:
- RESTORED: Dynamic dashed average lines on all charts.
- MAINTAINED: Fix for bracket syntax error on all percentage indices.
- MAINTAINED: Bound-safe column fetcher prevents IndexError on truncated Excel rows.
- MAINTAINED: Python-side date parsing normalizes regional text dates.
- MAINTAINED: Date range filter (7D/30D/ALL) with full chart & table syncing.

Usage:  python generate_dashboard.py
"""

import sys, json, urllib.request, csv
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","--quiet"])
    import openpyxl

# ╔══════════════════════════════════════════════════════════╗
# ║  SET THESE PATHS                                         ║
# ╚══════════════════════════════════════════════════════════╝
XLSX_PATH = Path(r"Y:\Veto Logs Backup\Dashboards\OverView\overview_report.xlsx")
HTML_OUT  = Path(r"Y:\Veto Logs Backup\Dashboards\OverView\overview_dashboard.html")
DEVICE_SNAPSHOT_CSV = Path(r"Y:\Veto Logs Backup\Dashboards\OverView\device_snapshot.csv")
DEVICE_DAILY_CSV    = Path(r"Y:\Veto Logs Backup\Dashboards\OverView\device_daily.csv")
# ══════════════════════════════════════════════════════════

if len(sys.argv) == 3:
    XLSX_PATH, HTML_OUT = Path(sys.argv[1]), Path(sys.argv[2])

COL = {
    'DATE': 2, 'BYTES': 3, 'ROWS': 4, 'IP_ROWS': 5,
    'DIST_IP': 6, 'DIST_IPUA': 7, 'DIST_DEV': 8, 'DIST_SESS': 9,
    'DIST_IP_R2': 10, 'DIST_IPUA_R2': 11, 
    'SESS_AVAIL': 12, 'SESS_NA': 13, 'SESS_NONE': 14,
    'PCT_IP': 15, 'PCT_IPUA': 16, 'PCT_SESS': 17, 'PCT_SESSNA': 18, 'PCT_NONE': 19
}
DATA_START_ROW = 13

def sn(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def get_val(row, col_num):
    """Safely fetch cell value by 1-based index to prevent unexpected IndexError on short rows"""
    idx = col_num - 1
    return row[idx] if idx < len(row) else None

def read_device_data(snapshot_path: Path, daily_path: Path):
    snapshot_rows = []
    device_idx = {}
    first_seen_counts = {}
    daily_map = {}
    if snapshot_path.exists():
        with snapshot_path.open("r", encoding="utf-8", newline="") as f:
            for r in csv.DictReader(f):
                device_id = r.get("device_id", "")
                if not device_id:
                    continue
                device_idx[device_id] = len(device_idx)
                first_seen = r.get("first_seen_utc_date", "")
                if first_seen:
                    first_seen_counts[first_seen] = first_seen_counts.get(first_seen, 0) + 1
                snapshot_rows.append({
                    "device_id": device_id,
                    "first_seen": r.get("first_seen_utc_date", ""),
                    "last_seen": r.get("last_seen_utc_date", ""),
                    "days_seen": int(sn(r.get("days_seen"))),
                    "total_rows": int(sn(r.get("total_rows"))),
                    "sessions": int(sn(r.get("distinct_sessions_day_sum"))),
                })
    if daily_path.exists():
        with daily_path.open("r", encoding="utf-8", newline="") as f:
            for r in csv.DictReader(f):
                d = r.get("utc_date", "")
                device_id = r.get("device_id", "")
                if not d or not device_id:
                    continue
                idx = device_idx.get(device_id)
                if idx is None:
                    idx = len(device_idx)
                    device_idx[device_id] = idx
                item = daily_map.setdefault(d, {"date": d, "devices": set(), "rows": 0, "sessions": 0})
                item["devices"].add(idx)
                item["rows"] += int(sn(r.get("rows_on_date")))
                item["sessions"] += int(sn(r.get("distinct_sessions")))
    daily_rows = []
    for k in sorted(daily_map):
        item = daily_map[k]
        devices = sorted(item["devices"])
        daily_rows.append({
            "date": k,
            "devices": devices,
            "active_devices": len(devices),
            "rows": item["rows"],
            "sessions": item["sessions"],
        })
    return snapshot_rows, daily_rows, first_seen_counts

CHARTJS_CACHE = Path(__file__).parent / ".chartjs_cache.js"
CHARTJS_URL   = "https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"

def get_chartjs():
    if CHARTJS_CACHE.exists():
        return CHARTJS_CACHE.read_text(encoding="utf-8")
    print("  Downloading Chart.js once...", end=" ", flush=True)
    try:
        req = urllib.request.Request(CHARTJS_URL, headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=20) as r:
            js = r.read().decode("utf-8")
        CHARTJS_CACHE.write_text(js, encoding="utf-8")
        print(f"done ({len(js)//1024}KB cached)")
        return js
    except Exception as e:
        print(f"failed: {e}\n  Charts will be disabled.")
        return None

print(f"Reading: {XLSX_PATH}")
if not XLSX_PATH.exists():
    print(f"  ERROR: File not found: {XLSX_PATH}"); sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=False, data_only=True)
ws = wb.active

meta = {}
data_time_range = ""
exclude_keywords = ["report generated", "filtered to", "lake folder", "duck db", "duckdb", "parquet files", "date range", "data time range", "total rows"]
for row in ws.iter_rows(min_row=2, max_row=9, values_only=True):
    k = str(row[1]).strip() if row[1] else ''
    v = str(row[2]).strip() if row[2] else ''
    if k and v and ("data time range" in k.lower() or "date range" in k.lower()):
        data_time_range = v
    if k and k not in ('None','nan') and not any(keyword in k.lower() for keyword in exclude_keywords):
        meta[k] = v

data_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
    if ws.row_dimensions[row_idx].hidden:
        continue
    raw_d = get_val(row, COL['DATE'])
    if isinstance(raw_d, datetime):
        ds = raw_d.strftime("%Y-%m-%d")
    elif raw_d:
        ds = str(raw_d).strip()
        if not ds or 'TOTAL' in ds.upper() or ds in ('None','nan'): continue
        
        # Normalize regional variations (e.g., 14/05/26 to standard 2026-05-14) on Python side
        parsed_dt = None
        for fmt_str in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y"):
            try:
                parsed_dt = datetime.strptime(ds, fmt_str)
                break
            except ValueError:
                continue
        if parsed_dt:
            ds = parsed_dt.strftime("%Y-%m-%d")
        else:
            if '/' not in ds and '-' not in ds: continue
    else:
        continue
        
    if not ds or 'TOTAL' in ds.upper(): continue

    ip  = int(sn(get_val(row, COL['IP_ROWS'])))
    sa  = int(sn(get_val(row, COL['SESS_AVAIL'])))
    sna = int(sn(get_val(row, COL['SESS_NA'])))
    sno = int(sn(get_val(row, COL['SESS_NONE'])))
    
    dist_ip = int(sn(get_val(row, COL['DIST_IP'])))
    dist_ipua = int(sn(get_val(row, COL['DIST_IPUA'])))
    
    dip_r2 = get_val(row, COL['DIST_IP_R2'])
    dipua_r2 = get_val(row, COL['DIST_IPUA_R2'])
    dist_ip_r2 = int(sn(dip_r2)) if dip_r2 not in (None, '', 'None', 'nan') else dist_ip
    dist_ipua_r2 = int(sn(dipua_r2)) if dipua_r2 not in (None, '', 'None', 'nan') else dist_ipua

    p_ip = get_val(row, COL['PCT_IP'])
    p_ipua = get_val(row, COL['PCT_IPUA'])
    p_sess = get_val(row, COL['PCT_SESS'])
    p_sessna = get_val(row, COL['PCT_SESSNA'])
    p_none = get_val(row, COL['PCT_NONE'])
    
    data_rows.append({
        "date":          ds,
        "bytes":         round(sn(get_val(row, COL['BYTES'])), 2),
        "rows":          int(sn(get_val(row, COL['ROWS']))),
        "ip_rows":       ip,
        "dist_ip":       dist_ip,
        "dist_ipua":     dist_ipua,
        "dist_dev":      int(sn(get_val(row, COL['DIST_DEV']))),
        "dist_sess":     int(sn(get_val(row, COL['DIST_SESS']))),
        "dist_ip_r2":    dist_ip_r2,
        "dist_ipua_r2":  dist_ipua_r2,
        "sess_avail":    sa,
        "sess_na":       sna,
        "sess_none":     sno,
        "pct_ip":        sn(p_ip) if (p_ip not in (None, '', 'None', 'nan')) else (round(dist_ip / ip, 6) if ip > 0 else 0),
        "pct_ipua":      sn(p_ipua) if (p_ipua not in (None, '', 'None', 'nan')) else (round(dist_ipua / ip, 6) if ip > 0 else 0),
        "pct_sess":      sn(p_sess) if (p_sess not in (None, '', 'None', 'nan')) else (round(sa / ip, 6) if ip > 0 else 0),
        "pct_sessna":    sn(p_sessna) if (p_sessna not in (None, '', 'None', 'nan')) else (round(sna / ip, 6) if ip > 0 else 0),
        "pct_none":      sn(p_none) if (p_none not in (None, '', 'None', 'nan')) else (round(sno / ip, 6) if ip > 0 else 0),
    })
wb.close()

# The latest date in the workbook is often an in-progress day. Match the Excel
# summary logic by excluding that newest partial date from the dashboard.
if data_rows:
    latest_date = max(r["date"] for r in data_rows)
    data_rows = [r for r in data_rows if r["date"] != latest_date]

device_snapshot_rows, device_daily_rows, device_first_seen_counts = read_device_data(DEVICE_SNAPSHOT_CSV, DEVICE_DAILY_CSV)
device_summary = {"total_devices": 0, "latest_date": "", "active_latest": 0, "new_latest": 0, "returning_latest": 0}
if device_snapshot_rows:
    device_summary["total_devices"] = len(device_snapshot_rows)
if device_daily_rows:
    latest_device_day = max(device_daily_rows, key=lambda d: d["date"])
    device_summary["latest_date"] = latest_device_day["date"]
    device_summary["active_latest"] = int(latest_device_day["active_devices"])
    device_summary["new_latest"] = sum(1 for d in device_snapshot_rows if d["first_seen"] == latest_device_day["date"])
    device_summary["returning_latest"] = max(0, device_summary["active_latest"] - device_summary["new_latest"])

generated_at = datetime.now()
data_blob = json.dumps({
    "meta": meta, "rows": data_rows,
    "device_daily": device_daily_rows,
    "device_first_seen_counts": device_first_seen_counts,
    "device_summary": device_summary,
    "generated": generated_at.strftime("%Y-%m-%d %H:%M:%S"),
    "report_date": generated_at.strftime("%Y-%m-%d"),
    "data_range": data_time_range,
}, ensure_ascii=False, separators=(',', ':'))

chartjs = get_chartjs()
chartjs_tag = f"<script>{chartjs}</script>" if chartjs else "<script>window.Chart=null;</script>"

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Overview Dashboard</title>
""" + chartjs_tag + """
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{--ink:#172033;--muted:#64748b;--panel:#fff;--line:#d8e1ee;--brand:#173b5c;--brand2:#1f6f8b;--accent:#c9a227;--bg:#eef3f8}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--ink);min-height:100vh;-webkit-text-size-adjust:100%}
.topbar{background:var(--brand);color:#fff;padding:18px 30px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:14px;border-bottom:4px solid var(--accent);box-shadow:0 8px 22px rgba(15,23,42,.16)}
.title-block{display:flex;flex-direction:column;gap:3px}
.topbar h1{font-size:20px;font-weight:700;letter-spacing:0}.topbar .eyebrow{font-size:11px;text-transform:uppercase;letter-spacing:.08em;color:#cbd5e1;font-weight:700}.header-meta{display:flex;flex-direction:column;align-items:flex-end;gap:3px;text-align:right;max-width:min(980px,72vw);line-height:1.35}.topbar .sub{font-size:12px;color:#dbe7f2}.range-pill{font-size:12px;font-weight:700;color:#fff}
#main{padding:22px 28px;max-width:1760px;margin:0 auto}
.meta-strip{background:var(--panel);border-radius:8px;border:1px solid var(--line);padding:12px 16px;margin-bottom:16px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px 18px;font-size:12px;color:#475569;box-shadow:0 1px 2px rgba(15,23,42,.05)}
.meta-items{display:flex;flex-wrap:wrap;gap:8px 26px;align-items:center}
.meta-strip b{color:var(--ink);font-weight:700}
.control-row{display:flex;flex-direction:column;gap:12px;margin-bottom:18px}
.cards{display:flex;flex-wrap:wrap;justify-content:center;gap:10px;width:100%}
.card{flex:0 1 calc((100% - 70px)/8);min-width:0;background:var(--panel);border-radius:8px;border:1px solid var(--line);padding:11px 14px;text-align:center;transition:all 0.3s;box-shadow:0 1px 2px rgba(15,23,42,.05);border-top:3px solid currentColor}
.clbl{font-size:10.5px;color:var(--muted);margin-bottom:7px;display:flex;align-items:flex-start;justify-content:center;text-align:center;gap:6px;font-weight:700;text-transform:uppercase;letter-spacing:.04em;line-height:1.15;min-height:24px}
.cdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.cval{font-size:25px;font-weight:800;line-height:1.05}
.charts{display:flex;flex-wrap:wrap;justify-content:center;align-items:stretch;gap:16px;margin-bottom:18px}
.cbox{flex:1 1 calc((100% - 16px)/2);min-width:320px;background:var(--panel);border-radius:8px;border:1px solid var(--line);padding:15px 17px;box-shadow:0 1px 2px rgba(15,23,42,.05)}
.cbox h3{font-size:13px;font-weight:700;margin-bottom:10px;display:flex;align-items:center;justify-content:space-between;gap:10px;color:#24324a}
.cbox canvas{width:100%!important;height:clamp(170px,22vh,240px)!important}
.chart-head{display:flex;align-items:center;justify-content:space-between;gap:8px;min-width:0}
.chart-actions{display:flex;align-items:center;gap:8px;flex-shrink:0}
.chart-days{color:#94a3b8;font-weight:500;font-size:11px}
.expand-btn{border:1px solid #cbd5e1;background:#fff;color:#1F3864;border-radius:5px;padding:3px 8px;font-size:10px;font-weight:800;cursor:pointer}
.expand-btn:hover{background:#edf2f7;border-color:#94a3b8}
.modal-backdrop{position:fixed;inset:0;background:rgba(15,23,42,.55);z-index:50;display:none;align-items:center;justify-content:center;padding:22px}
.modal-backdrop.open{display:flex}
.chart-modal{width:min(1180px,96vw);height:min(760px,92vh);background:#fff;border-radius:10px;border:1px solid #d8e1ee;box-shadow:0 24px 80px rgba(15,23,42,.28);display:flex;flex-direction:column;overflow:hidden}
.modal-top{padding:14px 18px;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;gap:14px;background:#f8fafc}
.modal-title{font-size:16px;font-weight:800;color:#1F3864}
.modal-close{border:1px solid #cbd5e1;background:#fff;color:#334155;border-radius:6px;padding:5px 10px;font-weight:800;cursor:pointer}
.modal-body{padding:14px 18px 18px;display:flex;flex-direction:column;gap:12px;min-height:0;flex:1}
.modal-controls{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.modal-controls input[type="date"]{border:1px solid #cbd5e1;padding:5px 7px;border-radius:5px;font-size:11px;font-family:inherit;color:#1F3864;font-weight:600}
.modal-range{font-size:11px;color:#64748b;font-weight:700;margin-left:auto}
.modal-stats{display:grid;grid-template-columns:repeat(4,minmax(120px,1fr));gap:8px}
.stat-chip{border:1px solid #e2e8f0;border-top:3px solid #1F3864;border-radius:7px;padding:8px 10px;background:#fff}
.stat-chip b{display:block;font-size:10px;color:#64748b;text-transform:uppercase;margin-bottom:3px}
.stat-chip span{font-size:16px;font-weight:800;color:#1e293b}
.modal-chart-wrap{min-height:0;flex:1;border:1px solid #e2e8f0;border-radius:8px;padding:12px;background:#fff}
.modal-chart-wrap canvas{width:100%!important;height:100%!important}
.table-wrap{background:var(--panel);border-radius:8px;border:1px solid var(--line);overflow:hidden;margin-bottom:18px;box-shadow:0 2px 8px rgba(15,23,42,.06)}
.thead{padding:13px 17px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid var(--line);flex-wrap:wrap;gap:15px;background:#f8fafc}
.thead h3{font-size:14px;font-weight:700;color:#24324a}
.drop-label{display:flex;flex-direction:column;align-items:center;gap:4px;line-height:1.1;font-weight:800}
.drop-label span{display:block}
.drop-select{width:100%;max-width:68px;border:1px solid #fdba74;background:#fff;color:#7c2d12;border-radius:6px;padding:3px 5px;font-size:9.5px;font-weight:800;outline:none;cursor:pointer;text-align:center;box-shadow:inset 0 0 0 1px rgba(255,255,255,.7)}
.drop-select:hover{border-color:#f97316;background:#fffbeb}
.drop-select:focus{border-color:#ea580c;box-shadow:0 0 0 2px rgba(249,115,22,.18)}
.filter-group{display:flex;align-items:center;gap:8px;background:#fff;padding:11px 12px;border-radius:8px;border:1px solid var(--line);border-top:3px solid var(--accent);box-shadow:0 1px 2px rgba(15,23,42,.05);width:100%;overflow-x:auto;white-space:nowrap}
.filter-group input[type="date"]{border:1px solid #cbd5e1;padding:5px 6px;border-radius:4px;font-size:11px;outline:none;font-family:inherit;color:#1F3864;font-weight:500;cursor:pointer;min-width:116px}
.filter-group span{font-size:11px;color:#64748b;font-weight:700;white-space:nowrap}
.filter-sep{width:1px;height:24px;background:#d8e1ee}
.btn-preset{background:#fff;border:1px solid #cbd5e1;padding:4px 10px;border-radius:4px;font-size:11px;font-weight:600;cursor:pointer;color:#555;transition:all 0.2s}
.btn-preset:hover{background:#edf2f7;color:#1F3864}
.btn-preset.active{background:#1F3864;color:#fff;border-color:#1F3864}
.tbl-c{overflow-x:auto;max-height:650px;overflow-y:auto;background:#fff}
table{width:100%;border-collapse:separate;border-spacing:0;font-size:10px;table-layout:fixed}
thead{position:sticky;top:0;z-index:5}
thead tr:first-child th{padding:5px 5px;font-size:9px;font-weight:700;color:#fff;text-align:center;white-space:normal;border-right:1px solid rgba(255,255,255,.25);border-bottom:1px solid #b8c4d4}
thead tr:last-child th{padding:6px 4px;text-align:center;font-size:9px;font-weight:600;color:#334155;border-right:1px solid #cbd5e1;border-bottom:1px solid #b8c4d4;white-space:normal;line-height:1.15;background:#f8fafc;cursor:pointer;user-select:none;overflow-wrap:anywhere;word-break:normal}
thead tr:last-child th:hover{background:#edf2f7}
thead tr:last-child th:first-child{text-align:center}
td{padding:6px 4px;text-align:center;vertical-align:middle;border-right:1px solid #dbe3ef;border-bottom:1px solid #dbe3ef;white-space:normal;line-height:1.2;overflow:visible;text-overflow:clip;overflow-wrap:anywhere}
th:first-child,td:first-child{border-left:1px solid #dbe3ef}
td:first-child{text-align:center;font-weight:500;position:sticky;left:0;background:#fff;z-index:1;width:72px}
tr:nth-child(even) td:first-child{background:#fafafa}
tr:hover td{background:#eef4ff!important}
tr.total-row td{background:#1F3864!important;color:#fff!important;font-weight:700}
tr.drop-row td{background:#fff7ed!important;color:#7c2d12!important;font-weight:700}
tr.drop-row td:first-child{text-align:center;padding:6px 5px}
.chg-pos{color:#166534!important}.chg-neg{color:#b91c1c!important}.chg-flat{color:#57534e!important}
.pct{color:#185FA5;font-weight:500}
.footer{text-align:center;font-size:11px;color:#94a3b8;padding:8px 0 20px}
th.b{background:#1F3864} th.g{background:#2e6b10} th.m{background:#7c3aed} th.a{background:#9a6000} th.p{background:#4a41a8}
@media (max-width:1450px){#main{padding:18px 20px}.topbar{padding:16px 22px}.cards{gap:8px}.card{flex-basis:calc((100% - 24px)/4);padding:10px 9px}.clbl{font-size:9.5px}.cval{font-size:21px}.charts{gap:12px}.cbox{padding:13px 14px}.tbl-c{max-height:610px}table{font-size:9.5px}thead tr:last-child th{font-size:8.5px;padding:5px 3px}td{padding:5px 3px}}
@media (max-width:1180px){.card{flex-basis:calc((100% - 24px)/4)}.cbox{flex-basis:calc((100% - 12px)/2);min-width:300px}.header-meta{max-width:100%;align-items:flex-start;text-align:left}}
@media (min-width:1500px){.cbox{flex-basis:calc((100% - 32px)/3)}.tbl-c{max-height:720px}}
@media (min-width:1900px){#main{max-width:1880px}.cbox{flex-basis:calc((100% - 64px)/5);min-width:0}.cbox canvas{height:clamp(190px,23vh,270px)!important}.tbl-c{max-height:760px}body{font-size:14px}}
@media (min-width:2200px){#main{max-width:2200px}.cbox canvas{height:clamp(220px,24vh,310px)!important}.tbl-c{max-height:880px}.chart-modal{width:min(1500px,94vw);height:min(900px,92vh)}}
@media (max-width:1050px){#main{padding:16px}.card{flex-basis:calc((100% - 16px)/3)}.cbox{flex-basis:100%;min-width:0}.topbar{padding:16px 18px}.topbar h1{font-size:18px}.header-meta{align-items:flex-start;max-width:100%;text-align:left}.meta-strip{align-items:flex-start}.filter-group{width:100%;overflow-x:visible;white-space:normal;flex-wrap:wrap}.control-row{align-items:stretch}}
@media (max-width:760px){.card{flex-basis:calc((100% - 8px)/2)}.modal-stats{grid-template-columns:repeat(2,minmax(0,1fr))}.modal-range{width:100%;margin-left:0}.chart-modal{height:94vh}.modal-backdrop{padding:10px}}
@media print{body{background:#fff}.topbar,.meta-strip,.card,.cbox,.table-wrap{box-shadow:none}.filter-group{display:none}.tbl-c{max-height:none;overflow:visible}}
</style>
</head>
<body>
<div class="topbar">
  <div class="title-block">
    <h1>Overview Dashboard Veto</h1>
  </div>
  <div class="header-meta">
    <span class="range-pill" id="rangeLbl"></span>
    <span class="sub" id="genLbl"></span>
  </div>
</div>
<div id="main">
  <div class="meta-strip" id="metaWrap">
    <div class="meta-items" id="metaStrip"></div>
  </div>
  <div class="control-row">
    <div class="cards" id="cards"></div>
    <div class="filter-group">
       <button class="btn-preset active" id="btnAll" onclick="setPreset('all')">ALL</button>
       <button class="btn-preset" id="btn30" onclick="setPreset(30)">30D</button>
       <button class="btn-preset" id="btn7" onclick="setPreset(7)">7D</button>
       <span class="filter-sep"></span>
       <input type="date" id="dateStart" onchange="handleCustomDate()" />
       <span>to</span>
       <input type="date" id="dateEnd" onchange="handleCustomDate()" />
    </div>
  </div>
  
  <div class="charts">
    <div class="cbox"><h3><span>Daily Row Volume</span><span class="chart-actions"><span class="chart-days" id="cH1"></span><button class="expand-btn" onclick="openChartModal('cRows')">Expand</button></span></h3><canvas id="cRows" height="170"></canvas></div>
    <div class="cbox"><h3><span>Data Density Profile</span><span class="chart-actions"><span class="chart-days" id="cH2"></span><button class="expand-btn" onclick="openChartModal('cBytes')">Expand</button></span></h3><canvas id="cBytes" height="170"></canvas></div>
    <div class="cbox"><h3><span>Session Not Found</span><span class="chart-actions"><span class="chart-days" id="cH3"></span><button class="expand-btn" onclick="openChartModal('cSessNone')">Expand</button></span></h3><canvas id="cSessNone" height="170"></canvas></div>
    <div class="cbox"><h3><span>Session Available & Missing</span><span class="chart-actions"><span class="chart-days" id="cH4"></span><button class="expand-btn" onclick="openChartModal('cSessFound')">Expand</button></span></h3><canvas id="cSessFound" height="170"></canvas></div>
    <div class="cbox"><h3><span>Unique Visitors & Devices</span><span class="chart-actions"><span class="chart-days" id="cH5"></span><button class="expand-btn" onclick="openChartModal('cIP')">Expand</button></span></h3><canvas id="cIP" height="170"></canvas></div>
  </div>

  <div class="table-wrap">
    <div class="thead">
      <h3>Day-by-Day Breakdown (Full Schema)</h3>
    </div>
    <div class="tbl-c"><table id="tbl"></table></div>
  </div>
</div>
<div class="modal-backdrop" id="chartModal">
  <div class="chart-modal">
    <div class="modal-top">
      <div class="modal-title" id="modalTitle">Chart</div>
      <button class="modal-close" onclick="closeChartModal()">Close</button>
    </div>
    <div class="modal-body">
      <div class="modal-controls">
        <button class="btn-preset" onclick="setModalPreset('dashboard')">Dashboard range</button>
        <button class="btn-preset" onclick="setModalPreset(7)">7D</button>
        <button class="btn-preset" onclick="setModalPreset(30)">30D</button>
        <button class="btn-preset" onclick="setModalPreset('all')">ALL</button>
        <span class="filter-sep"></span>
        <input type="date" id="modalStart" onchange="setModalCustom()" />
        <span style="font-size:11px;color:#64748b;font-weight:700">to</span>
        <input type="date" id="modalEnd" onchange="setModalCustom()" />
        <span class="modal-range" id="modalRange"></span>
      </div>
      <div class="modal-stats" id="modalStats"></div>
      <div class="modal-chart-wrap"><canvas id="modalChart"></canvas></div>
    </div>
  </div>
</div>
<script>
const {meta,rows,device_daily,device_first_seen_counts,device_summary,generated,report_date,data_range} = """ + data_blob + """;

function fmtDateTime12(value) {
  const m = String(value || '').trim().match(/^(\\d{4})-(\\d{2})-(\\d{2})(?:[ T](\\d{2}):(\\d{2})(?::(\\d{2}))?)?/);
  if(!m) return String(value || '');
  let hh = Number(m[4] || 0);
  const ap = hh >= 12 ? 'PM' : 'AM';
  hh = hh % 12 || 12;
  return `${m[3]}-${m[2]}-${String(m[1]).slice(2)} ${String(hh).padStart(2,'0')}:${m[5] || '00'}:${m[6] || '00'} ${ap}`;
}

function fmtDateRange12(value) {
  const parts = String(value || '').split(/\\s*(?:\u2192|->| to )\\s*/);
  if(parts.length >= 2) return `${fmtDateTime12(parts[0])} \u2192 ${fmtDateTime12(parts.slice(1).join(' to '))}`;
  return fmtDateTime12(value);
}

// Parse strictly using localized component metrics to remove browser engine discrepancies
rows.forEach(r => { 
    const [y, m, d] = r.date.split('-').map(Number);
    r.ts = new Date(y, m - 1, d).getTime(); 
});
rows.sort((a,b) => a.ts - b.ts);
const firstDataDate = rows.length ? rows[0].date : '';
const completeRows = rows.filter(r => r.date !== firstDataDate && r.date !== report_date);
const usedDateRange = completeRows.length ? `${fmtDateTime12(completeRows[0].date + ' 00:00:00')} to ${fmtDateTime12(completeRows[completeRows.length - 1].date + ' 23:59:59')} IST` : '';
const deviceDailyByDate = new Map(device_daily.map(d => [d.date, d]));

const headerRangeText = (data_range ? 'True data range: '+fmtDateRange12(data_range) : '') + (usedDateRange ? (data_range ? ' | ' : '')+'Used range: '+usedDateRange : '');
document.getElementById('rangeLbl').style.display = headerRangeText ? '' : 'none';
document.getElementById('rangeLbl').textContent = headerRangeText;
document.getElementById('genLbl').textContent='Last updated '+fmtDateTime12(generated);
const visibleMeta = Object.entries(meta).filter(([k]) => {
  const key = k.toLowerCase().replace(/\\s+/g, '');
  return !key.includes('duckdb') && !key.includes('totalrows') && !key.includes('datetimerange');
});
if(visibleMeta.length > 0) document.getElementById('metaStrip').innerHTML=visibleMeta.map(([k,v])=>`<span><b>${k}:</b> ${v}</span>`).join('');
else document.getElementById('metaWrap').style.display = 'none';

const fmt=(n,d=0)=>n==null||isNaN(n)?'\u2014':Number(n).toLocaleString('en-IN',{minimumFractionDigits:d,maximumFractionDigits:d});
const fmtData=gib=>{ if(gib==null||isNaN(gib))return'\u2014'; const tb=Number(gib)*1.073741824/1000; return tb>=1 ? fmt(Math.round(tb))+' TB' : fmt(Math.round(Number(gib)))+' GiB'; };
const fmtP=n=>{ if(n==null||isNaN(n))return'\u2014'; const p=Math.abs(n)<=1?n*100:n; return p.toFixed(2)+'%'; };

let viewRows = [...completeRows];
let sortKey = null;
let sortAsc = true;
let dropLookbackDays = 5;
const chartInstances = {};

const CD={tension:.35,fill:true,pointRadius:1.5,pointHoverRadius:4,borderWidth:1.5,animation:{duration:400}};
const mkOpt=(yfmt, expanded=false)=>({responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{font:{size:expanded?11:9.5},boxWidth:8,padding:6}}},
  scales:{x:{ticks:{font:{size:8.5},maxTicksLimit:12},grid:{color:'#f5f5f5'}},
          y:{ticks:{font:{size:8.5},callback:v=>yfmt==='gib'?fmtData(v):v>=1e6?(v/1e6).toFixed(1)+'M':v>=1e3?(v/1e3).toFixed(0)+'K':v},grid:{color:'#f5f5f5'}}}});

const chartDefs = {
  cRows:{title:'Daily Row Volume', keys:['rows'], yfmt:null, datasets:[
    {label:'Total rows', borderColor:'#185FA5', backgroundColor:'rgba(24,95,165,.04)', ...CD},
    {label:'Avg', borderColor:'rgba(24,95,165,.6)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]},
  cBytes:{title:'Data Density Profile', keys:['bytes'], yfmt:'gib', datasets:[
    {label:'Data Volume', borderColor:'#1D9E75', backgroundColor:'rgba(29,158,117,.04)', ...CD},
    {label:'Avg', borderColor:'rgba(29,158,117,.6)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]},
  cSessNone:{title:'Session Not Found', keys:['sess_none'], yfmt:null, datasets:[
    {label:'Sess Not Found', borderColor:'#b71c1c', backgroundColor:'transparent', ...CD},
    {label:'Avg', borderColor:'rgba(183,28,28,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]},
  cSessFound:{title:'Session Available & Missing', keys:['sess_avail','sess_na'], yfmt:null, datasets:[
    {label:'Sess Avail', borderColor:'#2e6b10', backgroundColor:'transparent', ...CD},
    {label:'Avg', borderColor:'rgba(46,107,16,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false},
    {label:'Sess Missing', borderColor:'#9a6000', backgroundColor:'transparent', ...CD},
    {label:'Avg', borderColor:'rgba(154,96,0,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]},
  cIP:{title:'Unique Visitors & Devices', keys:['dist_ip','dist_dev'], yfmt:null, datasets:[
    {label:'Distinct cliIP', borderColor:'#4a41a8', backgroundColor:'transparent', ...CD},
    {label:'Avg', borderColor:'rgba(74,65,168,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false},
    {label:'Distinct device', borderColor:'#0F6E56', backgroundColor:'transparent', ...CD},
    {label:'Avg', borderColor:'rgba(15,110,86,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]}
};

function cloneDatasets(datasets) {
  return datasets.map(d => ({...d, data:[]}));
}

function initCharts() {
  if(typeof Chart==='undefined') return;
  Object.entries(chartDefs).forEach(([id, def]) => {
    chartInstances[id] = new Chart(document.getElementById(id), {type:'line', data:{labels:[], datasets:cloneDatasets(def.datasets)}, options:mkOpt(def.yfmt)});
  });
}

function updateKPIs() {
    const ndays = viewRows.length || 1;
    const sortedVisible = [...viewRows].sort((a,b) => a.ts - b.ts);
    const rangeDates = new Set(sortedVisible.map(r => r.date));
    const activeDeviceSet = new Set();
    let newInRange = 0;
    rangeDates.forEach(date => {
      const day = deviceDailyByDate.get(date);
      if(day) day.devices.forEach(id => activeDeviceSet.add(id));
      newInRange += Number(device_first_seen_counts[date] || 0);
    });
    const activeInRange = activeDeviceSet.size;
    const returningInRange = Math.max(0, activeInRange - newInRange);
    const tBytes = viewRows.reduce((s, r) => s + r.bytes, 0);
    const tIP = viewRows.reduce((s, r) => s + r.dist_ip, 0);
    const tDev = viewRows.reduce((s, r) => s + r.dist_dev, 0);
    const tSess = viewRows.reduce((s, r) => s + r.dist_sess, 0);
    const tRows = viewRows.reduce((s, r) => s + r.rows, 0);

    const kpiHTML = [
        {l:'Avg Volume / Day',v:fmtData(tBytes/ndays),clr:'#0F6E56'},
        {l:'Avg Distinct IPs / Day',v:fmt(Math.round(tIP/ndays)),clr:'#4a41a8'},
        {l:'Avg Device IDs / Day',v:fmt(Math.round(tDev/ndays)),clr:'#1f6f8b'},
        {l:'Avg Sessions / Day',v:fmt(Math.round(tSess/ndays)),clr:'#9a6000'},
        {l:'Avg Rows / Day',v:fmt(Math.round(tRows/ndays)),clr:'#b71c1c'},
        {l:'Total Device Count',v:fmt(activeInRange),clr:'#173b5c'},
        {l:'Returning Devices',v:fmt(returningInRange),clr:'#2e6b10'},
        {l:'New Devices',v:fmt(newInRange),clr:'#c9a227'}
    ].map(c => `<div class="card"><div class="clbl"><span class="cdot" style="background:${c.clr}"></span>${c.l}</div><div class="cval" style="color:${c.clr}">${c.v}</div></div>`).join('');
    document.getElementById('cards').innerHTML = kpiHTML;
}

function updateCharts() {
  if(!chartInstances['cRows']) return;
  Object.entries(chartDefs).forEach(([id, def]) => syncChart(chartInstances[id], viewRows, def.keys));
  document.querySelectorAll('.chart-days').forEach(e => e.textContent = viewRows.length + ' days');
}

function syncChart(ch, rowsForChart, keys) {
  const lbl = rowsForChart.map(r=>r.date);
  const len = rowsForChart.length || 1;
  ch.data.labels = lbl;
  keys.forEach((k, i) => {
    const total = rowsForChart.reduce((s,r) => s + Number(r[k] || 0), 0);
    ch.data.datasets[i * 2].data = rowsForChart.map(r => r[k]);
    ch.data.datasets[(i * 2) + 1].data = Array(len).fill(total / len);
  });
  ch.update();
}

let modalChart = null;
let modalChartId = null;
let modalRows = [];

function dtFmt(ts) {
  const d = new Date(ts);
  return `${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,'0')}-${String(d.getDate()).padStart(2,'0')}`;
}

function formatMetric(k, v) {
  if(k === 'bytes') return fmtData(v);
  if(k.startsWith('pct_')) return fmtP(v);
  return fmt(Math.round(v));
}

function openChartModal(id) {
  if(typeof Chart==='undefined') return;
  modalChartId = id;
  const def = chartDefs[id];
  document.getElementById('modalTitle').textContent = def.title;
  document.getElementById('chartModal').classList.add('open');
  if(modalChart) modalChart.destroy();
  modalChart = new Chart(document.getElementById('modalChart'), {
    type:'line',
    data:{labels:[], datasets:cloneDatasets(def.datasets)},
    options:mkOpt(def.yfmt, true)
  });
  setModalPreset('dashboard');
}

function closeChartModal() {
  document.getElementById('chartModal').classList.remove('open');
  if(modalChart) { modalChart.destroy(); modalChart = null; }
  modalChartId = null;
}

document.getElementById('chartModal').addEventListener('click', e => {
  if(e.target.id === 'chartModal') closeChartModal();
});
document.addEventListener('keydown', e => {
  if(e.key === 'Escape' && document.getElementById('chartModal').classList.contains('open')) closeChartModal();
});

function setModalPreset(mode) {
  if(!modalChartId) return;
  if(mode === 'dashboard') modalRows = [...viewRows].sort((a,b) => a.ts - b.ts);
  else if(mode === 'all') modalRows = [...completeRows];
  else modalRows = completeRows.slice(-Number(mode));
  document.getElementById('modalStart').value = modalRows.length ? modalRows[0].date : '';
  document.getElementById('modalEnd').value = modalRows.length ? modalRows[modalRows.length - 1].date : '';
  renderModalChart();
}

function setModalCustom() {
  if(!modalChartId) return;
  const startStr = document.getElementById('modalStart').value;
  const endStr = document.getElementById('modalEnd').value;
  let sTs = -Infinity, eTs = Infinity;
  if(startStr) {
    const [y,m,d] = startStr.split('-').map(Number);
    sTs = new Date(y, m - 1, d).getTime();
  }
  if(endStr) {
    const [y,m,d] = endStr.split('-').map(Number);
    eTs = new Date(y, m - 1, d + 1).getTime() - 1;
  }
  modalRows = completeRows.filter(r => r.ts >= sTs && r.ts <= eTs);
  renderModalChart();
}

function renderModalChart() {
  const def = chartDefs[modalChartId];
  syncChart(modalChart, modalRows, def.keys);
  const rangeText = modalRows.length ? `${modalRows[0].date} to ${modalRows[modalRows.length - 1].date} | ${modalRows.length} days` : 'No dates selected';
  document.getElementById('modalRange').textContent = rangeText;
  const chips = [];
  def.keys.forEach(k => {
    const vals = modalRows.map(r => Number(r[k] || 0)).filter(v => !isNaN(v));
    if(!vals.length) return;
    const latest = vals[vals.length - 1];
    const avg = vals.reduce((s,v) => s + v, 0) / vals.length;
    chips.push({l:chartDefs[modalChartId].datasets[def.keys.indexOf(k) * 2].label + ' Latest', v:formatMetric(k, latest)});
    chips.push({l:chartDefs[modalChartId].datasets[def.keys.indexOf(k) * 2].label + ' Avg', v:formatMetric(k, avg)});
  });
  document.getElementById('modalStats').innerHTML = chips.slice(0,4).map(c => `<div class="stat-chip"><b>${c.l}</b><span>${c.v}</span></div>`).join('');
}

const subH=['Date', 'Data', 'Total Rows', 'cliIP rows', 'Dist IP', 'Dist IP+UA', 'Devices', 'Sessions', 'Dist IP', 'Dist IP+UA', 'Sess Avail', 'Sess Missing', 'Sess Not Found', 'Dist IP', 'Dist IP+UA', 'Sess Avail', 'Sess Missing', 'Sess Not Found'];
const fields=['date','bytes','rows','ip_rows','dist_ip','dist_ipua','dist_dev','dist_sess','dist_ip_r2','dist_ipua_r2','sess_avail','sess_na','sess_none','pct_ip','pct_ipua','pct_sess','pct_sessna','pct_none'];
const columnGroups = {
  b:{fields:['bytes','rows','ip_rows'], color:'31,56,100'},
  g:{fields:['dist_ip','dist_ipua','dist_dev','dist_sess'], color:'46,107,16'},
  m:{fields:['dist_ip_r2','dist_ipua_r2'], color:'124,58,237'},
  a:{fields:['sess_avail','sess_na','sess_none'], color:'154,96,0'},
  p:{fields:['pct_ip','pct_ipua','pct_sess','pct_sessna','pct_none'], color:'74,65,168'}
};
const fieldGroup = {};
Object.entries(columnGroups).forEach(([groupKey, group]) => group.fields.forEach(f => fieldGroup[f] = groupKey));

function getScaleStats() {
  const stats = {};
  fields.filter(f => f !== 'date').forEach(k => {
    const vals = viewRows.map(r => Number(r[k])).filter(v => !isNaN(v));
    stats[k] = vals.length ? {min: Math.min(...vals), max: Math.max(...vals)} : {min: 0, max: 0};
  });
  return stats;
}

function blend(a, b, t) {
  const ah = a.match(/\\w\\w/g).map(x => parseInt(x, 16));
  const bh = b.match(/\\w\\w/g).map(x => parseInt(x, 16));
  const rgb = ah.map((v, i) => Math.round(v + (bh[i] - v) * t));
  return '#' + rgb.map(v => v.toString(16).padStart(2, '0')).join('');
}

function columnGradient(k, v, stats) {
  const n = Number(v);
  const s = stats[k];
  const group = columnGroups[fieldGroup[k]];
  if(!group || !s || isNaN(n)) return '';
  if(s.max === s.min) return ` style="background:rgba(${group.color},.08)"`;
  const t = Math.max(0, Math.min(1, (n - s.min) / (s.max - s.min)));
  const alpha = (.08 + (t * .30)).toFixed(3);
  return ` style="background:rgba(${group.color},${alpha})"`;
}

function dataCell(d, k, body, cls='', stats=null) {
  return `<td${cls ? ` class="${cls}"` : ''}${stats ? columnGradient(k, d[k], stats) : ''}>${body}</td>`;
}

function fmtChg(n) {
  if(n == null || isNaN(n)) return '\u2014';
  const cls = n > 0 ? 'chg-pos' : n < 0 ? 'chg-neg' : 'chg-flat';
  const sign = n > 0 ? '+' : '';
  return `<span class="${cls}">${sign}${(n * 100).toFixed(2)}%</span>`;
}

function getDropGain() {
  const byDate = [...viewRows].sort((a, b) => a.ts - b.ts);
  if(!byDate.length) return {};
  const target = byDate[byDate.length - 1];
  const historyRows = completeRows.filter(r => r.ts < target.ts).sort((a, b) => a.ts - b.ts);
  if(historyRows.length < dropLookbackDays) return {};
  const baseRows = historyRows.slice(-dropLookbackDays);
  const calc = k => {
    const avg = baseRows.reduce((s, r) => s + Number(r[k] || 0), 0) / baseRows.length;
    return avg ? (Number(target[k] || 0) / avg) - 1 : null;
  };
  return {
    bytes: calc('bytes'), rows: calc('rows'), ip_rows: calc('ip_rows'),
    dist_ip: calc('dist_ip'), dist_ipua: calc('dist_ipua'), dist_dev: calc('dist_dev'), dist_sess: calc('dist_sess'),
    dist_ip_r2: calc('dist_ip_r2'), dist_ipua_r2: calc('dist_ipua_r2'),
    sess_avail: calc('sess_avail'), sess_na: calc('sess_na'), sess_none: calc('sess_none'),
    pct_sess: calc('pct_sess'), pct_sessna: calc('pct_sessna'), pct_none: calc('pct_none')
  };
}

function renderTable() {
  const scaleStats = getScaleStats();
  const dropOptions = [3, 5, 7, 14, 30].map(n => `<option value="${n}"${n === dropLookbackDays ? ' selected' : ''}>${n} days</option>`).join('');
  let h = `<thead>
    <tr><th class="b" style="background:#f8fafc;"></th><th colspan="3" class="b">General Volumes</th><th colspan="4" class="g">Absolute Counts</th><th colspan="2" class="m">Unique Subsets</th><th colspan="3" class="a">Session Identifiers</th><th colspan="5" class="p">% Data of TOTAL Segment</th></tr>
    <tr>${subH.map((title, idx) => `<th onclick="handleSort('${fields[idx]}')">${title}${sortKey === fields[idx] ? (sortAsc ? ' \u25B2' : ' \u25BC') : ''}</th>`).join('')}</tr>
  </thead><tbody>`;

  h += viewRows.map(d => `<tr>
    <td>${d.date}</td>${dataCell(d,'bytes',fmtData(d.bytes),'',scaleStats)}${dataCell(d,'rows',fmt(d.rows),'',scaleStats)}${dataCell(d,'ip_rows',fmt(d.ip_rows),'',scaleStats)}
    ${dataCell(d,'dist_ip',fmt(d.dist_ip),'',scaleStats)}${dataCell(d,'dist_ipua',fmt(d.dist_ipua),'',scaleStats)}${dataCell(d,'dist_dev',fmt(d.dist_dev),'',scaleStats)}${dataCell(d,'dist_sess',fmt(d.dist_sess),'',scaleStats)}
    ${dataCell(d,'dist_ip_r2',fmt(d.dist_ip_r2),'',scaleStats)}${dataCell(d,'dist_ipua_r2',fmt(d.dist_ipua_r2),'',scaleStats)}
    ${dataCell(d,'sess_avail',fmt(d.sess_avail),'',scaleStats)}${dataCell(d,'sess_na',fmt(d.sess_na),'',scaleStats)}${dataCell(d,'sess_none',fmt(d.sess_none),'',scaleStats)}
    ${dataCell(d,'pct_ip',fmtP(d.pct_ip),'pct',scaleStats)}${dataCell(d,'pct_ipua',fmtP(d.pct_ipua),'pct',scaleStats)}${dataCell(d,'pct_sess',fmtP(d.pct_sess),'pct',scaleStats)}${dataCell(d,'pct_sessna',fmtP(d.pct_sessna),'pct',scaleStats)}${dataCell(d,'pct_none',fmtP(d.pct_none),'pct',scaleStats)}
  </tr>`).join('');

  const tKeys = ["bytes","rows","ip_rows","dist_ip","dist_ipua","dist_dev","dist_sess","dist_ip_r2","dist_ipua_r2","sess_avail","sess_na","sess_none"];
  const tVals = {}; tKeys.forEach(k => tVals[k] = viewRows.reduce((sum, r) => sum + r[k], 0));
  const tPct = {
    pct_ip: tVals.ip_rows > 0 ? tVals.dist_ip / tVals.ip_rows : 0,
    pct_ipua: tVals.ip_rows > 0 ? tVals.dist_ipua / tVals.ip_rows : 0,
    pct_sess: tVals.ip_rows > 0 ? tVals.sess_avail / tVals.ip_rows : 0,
    pct_sessna: tVals.ip_rows > 0 ? tVals.sess_na / tVals.ip_rows : 0,
    pct_none: tVals.ip_rows > 0 ? tVals.sess_none / tVals.ip_rows : 0
  };
  const dropGain = getDropGain();

  h += `<tr class="total-row">
    <td>TOTAL</td><td>${fmtData(tVals.bytes)}</td><td>${fmt(tVals.rows)}</td><td>${fmt(tVals.ip_rows)}</td>
    <td>${fmt(tVals.dist_ip)}</td><td>${fmt(tVals.dist_ipua)}</td><td>${fmt(tVals.dist_dev)}</td><td>${fmt(tVals.dist_sess)}</td>
    <td>${fmt(tVals.dist_ip_r2)}</td><td>${fmt(tVals.dist_ipua_r2)}</td>
    <td>${fmt(tVals.sess_avail)}</td><td>${fmt(tVals.sess_na)}</td><td>${fmt(tVals.sess_none)}</td>
    <td class="pct">${fmtP(tPct.pct_ip)}</td><td class="pct">${fmtP(tPct.pct_ipua)}</td><td class="pct">${fmtP(tPct.pct_sess)}</td><td class="pct">${fmtP(tPct.pct_sessna)}</td><td class="pct">${fmtP(tPct.pct_none)}</td>
  </tr>
  <tr class="drop-row">
    <td><div class="drop-label"><span>% Drop/Gain</span><span>vs previous</span><select class="drop-select" onchange="setDropLookback(this.value)">${dropOptions}</select></div></td><td>${fmtChg(dropGain.bytes)}</td><td>${fmtChg(dropGain.rows)}</td><td>${fmtChg(dropGain.ip_rows)}</td>
    <td>${fmtChg(dropGain.dist_ip)}</td><td>${fmtChg(dropGain.dist_ipua)}</td><td>${fmtChg(dropGain.dist_dev)}</td><td>${fmtChg(dropGain.dist_sess)}</td>
    <td>${fmtChg(dropGain.dist_ip_r2)}</td><td>${fmtChg(dropGain.dist_ipua_r2)}</td>
    <td>${fmtChg(dropGain.sess_avail)}</td><td>${fmtChg(dropGain.sess_na)}</td><td>${fmtChg(dropGain.sess_none)}</td>
    <td>\u2014</td><td>\u2014</td><td>${fmtChg(dropGain.pct_sess)}</td><td>${fmtChg(dropGain.pct_sessna)}</td><td>${fmtChg(dropGain.pct_none)}</td>
  </tr></tbody>`;
  document.getElementById('tbl').innerHTML = h;
}

function setDropLookback(days) {
  dropLookbackDays = Math.max(1, Number(days) || 5);
  renderTable();
}

function handleSort(key) {
  if (sortKey === key) sortAsc = !sortAsc;
  else { sortKey = key; sortAsc = true; }
  executeSortLogic(); renderTable();
}

function executeSortLogic() {
  viewRows.sort((a, b) => {
    let va = a[sortKey], vb = b[sortKey];
    if(sortKey === 'date') return sortAsc ? a.ts - b.ts : b.ts - a.ts;
    if (typeof va === 'string') return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
    return sortAsc ? va - vb : vb - va;
  });
}

function syncUIState(presetMode) {
    document.querySelectorAll('.btn-preset').forEach(b => b.classList.remove('active'));
    if(presetMode === 'all') document.getElementById('btnAll').classList.add('active');
    else if(presetMode === 30) document.getElementById('btn30').classList.add('active');
    else if(presetMode === 7) document.getElementById('btn7').classList.add('active');
}

function setPreset(days) {
    syncUIState(days);
    if(days === 'all' || completeRows.length === 0) {
        viewRows = [...completeRows];
        document.getElementById('dateStart').value = '';
        document.getElementById('dateEnd').value = '';
    } else {
        viewRows = completeRows.slice(-days);
        
        const dtFmt = (ts) => { 
            const d = new Date(ts); 
            return `${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,'0')}-${String(d.getDate()).padStart(2,'0')}`; 
        };
        if(viewRows.length) {
            document.getElementById('dateStart').value = dtFmt(viewRows[0].ts);
            document.getElementById('dateEnd').value = dtFmt(viewRows[viewRows.length-1].ts);
        }
    }
    if(sortKey) executeSortLogic();
    renderTable(); updateCharts(); updateKPIs();
}

function handleCustomDate() {
    syncUIState(null);
    const startStr = document.getElementById('dateStart').value;
    const endStr = document.getElementById('dateEnd').value;
    
    let sTs = -Infinity;
    if (startStr) {
        const [y, m, d] = startStr.split('-').map(Number);
        sTs = new Date(y, m - 1, d).getTime();
    }
    let eTs = Infinity;
    if (endStr) {
        const [y, m, d] = endStr.split('-').map(Number);
        eTs = new Date(y, m - 1, d + 1).getTime() - 1; 
    }
    
    viewRows = completeRows.filter(r => r.ts >= sTs && r.ts <= eTs);
    if(sortKey) executeSortLogic();
    renderTable(); updateCharts(); updateKPIs();
}

initCharts();
setPreset(7);
</script>
</body>
</html>"""

HTML_OUT.parent.mkdir(parents=True, exist_ok=True)
HTML_OUT.write_text(HTML, encoding='utf-8')
print(f"\n[Success] Dashboard output to: {HTML_OUT}")
