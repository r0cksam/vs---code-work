"""
generate_dashboard.py
─────────────────────
Reads overview_report.xlsx (same folder as this script)
Writes overview_dashboard.html (same folder) — 100% self-contained.

Your daily workflow:
    1. Run your overview_generator.py   →  overview_report.xlsx updated
    2. Run:  python generate_dashboard.py
    3. Done — stakeholders double-click overview_dashboard.html

Both files live in the same folder on the network drive.
No internet needed at view-time (Chart.js is cached locally on first run).
"""

import sys, json, urllib.request
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl

# ── Paths: same folder as this script ───────────────────────────
HERE      = Path(__file__).parent
XLSX_PATH = HERE / "Y:\Veto Logs Backup\Dashboards\OverView\overview_report.xlsx"
HTML_OUT  = HERE / "Y:\Veto Logs Backup\Dashboards\OverView\overview_dashboard.html"

# CLI override: python generate_dashboard.py <xlsx> <out.html>
if len(sys.argv) == 3:
    XLSX_PATH = Path(sys.argv[1])
    HTML_OUT  = Path(sys.argv[2])

# ── Column map (1-indexed, matches overview_generator.py) ───────
COL = dict(
    DATE=2, BYTES=3, ROWS=4, IP_ROWS=5,
    DIST_IP=6, DIST_IPUA=7, DIST_DEV=8, DIST_SESS=9,
    UNIQ_IP=10, UNIQ_IPUA=11,
    SESS_AVAIL=12, SESS_NA=13, SESS_NONE=14,
    PCT_IP=15, PCT_IPUA=16, PCT_SESS=17, PCT_SESSNA=18, PCT_NONE=19,
)
DATA_START_ROW = 13

def safe_num(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def safe_str(v):
    return str(v).strip() if v is not None else ""

# ── Chart.js: downloaded once, cached next to this script ───────
CHARTJS_CACHE = HERE / ".chartjs_cache.js"
CHARTJS_URL   = "https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"

def get_chartjs():
    if CHARTJS_CACHE.exists():
        print("  Chart.js: using cached copy")
        return CHARTJS_CACHE.read_text(encoding="utf-8")
    print("  Chart.js: downloading once from CDN…", end=" ", flush=True)
    try:
        with urllib.request.urlopen(CHARTJS_URL, timeout=15) as r:
            js = r.read().decode("utf-8")
        CHARTJS_CACHE.write_text(js, encoding="utf-8")
        print(f"done ({len(js)//1024} KB cached)")
        return js
    except Exception as e:
        print(f"\n  WARNING: Could not download Chart.js: {e}")
        print("  Charts will be disabled. Run once with internet to cache it.")
        return None

# ── Read Excel ───────────────────────────────────────────────────
print(f"\nReading: {XLSX_PATH}")
if not XLSX_PATH.exists():
    print(f"  ERROR: File not found: {XLSX_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb.active

meta = {}
for r in range(2, 10):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    k, v = safe_str(row[1]), safe_str(row[2])
    if k: meta[k] = v

data_rows, total_row = [], None
for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
    ds = safe_str(row[COL['DATE'] - 1])
    if not ds: continue
    if 'TOTAL' in ds.upper(): total_row = row; continue
    if '/' not in ds and len(ds) < 5: continue
    n = lambda c: safe_num(row[COL[c] - 1])
    data_rows.append({
        "date": ds,        "bytes": round(n('BYTES'), 2),
        "rows": int(n('ROWS')),     "ip_rows": int(n('IP_ROWS')),
        "dist_ip":    int(n('DIST_IP')),   "dist_ipua": int(n('DIST_IPUA')),
        "dist_dev":   int(n('DIST_DEV')),  "dist_sess": int(n('DIST_SESS')),
        "sess_avail": int(n('SESS_AVAIL')),"sess_na":   int(n('SESS_NA')),
        "sess_none":  int(n('SESS_NONE')),
        "pct_ip":     round(n('PCT_IP'),6),    "pct_ipua":   round(n('PCT_IPUA'),6),
        "pct_sess":   round(n('PCT_SESS'),6),  "pct_sessna": round(n('PCT_SESSNA'),6),
        "pct_none":   round(n('PCT_NONE'),6),
    })
wb.close()

tot = {}
if total_row:
    tn = lambda c: safe_num(total_row[COL[c] - 1])
    tot = {k.lower(): (round(tn('BYTES'), 2) if k == 'BYTES' else int(tn(k)))
           for k in ['ROWS','IP_ROWS','DIST_IP','DIST_IPUA','DIST_DEV',
                     'DIST_SESS','SESS_AVAIL','SESS_NA','SESS_NONE','BYTES']}

generated_at = datetime.now().strftime("%d %b %Y, %H:%M")
print(f"  {len(data_rows)} data rows found")

chartjs   = get_chartjs()
data_blob = json.dumps({
    "meta": meta, "rows": data_rows, "total": tot,
    "generated": generated_at, "source": XLSX_PATH.name,
    "has_charts": chartjs is not None,
}, ensure_ascii=False)

chartjs_tag = f"<script>{chartjs}</script>" if chartjs else "<script>window.Chart=null;</script>"

# ── Build HTML ───────────────────────────────────────────────────
html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Overview Report Dashboard</title>
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
{chartjs_tag}
<style>
:root {{
  --bg:       #07090f;
  --surface:  #0e1117;
  --surface2: #151a24;
  --border:   #1e2736;
  --border2:  #263047;
  --text:     #d8e0ee;
  --muted:    #5a6a84;
  --faint:    #2a3548;
  --accent:   #4f8ef7;
  --green:    #43e07a;
  --orange:   #f5a623;
  --red:      #f75b5b;
  --purple:   #b57bee;
  --teal:     #30d9a0;
  --r:        14px;
}}
*, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
html {{ scroll-behavior: smooth; }}
body {{
  font-family: 'Sora', sans-serif;
  background: var(--bg);
  color: var(--text);
  min-height: 100vh;
  font-size: 13.5px;
  line-height: 1.6;
}}
body::before {{
  content: '';
  position: fixed; inset: 0; z-index: 0;
  background-image:
    linear-gradient(rgba(79,142,247,.03) 1px, transparent 1px),
    linear-gradient(90deg, rgba(79,142,247,.03) 1px, transparent 1px);
  background-size: 40px 40px;
  pointer-events: none;
}}

/* ── Header ── */
.header {{
  position: sticky; top: 0; z-index: 200;
  background: rgba(7,9,15,.9);
  backdrop-filter: blur(16px);
  border-bottom: 1px solid var(--border);
  padding: 14px 32px;
  display: flex; align-items: center; justify-content: space-between;
  flex-wrap: wrap; gap: 10px;
}}
.header-brand {{ display: flex; align-items: center; gap: 12px; }}
.brand-icon {{
  width: 34px; height: 34px; border-radius: 9px; flex-shrink: 0;
  background: linear-gradient(135deg, var(--accent), var(--purple));
  display: flex; align-items: center; justify-content: center; font-size: 15px;
}}
.brand-name {{ font-size: 15px; font-weight: 600; }}
.brand-sub  {{ font-size: 11px; color: var(--muted); margin-top: 1px; }}
.pill {{
  background: rgba(79,142,247,.12); color: var(--accent);
  border: 1px solid rgba(79,142,247,.25);
  border-radius: 20px; padding: 4px 12px; font-size: 11px; font-weight: 500;
}}

/* ── Layout ── */
#main {{ position: relative; z-index: 1; max-width: 1480px; margin: 0 auto; padding: 28px 32px 56px; }}

/* ── Meta strip ── */
.meta-strip {{
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--r); padding: 11px 18px; margin-bottom: 22px;
  display: flex; flex-wrap: wrap; gap: 6px 24px;
  font-size: 11.5px; color: var(--muted);
}}
.meta-strip b {{ color: var(--text); font-weight: 500; }}

/* ── Section label ── */
.sec-label {{
  font-size: 10px; font-weight: 700; letter-spacing: .1em;
  text-transform: uppercase; color: var(--muted);
  margin: 0 0 14px; padding-bottom: 10px;
  border-bottom: 1px solid var(--border);
}}

/* ── KPI Cards ── */
.cards {{
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(195px, 1fr));
  gap: 14px; margin-bottom: 28px;
}}
.card {{
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--r); padding: 20px 20px 16px;
  position: relative; overflow: hidden;
  transition: border-color .2s, transform .15s, box-shadow .2s;
}}
.card:hover {{ transform: translateY(-3px); box-shadow: 0 8px 32px rgba(0,0,0,.4); border-color: var(--border2); }}
.card::after {{
  content: '';
  position: absolute; inset: 0;
  background: linear-gradient(135deg, rgba(255,255,255,.025) 0%, transparent 60%);
  pointer-events: none;
}}
.card-glow {{
  position: absolute; top: -20px; right: -20px;
  width: 80px; height: 80px; border-radius: 50%;
  opacity: .12; filter: blur(20px);
  background: var(--c, var(--accent));
}}
.card-label {{ font-size: 11px; color: var(--muted); font-weight: 500; margin-bottom: 10px; }}
.card-value {{
  font-size: 30px; font-weight: 700; line-height: 1;
  color: var(--c, var(--accent));
  font-family: 'JetBrains Mono', monospace;
}}
.card-sub {{ font-size: 10.5px; color: var(--muted); margin-top: 7px; }}
.card-bar {{
  position: absolute; bottom: 0; left: 0; right: 0; height: 3px;
  background: linear-gradient(90deg, var(--c, var(--accent)), transparent);
}}

/* ── Charts ── */
.charts-grid {{
  display: grid; grid-template-columns: 1fr 1fr;
  gap: 16px; margin-bottom: 28px;
}}
@media(max-width:860px) {{ .charts-grid {{ grid-template-columns: 1fr; }} }}
.chart-box {{
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--r); padding: 18px 20px 14px;
}}
.chart-box h3 {{ font-size: 12.5px; font-weight: 600; margin-bottom: 16px; }}
.no-chart {{
  display: none; align-items: center; justify-content: center;
  height: 160px; color: var(--muted); font-size: 12px;
  border: 1px dashed var(--border2); border-radius: 8px;
}}

/* ── Table ── */
.table-wrap {{
  background: var(--surface); border: 1px solid var(--border);
  border-radius: var(--r); overflow: hidden;
}}
.table-top {{
  padding: 15px 20px; border-bottom: 1px solid var(--border);
  display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 6px;
}}
.table-top h3 {{ font-size: 13.5px; font-weight: 600; }}
.date-range {{ font-size: 11px; color: var(--muted); font-family: 'JetBrains Mono', monospace; }}
.tbl-scroll {{ overflow-x: auto; }}
table {{ width: 100%; border-collapse: collapse; font-size: 11.5px; font-family: 'JetBrains Mono', monospace; }}
thead th {{
  padding: 9px 11px; text-align: right;
  font-size: 9.5px; font-weight: 600; letter-spacing: .06em;
  text-transform: uppercase; color: #fff; white-space: nowrap;
  background: var(--surface2); border-bottom: 1px solid var(--border);
  position: sticky; top: 0;
}}
thead th:first-child {{ text-align: left; position: sticky; left: 0; z-index: 4; background: var(--surface2); }}
thead tr:first-child th {{ border-bottom: none; font-size: 10px; }}
thead tr:first-child th.hb {{ background: #0f2454; }}
thead tr:first-child th.hg {{ background: #0b2e18; }}
thead tr:first-child th.ho {{ background: #2d1e04; }}
thead tr:first-child th.hp {{ background: #1e1040; }}
thead tr:first-child th:first-child {{ background: var(--surface2); }}
tbody td {{
  padding: 8px 11px; text-align: right;
  border-bottom: 1px solid rgba(30,39,54,.8);
  color: var(--text); white-space: nowrap;
}}
tbody td:first-child {{ text-align: left; font-weight: 500; position: sticky; left: 0; background: var(--surface); z-index: 1; }}
tbody tr:nth-child(even) td {{ background: rgba(21,26,36,.7); }}
tbody tr:nth-child(even) td:first-child {{ background: rgba(21,26,36,.95); }}
tbody tr:hover td {{ background: rgba(79,142,247,.07) !important; }}
.pct {{ color: var(--muted); }}
.pos {{ color: var(--green); font-weight: 600; }}
.neg {{ color: var(--red);   font-weight: 600; }}
tr.drop-row td {{ background: rgba(181,123,238,.06) !important; border-top: 1px solid rgba(181,123,238,.15); }}
tr.drop-row td:first-child {{ background: rgba(181,123,238,.1) !important; }}
tr.total-row td {{ background: rgba(79,142,247,.08) !important; color: var(--accent) !important; font-weight: 700; border-top: 1px solid rgba(79,142,247,.2); }}
tr.total-row td:first-child {{ color: var(--accent) !important; background: rgba(79,142,247,.12) !important; }}

/* ── Footer ── */
.footer {{ text-align: center; font-size: 11px; color: var(--muted); padding: 24px 0 8px; }}
</style>
</head>
<body>

<div class="header">
  <div class="header-brand">
    <div class="brand-icon">&#9641;</div>
    <div>
      <div class="brand-name">Overview Report Dashboard</div>
      <div class="brand-sub" id="genLbl"></div>
    </div>
  </div>
  <span class="pill" id="daysPill"></span>
</div>

<div id="main">
  <div class="meta-strip" id="metaStrip"></div>

  <div class="sec-label">Key Metrics</div>
  <div class="cards" id="cards"></div>

  <div class="sec-label">Trends</div>
  <div class="charts-grid">
    <div class="chart-box"><h3>Daily Row Volume</h3>
      <canvas id="cRows" height="160"></canvas>
      <div class="no-chart" id="nc1">Charts unavailable — run once with internet to cache Chart.js</div>
    </div>
    <div class="chart-box"><h3>Unique Visitors — Distinct cliIP &amp; Device</h3>
      <canvas id="cIP" height="160"></canvas>
      <div class="no-chart" id="nc2">Charts unavailable</div>
    </div>
    <div class="chart-box"><h3>Session Breakdown</h3>
      <canvas id="cSess" height="160"></canvas>
      <div class="no-chart" id="nc3">Charts unavailable</div>
    </div>
    <div class="chart-box"><h3>Daily Data Transferred (GiB)</h3>
      <canvas id="cBytes" height="160"></canvas>
      <div class="no-chart" id="nc4">Charts unavailable</div>
    </div>
  </div>

  <div class="sec-label">Day-by-Day Breakdown</div>
  <div class="table-wrap">
    <div class="table-top">
      <h3>Full Data Table</h3>
      <span class="date-range" id="drLbl"></span>
    </div>
    <div class="tbl-scroll"><table id="tbl"></table></div>
  </div>

  <div class="footer" id="footer"></div>
</div>

<script>
const D = DATA_BLOB_PLACEHOLDER;
const {{meta, rows, total: tot, generated, source, has_charts}} = D;

document.getElementById('genLbl').textContent = 'Last updated ' + generated + '  ·  ' + source;
document.getElementById('daysPill').textContent = rows.length + ' days';
document.getElementById('footer').textContent = 'To refresh: run generate_dashboard.py after updating ' + source;

document.getElementById('metaStrip').innerHTML =
  Object.entries(meta).map(([k,v]) => `<span><b>${{k}}:</b> ${{v}}</span>`).join('');

const fmt  = (n, d=0) => n == null || isNaN(n) ? '—' :
  Number(n).toLocaleString('en-IN', {{minimumFractionDigits:d, maximumFractionDigits:d}});
const fmtP = n => {{ if(n==null||isNaN(n)) return '—'; const p=Math.abs(n)<1?n*100:n; return p.toFixed(2)+'%'; }};
const ndays = rows.length;

[
  {{l:'Total Rows',         v:fmt(tot.rows),                      c:'#4f8ef7', s:'from Parquet metadata'}},
  {{l:'Total GiB',          v:fmt(tot.bytes,2),                   c:'#30d9a0', s:'data transferred'}},
  {{l:'Distinct cliIPs',    v:fmt(tot.dist_ip),                   c:'#b57bee', s:'unique visitors'}},
  {{l:'Distinct Sessions',  v:fmt(tot.dist_sess),                 c:'#f5a623', s:'session_id values'}},
  {{l:'Days Loaded',        v:ndays,                              c:'#43e07a', s:'date records'}},
  {{l:'Avg Rows / Day',     v:fmt(Math.round(tot.rows/ndays)),    c:'#6ee7f7', s:'daily average'}},
].forEach(k => {{
  document.getElementById('cards').insertAdjacentHTML('beforeend', `
    <div class="card" style="--c:${{k.c}}">
      <div class="card-glow"></div>
      <div class="card-label">${{k.l}}</div>
      <div class="card-value">${{k.v}}</div>
      <div class="card-sub">${{k.s}}</div>
      <div class="card-bar"></div>
    </div>`);
}});

if (rows.length >= 2)
  document.getElementById('drLbl').textContent = rows[0].date + ' → ' + rows[rows.length-1].date;

const CD  = {{tension:.35, fill:true, pointRadius:2, pointHoverRadius:5, borderWidth:2}};
const opt = (yfmt) => {{
  const cb = v => yfmt==='gib' ? v.toFixed(1)+' GiB'
    : v>=1e6 ? (v/1e6).toFixed(1)+'M' : v>=1e3 ? (v/1e3).toFixed(0)+'K' : v;
  return {{
    responsive: true,
    plugins: {{
      legend: {{position:'bottom', labels:{{color:'#5a6a84', font:{{size:10, family:"'Sora'"}}, boxWidth:10}}}},
      tooltip: {{backgroundColor:'#0e1117', borderColor:'#1e2736', borderWidth:1,
                 titleColor:'#d8e0ee', bodyColor:'#5a6a84', padding:10}}
    }},
    scales: {{
      x: {{ticks:{{color:'#5a6a84', font:{{size:9}}, maxTicksLimit:14}}, grid:{{color:'rgba(30,39,54,.6)'}}}},
      y: {{ticks:{{color:'#5a6a84', font:{{size:9}}, callback:cb}},      grid:{{color:'rgba(30,39,54,.6)'}}}}
    }}
  }};
}};

if (has_charts && typeof Chart !== 'undefined') {{
  const lbl = rows.map(r => r.date);
  new Chart(document.getElementById('cRows'),  {{type:'line', data:{{labels:lbl, datasets:[
    {{label:'Total rows', data:rows.map(r=>r.rows), borderColor:'#4f8ef7', backgroundColor:'rgba(79,142,247,.07)', ...CD}}
  ]}}, options:opt()}});
  new Chart(document.getElementById('cIP'),    {{type:'line', data:{{labels:lbl, datasets:[
    {{label:'Distinct cliIP',  data:rows.map(r=>r.dist_ip),  borderColor:'#b57bee', backgroundColor:'rgba(181,123,238,.07)', ...CD}},
    {{label:'Distinct device', data:rows.map(r=>r.dist_dev), borderColor:'#30d9a0', backgroundColor:'rgba(48,217,160,.05)',  ...CD}}
  ]}}, options:opt()}});
  new Chart(document.getElementById('cSess'),  {{type:'line', data:{{labels:lbl, datasets:[
    {{label:'Found+ID avail',   data:rows.map(r=>r.sess_avail), borderColor:'#43e07a', backgroundColor:'rgba(67,224,122,.07)',  ...CD}},
    {{label:'Found-ID missing', data:rows.map(r=>r.sess_na),    borderColor:'#f5a623', backgroundColor:'rgba(245,166,35,.07)',  ...CD}},
    {{label:'Not found',        data:rows.map(r=>r.sess_none),  borderColor:'#f75b5b', backgroundColor:'rgba(247,91,91,.06)',   ...CD}}
  ]}}, options:opt()}});
  new Chart(document.getElementById('cBytes'), {{type:'line', data:{{labels:lbl, datasets:[
    {{label:'GiB transferred', data:rows.map(r=>r.bytes), borderColor:'#30d9a0', backgroundColor:'rgba(48,217,160,.07)', ...CD}}
  ]}}, options:opt('gib')}});
}} else {{
  ['cRows','cIP','cSess','cBytes'].forEach(id => {{ const c=document.getElementById(id); if(c) c.style.display='none'; }});
  ['nc1','nc2','nc3','nc4'].forEach(id  => {{ const c=document.getElementById(id);    if(c) c.style.display='flex'; }});
}}

const colHdrs = [
  'Date (IST)','Bytes (GiB)','Total Rows','cliIP Rows',
  'Dist. cliIP','Dist. cliIP+UA','Dist. Device','Dist. Session',
  'Sess+ID Avail','Sess–ID Miss','Sess Not Found',
  '% cliIP','% cliIP+UA','% Sess+ID','% Sess–ID','% Sess None'
];
let h = '<thead>'
  + '<tr><th style="background:var(--surface2)"></th>'
  + '<th colspan="4" class="hb">Day-by-Day</th>'
  + '<th colspan="4" class="hg">Unique Set</th>'
  + '<th colspan="3" class="ho">Sessions</th>'
  + '<th colspan="5" class="hp">% of Total</th></tr>'
  + '<tr>' + colHdrs.map(s=>`<th>${{s}}</th>`).join('') + '</tr>'
  + '</thead><tbody>';

rows.forEach(d => {{
  h += `<tr>
    <td>${{d.date}}</td>
    <td>${{fmt(d.bytes,2)}}</td><td>${{fmt(d.rows)}}</td><td>${{fmt(d.ip_rows)}}</td>
    <td>${{fmt(d.dist_ip)}}</td><td>${{fmt(d.dist_ipua)}}</td><td>${{fmt(d.dist_dev)}}</td><td>${{fmt(d.dist_sess)}}</td>
    <td>${{fmt(d.sess_avail)}}</td><td>${{fmt(d.sess_na)}}</td><td>${{fmt(d.sess_none)}}</td>
    <td class="pct">${{fmtP(d.pct_ip)}}</td><td class="pct">${{fmtP(d.pct_ipua)}}</td>
    <td class="pct">${{fmtP(d.pct_sess)}}</td><td class="pct">${{fmtP(d.pct_sessna)}}</td>
    <td class="pct">${{fmtP(d.pct_none)}}</td>
  </tr>`;
}});

if (rows.length >= 6) {{
  const last=rows[rows.length-2], prev4=rows.slice(-6,-2);
  const avg =k => prev4.reduce((s,r)=>s+r[k],0)/prev4.length;
  const drop=k => {{ const d=last[k]/avg(k)-1; if(!isFinite(d)) return '—';
    return `<span class="${{d>=0?'pos':'neg'}}">${{d>=0?'+':''}}${{(d*100).toFixed(2)}}%</span>`; }};
  h += `<tr class="drop-row">
    <td style="font-size:10px;color:var(--purple);font-weight:600">Δ vs prev 4-day avg<br>
    <span style="font-weight:400;color:var(--muted);font-size:9px">drop / gain</span></td>
    <td>${{drop('bytes')}}</td><td>${{drop('rows')}}</td><td>${{drop('ip_rows')}}</td>
    <td>${{drop('dist_ip')}}</td><td>${{drop('dist_ipua')}}</td><td>${{drop('dist_dev')}}</td><td>${{drop('dist_sess')}}</td>
    <td>${{drop('sess_avail')}}</td><td>${{drop('sess_na')}}</td><td>${{drop('sess_none')}}</td>
    <td colspan="5" style="text-align:center;color:var(--muted);font-size:9px">— per-row cols only —</td>
  </tr>`;
}}
if (tot && tot.rows) {{
  h += `<tr class="total-row">
    <td>TOTAL</td>
    <td>${{fmt(tot.bytes,2)}}</td><td>${{fmt(tot.rows)}}</td><td>${{fmt(tot.ip_rows)}}</td>
    <td>${{fmt(tot.dist_ip)}}</td><td>${{fmt(tot.dist_ipua)}}</td><td>${{fmt(tot.dist_dev)}}</td><td>${{fmt(tot.dist_sess)}}</td>
    <td>${{fmt(tot.sess_avail)}}</td><td>${{fmt(tot.sess_na)}}</td><td>${{fmt(tot.sess_none)}}</td>
    <td colspan="5" style="text-align:center;opacity:.4;font-size:9px">— % cols are per-row —</td>
  </tr>`;
}}
h += '</tbody>';
document.getElementById('tbl').innerHTML = h;
</script>
</body>
</html>"""

# Inject data
html_final = html.replace("DATA_BLOB_PLACEHOLDER", data_blob)

HTML_OUT.write_text(html_final, encoding="utf-8")
size_kb = HTML_OUT.stat().st_size // 1024
print(f"\nDone → {HTML_OUT}")
print(f"  Size : {size_kb} KB")
print(f"  Days : {len(data_rows)}")
print(f"  Built: {generated_at}")
print(f"\n  Stakeholders open: {HTML_OUT.name}")
