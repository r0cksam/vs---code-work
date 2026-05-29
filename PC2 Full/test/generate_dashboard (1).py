"""
generate_dashboard.py
─────────────────────
Reads overview_report.xlsx → writes a 100% self-contained HTML file.
No CDN, no internet needed at view-time. Works on network shares.

Workflow:
    1. Run overview_generator.py  (updates the Excel as normal)
    2. Run this script            (bakes Excel → HTML, ~2 seconds)
    3. Stakeholder opens dashboard/overview_dashboard.html — done.

Configure the two paths below, then never touch this script again.
"""

import sys, json, urllib.request
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","--quiet"])
    import openpyxl

# ╔══════════════════════════════════════════════════════════════╗
# ║  CONFIGURE THESE TWO PATHS                                   ║
# ╚══════════════════════════════════════════════════════════════╝
XLSX_PATH = Path(r"\\YOUR_SERVER\share\overview_report.xlsx")   # ← your Excel
HTML_OUT  = Path(r"\\YOUR_SERVER\share\dashboard\overview_dashboard.html")  # ← output
# ════════════════════════════════════════════════════════════════

# Override via CLI: python generate_dashboard.py <xlsx> <out.html>
if len(sys.argv) == 3:
    XLSX_PATH = Path(sys.argv[1])
    HTML_OUT  = Path(sys.argv[2])

# ── Column map (1-indexed, matches overview_generator.py) ───────
COL = dict(
    DATE=2, BYTES=3, ROWS=4, IP_ROWS=5,
    DIST_IP=6, DIST_IPUA=7, DIST_DEV=8, DIST_SESS=9,
    UNIQ_IP=10, UNIQ_IPUA=11,
    SESS_AVAIL=12, SESS_NA=13, SESS_NONE=14,
    PCT_IP=15, PCT_IPUA=16, PCT_SESS=17, PCT_SESSNA=18, PCT_NONE=19,
)
DATA_START_ROW = 13

# ── Helpers ──────────────────────────────────────────────────────
def safe_num(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def safe_str(v):
    return str(v).strip() if v is not None else ""


# ── Step 1: Get Chart.js (cache locally next to this script) ────
CHARTJS_CACHE = Path(__file__).parent / ".chartjs_cache.js"
CHARTJS_URL   = "https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"

def get_chartjs() -> str:
    if CHARTJS_CACHE.exists():
        print("  Chart.js: using cached copy")
        return CHARTJS_CACHE.read_text(encoding="utf-8")
    print("  Chart.js: downloading once from CDN…", end=" ", flush=True)
    try:
        with urllib.request.urlopen(CHARTJS_URL, timeout=15) as r:
            js = r.read().decode("utf-8")
        CHARTJS_CACHE.write_text(js, encoding="utf-8")
        print(f"done ({len(js)//1024} KB cached)")
        return js
    except Exception as e:
        print(f"\n  ⚠️  Could not download Chart.js: {e}")
        print("  Charts will be disabled. Run with internet once to cache it.")
        return None


# ── Step 2: Read Excel ───────────────────────────────────────────
print(f"\nReading: {XLSX_PATH}")
if not XLSX_PATH.exists():
    print(f"  ❌  File not found: {XLSX_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb.active

# Metadata rows 2-9
meta = {}
for r in range(2, 10):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    k, v = safe_str(row[1]), safe_str(row[2])
    if k: meta[k] = v

# Data rows
data_rows, total_row = [], None
for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
    ds = safe_str(row[COL['DATE'] - 1])
    if not ds: continue
    if 'TOTAL' in ds.upper(): total_row = row; continue
    if '/' not in ds and len(ds) < 5: continue
    n = lambda c: safe_num(row[COL[c] - 1])
    data_rows.append({
        "date": ds, "bytes": round(n('BYTES'),2),
        "rows": int(n('ROWS')), "ip_rows": int(n('IP_ROWS')),
        "dist_ip": int(n('DIST_IP')), "dist_ipua": int(n('DIST_IPUA')),
        "dist_dev": int(n('DIST_DEV')), "dist_sess": int(n('DIST_SESS')),
        "sess_avail": int(n('SESS_AVAIL')), "sess_na": int(n('SESS_NA')),
        "sess_none": int(n('SESS_NONE')),
        "pct_ip": round(n('PCT_IP'),6), "pct_ipua": round(n('PCT_IPUA'),6),
        "pct_sess": round(n('PCT_SESS'),6), "pct_sessna": round(n('PCT_SESSNA'),6),
        "pct_none": round(n('PCT_NONE'),6),
    })
wb.close()

tot = {}
if total_row:
    tn = lambda c: safe_num(total_row[COL[c]-1])
    tot = {k: int(tn(k)) if k != 'bytes' else round(tn('BYTES'),2)
           for k in ['ROWS','IP_ROWS','DIST_IP','DIST_IPUA','DIST_DEV',
                     'DIST_SESS','SESS_AVAIL','SESS_NA','SESS_NONE']}
    tot['bytes'] = round(safe_num(total_row[COL['BYTES']-1]), 2)
    tot = {k.lower(): v for k, v in tot.items()}

generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
print(f"  {len(data_rows)} data rows found")

chartjs = get_chartjs()
data_blob = json.dumps({
    "meta": meta, "rows": data_rows, "total": tot,
    "generated": generated_at, "source": XLSX_PATH.name,
    "has_charts": chartjs is not None,
}, ensure_ascii=False)

# ── Step 3: Build HTML ───────────────────────────────────────────
chartjs_tag = f"<script>{chartjs}</script>" if chartjs else "<script>window.Chart=null;</script>"

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Overview Dashboard</title>
{chartjs_tag}
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f4f6f9;color:#1a1a2e;min-height:100vh}}
.topbar{{background:#1F3864;color:#fff;padding:13px 26px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
.topbar h1{{font-size:16px;font-weight:600}}
.topbar .sub{{font-size:11.5px;opacity:.6}}
#main{{padding:20px 26px;max-width:1440px;margin:0 auto}}
.meta-strip{{background:#fff;border-radius:10px;border:1px solid #e2e8f0;padding:12px 18px;margin-bottom:16px;display:flex;flex-wrap:wrap;gap:8px 26px;font-size:12px;color:#666}}
.meta-strip b{{color:#1a1a2e;font-weight:600}}
.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:20px}}
.card{{background:#fff;border-radius:10px;border:1px solid #e2e8f0;padding:14px 16px}}
.lbl{{font-size:10.5px;color:#888;margin-bottom:5px;display:flex;align-items:center;gap:5px}}
.dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0}}
.val{{font-size:24px;font-weight:600;line-height:1}}
.sub{{font-size:10px;color:#bbb;margin-top:3px}}
.charts{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:20px}}
@media(max-width:820px){{.charts{{grid-template-columns:1fr}}}}
.cbox{{background:#fff;border-radius:10px;border:1px solid #e2e8f0;padding:14px 16px}}
.cbox h3{{font-size:12px;font-weight:600;margin-bottom:10px;color:#1a1a2e}}
.no-chart{{display:flex;align-items:center;justify-content:center;height:160px;color:#bbb;font-size:12px;border:1px dashed #e2e8f0;border-radius:8px}}
.table-wrap{{background:#fff;border-radius:10px;border:1px solid #e2e8f0;overflow:hidden;margin-bottom:18px}}
.th{{padding:12px 16px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid #e2e8f0}}
.th h3{{font-size:12.5px;font-weight:600}}
.tbl-c{{overflow-x:auto}}
table{{width:100%;border-collapse:collapse;font-size:11.5px}}
thead tr:first-child th{{padding:5px 10px;font-size:9.5px;font-weight:700;color:#fff;text-align:center;white-space:nowrap}}
thead tr:last-child th{{padding:6px 10px;text-align:right;font-size:10px;font-weight:600;color:#555;border-bottom:1px solid #e2e8f0;white-space:nowrap;background:#f8fafc;position:sticky;top:0}}
thead tr:last-child th:first-child{{text-align:left;position:sticky;left:0;background:#f8fafc;z-index:3}}
thead tr:first-child th:first-child{{position:sticky;left:0;z-index:3;background:#f8fafc}}
td{{padding:6px 10px;text-align:right;border-bottom:.5px solid #f0f0f0;white-space:nowrap}}
td:first-child{{text-align:left;font-weight:500;position:sticky;left:0;background:inherit;z-index:1}}
tr:nth-child(even) td{{background:#fafafa}}
tr:hover td{{background:#eef5ff!important}}
tr.total-row td{{background:#1F3864!important;color:#fff!important;font-weight:700}}
tr.drop-row td{{background:#f8f9ff}}
.pct{{color:#185FA5;font-weight:500}}
.up{{color:#3B6D11;font-weight:600}}
.dn{{color:#A32D2D;font-weight:600}}
.footer{{text-align:center;font-size:10.5px;color:#ccc;padding:8px 0 20px}}
th.b{{background:#1F3864}} th.g{{background:#3B6D11}}
th.a{{background:#BA7517}} th.p{{background:#534AB7}}
</style>
</head>
<body>
<div class="topbar">
  <h1>&#128202;&nbsp; Overview Report Dashboard</h1>
  <span class="sub" id="genLbl"></span>
</div>
<div id="main">
  <div class="meta-strip" id="metaStrip"></div>
  <div class="cards" id="cards"></div>
  <div class="charts">
    <div class="cbox"><h3>Daily row volume</h3>
      <canvas id="cRows" height="180"></canvas><div class="no-chart" id="nc1" style="display:none">Charts unavailable (Chart.js not cached)</div></div>
    <div class="cbox"><h3>Unique visitors (distinct cliIP)</h3>
      <canvas id="cIP" height="180"></canvas><div class="no-chart" id="nc2" style="display:none">Charts unavailable</div></div>
    <div class="cbox"><h3>Session breakdown</h3>
      <canvas id="cSess" height="180"></canvas><div class="no-chart" id="nc3" style="display:none">Charts unavailable</div></div>
    <div class="cbox"><h3>Daily GiB transferred</h3>
      <canvas id="cBytes" height="180"></canvas><div class="no-chart" id="nc4" style="display:none">Charts unavailable</div></div>
  </div>
  <div class="table-wrap">
    <div class="th">
      <h3>Day-by-Day Breakdown</h3>
      <span style="font-size:11px;color:#aaa" id="drLbl"></span>
    </div>
    <div class="tbl-c"><table id="tbl"></table></div>
  </div>
  <div class="footer" id="footer"></div>
</div>
<script>
const D = {data_blob};
const {{meta,rows,total:tot,generated,source,has_charts}} = D;

document.getElementById('genLbl').textContent =
  'Last updated ' + generated + '\u2003Source: ' + source;
document.getElementById('footer').textContent =
  'To refresh: run generate_dashboard.py after updating ' + source;

document.getElementById('metaStrip').innerHTML =
  Object.entries(meta).map(([k,v])=>`<span><b>${{k}}:</b> ${{v}}</span>`).join('');

const fmt=(n,d=0)=>n==null||isNaN(n)?'\u2014':
  Number(n).toLocaleString('en-IN',{{minimumFractionDigits:d,maximumFractionDigits:d}});
const fmtP=n=>{{if(n==null||isNaN(n))return'\u2014';const p=Math.abs(n)<1?n*100:n;return p.toFixed(2)+'%';}};

const ndays=rows.length;
[
  {{l:'Total rows',      v:fmt(tot.rows),                clr:'#1F3864',s:'Parquet metadata'}},
  {{l:'Total GiB',       v:fmt(tot.bytes,2),             clr:'#0F6E56',s:'data transferred'}},
  {{l:'Distinct cliIPs', v:fmt(tot.dist_ip),             clr:'#534AB7',s:'unique visitors'}},
  {{l:'Distinct sessions',v:fmt(tot.dist_sess),          clr:'#BA7517',s:'session_id values'}},
  {{l:'Days loaded',     v:ndays,                        clr:'#3B6D11',s:'date records'}},
  {{l:'Avg rows / day',  v:fmt(Math.round(tot.rows/ndays)),clr:'#A32D2D',s:'daily average'}},
].forEach(c=>{{
  document.getElementById('cards').insertAdjacentHTML('beforeend',
    `<div class="card"><div class="lbl"><span class="dot" style="background:${{c.clr}}"></span>${{c.l}}</div>
     <div class="val" style="color:${{c.clr}}">${{c.v}}</div><div class="sub">${{c.s}}</div></div>`);
}});

if(rows.length>=2)
  document.getElementById('drLbl').textContent=rows[0].date+' \u2192 '+rows[rows.length-1].date;

if(has_charts && typeof Chart!=='undefined'){{
  const lbl=rows.map(r=>r.date);
  const CD={{tension:.35,fill:true,pointRadius:3,pointHoverRadius:5,borderWidth:2}};
  const opt=(yfmt)=>{{
    const cb=v=>yfmt==='gib'?v.toFixed(1)+' GiB':v>=1e6?(v/1e6).toFixed(1)+'M':v>=1e3?(v/1e3).toFixed(0)+'K':v;
    return{{responsive:true,
      plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}},boxWidth:10}}}}}},
      scales:{{x:{{ticks:{{font:{{size:9}},maxTicksLimit:14}},grid:{{color:'#f0f0f0'}}}},
               y:{{ticks:{{font:{{size:9}},callback:cb}},grid:{{color:'#f0f0f0'}}}}}}}};
  }};
  new Chart(document.getElementById('cRows'),{{type:'line',data:{{labels:lbl,
    datasets:[{{label:'Total rows',data:rows.map(r=>r.rows),borderColor:'#185FA5',backgroundColor:'rgba(24,95,165,.08)',...CD}}]}},options:opt()}});
  new Chart(document.getElementById('cIP'),{{type:'line',data:{{labels:lbl,
    datasets:[
      {{label:'Distinct cliIP',data:rows.map(r=>r.dist_ip),borderColor:'#534AB7',backgroundColor:'rgba(83,74,183,.08)',...CD}},
      {{label:'Distinct device',data:rows.map(r=>r.dist_dev),borderColor:'#0F6E56',backgroundColor:'rgba(15,110,86,.06)',...CD}}
    ]}},options:opt()}});
  new Chart(document.getElementById('cSess'),{{type:'line',data:{{labels:lbl,
    datasets:[
      {{label:'Found+ID avail',data:rows.map(r=>r.sess_avail),borderColor:'#3B6D11',backgroundColor:'rgba(59,109,17,.08)',...CD}},
      {{label:'Found-ID missing',data:rows.map(r=>r.sess_na),borderColor:'#BA7517',backgroundColor:'rgba(186,117,23,.07)',...CD}},
      {{label:'Not found',data:rows.map(r=>r.sess_none),borderColor:'#A32D2D',backgroundColor:'rgba(163,45,45,.07)',...CD}}
    ]}},options:opt()}});
  new Chart(document.getElementById('cBytes'),{{type:'line',data:{{labels:lbl,
    datasets:[{{label:'GiB',data:rows.map(r=>r.bytes),borderColor:'#1D9E75',backgroundColor:'rgba(29,158,117,.09)',...CD}}]}},options:opt('gib')}});
}}else{{
  ['nc1','nc2','nc3','nc4'].forEach(id=>{{
    document.getElementById(id).style.display='flex';
    document.getElementById(id.replace('nc','cRows')||id.replace(/nc\d/,m=>(['cRows','cIP','cSess','cBytes'][+m.slice(2)-1]))).style.display='none';
  }});
  ['cRows','cIP','cSess','cBytes'].forEach(id=>{{const c=document.getElementById(id);if(c)c.style.display='none';}});
  ['nc1','nc2','nc3','nc4'].forEach(id=>{{const c=document.getElementById(id);if(c)c.style.display='flex';}});
}}

const subH=['Date (IST)','Bytes (GiB)','Total Rows','cliIP rows',
  'Distinct cliIP','Dist cliIP+UA','Dist device','Dist session',
  'Sess+ID avail','Sess-ID missing','Sess not found',
  '% cliIP','% cliIP+UA','% Sess+ID','% Sess-ID','% Sess none'];
let h='<thead><tr><th class="b" style="background:#f8fafc"></th>'
  +'<th colspan="4" class="b">Day-by-Day Breakdown</th>'
  +'<th colspan="4" class="g">Unique Set</th>'
  +'<th colspan="3" class="a">Rows Against</th>'
  +'<th colspan="5" class="p">% of Total</th></tr>'
  +'<tr>'+subH.map(s=>`<th>${{s}}</th>`).join('')+'</tr></thead><tbody>';

rows.forEach((d,i)=>{{
  h+=`<tr><td>${{d.date}}</td><td>${{fmt(d.bytes,2)}}</td><td>${{fmt(d.rows)}}</td><td>${{fmt(d.ip_rows)}}</td>
    <td>${{fmt(d.dist_ip)}}</td><td>${{fmt(d.dist_ipua)}}</td><td>${{fmt(d.dist_dev)}}</td><td>${{fmt(d.dist_sess)}}</td>
    <td>${{fmt(d.sess_avail)}}</td><td>${{fmt(d.sess_na)}}</td><td>${{fmt(d.sess_none)}}</td>
    <td class="pct">${{fmtP(d.pct_ip)}}</td><td class="pct">${{fmtP(d.pct_ipua)}}</td>
    <td class="pct">${{fmtP(d.pct_sess)}}</td><td class="pct">${{fmtP(d.pct_sessna)}}</td>
    <td class="pct">${{fmtP(d.pct_none)}}</td></tr>`;
}});

if(rows.length>=6){{
  const last=rows[rows.length-2], prev=rows.slice(-6,-2);
  const avg=k=>prev.reduce((s,r)=>s+r[k],0)/prev.length;
  const drop=k=>{{const d=last[k]/avg(k)-1;if(!isFinite(d))return'\u2014';
    return`<span class="${{d>=0?'up':'dn'}}">${{d>=0?'+':''}}${{(d*100).toFixed(2)}}%</span>`;}};
  h+=`<tr class="drop-row"><td style="font-weight:600;font-size:10px;color:#534AB7">% Drop/Gain<br>
    <small style="font-weight:400;color:#bbb">vs prev 4-day avg</small></td>
    <td>${{drop('bytes')}}</td><td>${{drop('rows')}}</td><td>${{drop('ip_rows')}}</td>
    <td>${{drop('dist_ip')}}</td><td>${{drop('dist_ipua')}}</td><td>${{drop('dist_dev')}}</td><td>${{drop('dist_sess')}}</td>
    <td>${{drop('sess_avail')}}</td><td>${{drop('sess_na')}}</td><td>${{drop('sess_none')}}</td>
    <td colspan="5" style="text-align:center;color:#ccc;font-size:9.5px">— per-row only —</td></tr>`;
}}
if(tot&&tot.rows){{
  h+=`<tr class="total-row"><td>TOTAL</td><td>${{fmt(tot.bytes,2)}}</td><td>${{fmt(tot.rows)}}</td><td>${{fmt(tot.ip_rows)}}</td>
    <td>${{fmt(tot.dist_ip)}}</td><td>${{fmt(tot.dist_ipua)}}</td><td>${{fmt(tot.dist_dev)}}</td><td>${{fmt(tot.dist_sess)}}</td>
    <td>${{fmt(tot.sess_avail)}}</td><td>${{fmt(tot.sess_na)}}</td><td>${{fmt(tot.sess_none)}}</td>
    <td colspan="5" style="text-align:center;opacity:.6;font-size:9.5px">— % cols per-row —</td></tr>`;
}}
h+='</tbody>';
document.getElementById('tbl').innerHTML=h;
</script>
</body>
</html>"""

html_final = html.replace("{data_blob}", data_blob)

HTML_OUT.parent.mkdir(parents=True, exist_ok=True)
HTML_OUT.write_text(html_final, encoding="utf-8")
size_kb = HTML_OUT.stat().st_size // 1024
print(f"\nDone \u2192 {HTML_OUT}")
print(f"  Size: {size_kb} KB | {len(data_rows)} days | generated {generated_at}")
print(f"  Stakeholder opens: {HTML_OUT.name}")
