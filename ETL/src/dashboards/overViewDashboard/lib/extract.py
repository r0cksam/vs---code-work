"""
lib/extract.py — Data extraction from Excel and CSV sources.
"""

import csv
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "Missing Python dependency: openpyxl. Install ETL\\requirements.txt before running the dashboard."
    ) from exc

# ── Column map (1-based) ──────────────────────────────────────────────────────
COL = {
    'DATE': 2, 'BYTES': 3, 'ROWS': 4, 'IP_ROWS': 5,
    'DIST_IP': 6, 'DIST_IPUA': 7, 'DIST_DEV': 8, 'DIST_SESS': 9,
    'DIST_IP_R2': 10, 'DIST_IPUA_R2': 11,
    'SESS_AVAIL': 12, 'SESS_NA': 13, 'SESS_NONE': 14,
    'PCT_IP': 15, 'PCT_IPUA': 16, 'PCT_SESS': 17, 'PCT_SESSNA': 18, 'PCT_NONE': 19,
}
DATA_START_ROW = 13

# Date formats tried in order
_DATE_FMTS = ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y")

# Meta keys to exclude from the strip
_EXCLUDE_META = frozenset([
    "report generated", "filtered to", "lake folder", "duck db", "duckdb",
    "parquet files", "date range", "data time range", "total rows",
])


# ── Helpers ───────────────────────────────────────────────────────────────────

def coerce_float(v) -> float:
    """Safely coerce any cell value to float, returning 0.0 on failure."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def get_cell(row: tuple, col_num: int):
    """Safely fetch cell value by 1-based column index."""
    idx = col_num - 1
    return row[idx] if idx < len(row) else None


def parse_date(raw) -> str | None:
    """
    Normalize a raw cell value to 'YYYY-MM-DD'.
    Returns None for blank, TOTAL, or unparseable values.
    """
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")

    s = str(raw).strip() if raw is not None else ""
    if not s or s in ("None", "nan") or "TOTAL" in s.upper():
        return None

    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    # No slash or dash at all — unrecognisable
    if "/" not in s and "-" not in s:
        return None
    return None


def _pct_val(raw, numerator: float, denominator: float) -> float:
    """Return raw percentage if present, else compute numerator/denominator."""
    if raw not in (None, "", "None", "nan"):
        try:
            return float(raw)
        except (ValueError, TypeError):
            pass
    return round(numerator / denominator, 6) if denominator > 0 else 0.0


# ── Excel reader ─────────────────────────────────────────────────────────────

def read_excel(xlsx_path: Path) -> tuple[dict, str, list[dict]]:
    """
    Returns (meta, data_time_range, data_rows).
    Excludes the current report date automatically, if present.
    """
    print(f"Reading: {xlsx_path}")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    # read_only=False required to access row_dimensions (hidden row check)
    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb.active

    # ── Meta rows 2-9 ────────────────────────────────────────────────────────
    meta: dict[str, str] = {}
    data_time_range = ""
    for row in ws.iter_rows(min_row=2, max_row=9, values_only=True):
        k = str(row[1]).strip() if row[1] else ""
        v = str(row[2]).strip() if row[2] else ""
        k_lower = k.lower()
        if k and v and ("data time range" in k_lower or "date range" in k_lower):
            data_time_range = v
        if k and k not in ("None", "nan") and not any(ex in k_lower for ex in _EXCLUDE_META):
            meta[k] = v

    # ── Data rows ────────────────────────────────────────────────────────────
    data_rows: list[dict] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
    ):
        if ws.row_dimensions[row_idx].hidden:
            continue

        ds = parse_date(get_cell(row, COL["DATE"]))
        if ds is None:
            continue

        ip = int(coerce_float(get_cell(row, COL["IP_ROWS"])))
        sa = int(coerce_float(get_cell(row, COL["SESS_AVAIL"])))
        sna = int(coerce_float(get_cell(row, COL["SESS_NA"])))
        sno = int(coerce_float(get_cell(row, COL["SESS_NONE"])))

        dist_ip = int(coerce_float(get_cell(row, COL["DIST_IP"])))
        dist_ipua = int(coerce_float(get_cell(row, COL["DIST_IPUA"])))

        raw_dip_r2 = get_cell(row, COL["DIST_IP_R2"])
        raw_dipua_r2 = get_cell(row, COL["DIST_IPUA_R2"])
        dist_ip_r2 = (
            int(coerce_float(raw_dip_r2))
            if raw_dip_r2 not in (None, "", "None", "nan")
            else dist_ip
        )
        dist_ipua_r2 = (
            int(coerce_float(raw_dipua_r2))
            if raw_dipua_r2 not in (None, "", "None", "nan")
            else dist_ipua
        )

        data_rows.append({
            "date":         ds,
            "bytes":        round(coerce_float(get_cell(row, COL["BYTES"])), 2),
            "rows":         int(coerce_float(get_cell(row, COL["ROWS"]))),
            "ip_rows":      ip,
            "dist_ip":      dist_ip,
            "dist_ipua":    dist_ipua,
            "dist_dev":     int(coerce_float(get_cell(row, COL["DIST_DEV"]))),
            "dist_sess":    int(coerce_float(get_cell(row, COL["DIST_SESS"]))),
            "dist_ip_r2":   dist_ip_r2,
            "dist_ipua_r2": dist_ipua_r2,
            "sess_avail":   sa,
            "sess_na":      sna,
            "sess_none":    sno,
            "pct_ip":       _pct_val(get_cell(row, COL["PCT_IP"]),    dist_ip,   ip),
            "pct_ipua":     _pct_val(get_cell(row, COL["PCT_IPUA"]),  dist_ipua, ip),
            "pct_sess":     _pct_val(get_cell(row, COL["PCT_SESS"]),  sa,        ip),
            "pct_sessna":   _pct_val(get_cell(row, COL["PCT_SESSNA"]), sna,      ip),
            "pct_none":     _pct_val(get_cell(row, COL["PCT_NONE"]),  sno,       ip),
        })

    wb.close()

    # Drop only today's date. Yesterday's downloaded folder is normally complete
    # and must remain visible in the stakeholder dashboard.
    if data_rows:
        today = datetime.now().strftime("%Y-%m-%d")
        data_rows = [r for r in data_rows if r["date"] != today]

    return meta, data_time_range, data_rows


def read_source_daily(path: Path) -> list[dict]:
    """Read optional source/day rows for dashboard source filtering."""
    if not path.exists():
        return []
    out: list[dict] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            ds = parse_date(row.get("date"))
            if not ds:
                continue
            ip = int(coerce_float(row.get("ip_rows")))
            dist_ip = int(coerce_float(row.get("dist_ip")))
            dist_ipua = int(coerce_float(row.get("dist_ipua")))
            sess_avail = int(coerce_float(row.get("sess_avail")))
            sess_na = int(coerce_float(row.get("sess_na")))
            sess_none = int(coerce_float(row.get("sess_none")))
            out.append({
                "source":       str(row.get("source") or "stream").strip().lower() or "stream",
                "date":         ds,
                "bytes":        round(coerce_float(row.get("bytes")), 2),
                "rows":         int(coerce_float(row.get("rows"))),
                "ip_rows":      ip,
                "dist_ip":      dist_ip,
                "dist_ipua":    dist_ipua,
                "dist_dev":     int(coerce_float(row.get("dist_dev"))),
                "dist_sess":    int(coerce_float(row.get("dist_sess"))),
                "dist_ip_r2":   int(coerce_float(row.get("dist_ip_r2"))) or dist_ip,
                "dist_ipua_r2": int(coerce_float(row.get("dist_ipua_r2"))) or dist_ipua,
                "sess_avail":   sess_avail,
                "sess_na":      sess_na,
                "sess_none":    sess_none,
                "pct_ip":       _pct_val(row.get("pct_ip"), dist_ip, ip),
                "pct_ipua":     _pct_val(row.get("pct_ipua"), dist_ipua, ip),
                "pct_sess":     _pct_val(row.get("pct_sess"), sess_avail, ip),
                "pct_sessna":   _pct_val(row.get("pct_sessna"), sess_na, ip),
                "pct_none":     _pct_val(row.get("pct_none"), sess_none, ip),
            })
    return out
