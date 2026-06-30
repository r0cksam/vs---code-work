#!/usr/bin/env python3
"""
overview_generator.py
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
Standalone, append-aware Overview sheet generator.
Reproduces the exact Overview format with colored section headers.

First run  â†’ creates  <output_dir>/overview_report.xlsx
Later runs â†’ queries only new dates, merges with existing data,
             rewrites the file so formatting stays consistent.

Run:
    python overview_generator.py
    python overview_generator.py <lake_folder>

Requires:
    duckdb, pandas, pyarrow, openpyxl, psutil
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import psutil
import duckdb
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# LOGGING
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

logging.basicConfig(level=logging.INFO, format="  %(levelname)-7s %(message)s")
log = logging.getLogger("overview_generator")


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# CONSTANTS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

IST_OFFSET        = timedelta(hours=5, minutes=30)
IST_OFFSET_SECONDS = int(IST_OFFSET.total_seconds())
OVERVIEW_FILENAME = "overview_report.xlsx"
SOURCE_DAILY_FILENAME = "overview_source_daily.csv"
OVERVIEW_SHEET    = "Overview"

# â”€â”€ Column indices (1-indexed for openpyxl) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#   A (1) = reserved for % Drop label
#   B (2) = Date (IST)  â€¦  S (19) = % Session Not Found
COL_DATE         = 2   # B
COL_BYTES        = 3   # C â€“ Total Bytes (GiB)
COL_ROWS         = 4   # D â€“ Total Rows* (from Parquet metadata)
COL_IP_ROWS      = 5   # E â€“ cliIP rows
COL_DIST_IP      = 6   # F â€“ Distinct cliIP
COL_DIST_IP_UA   = 7   # G â€“ Distinct cliIP & UA
COL_DIST_DEV     = 8   # H â€“ Distinct device_id
COL_DIST_SESS    = 9   # I â€“ Distinct session_id
COL_UNIQ_IP      = 10  # J â€“ UNIQUE SET: Distinct cliIP  (formula =E)
COL_UNIQ_IP_UA   = 11  # K â€“ UNIQUE SET: Distinct cliIP & UA  (formula =E)
COL_SESS_AVAIL   = 12  # L â€“ ROWS AGAINST: Session Found - ID Available
COL_SESS_NA      = 13  # M â€“ ROWS AGAINST: Session Found - ID NOT Available
COL_SESS_NONE    = 14  # N â€“ ROWS AGAINST: Session Not Found
COL_PCT_IP       = 15  # O â€“ % Distinct cliIP  (formula =J/$E)
COL_PCT_IP_UA    = 16  # P â€“ % Distinct cliIP & UA  (formula =K/$E)
COL_PCT_SESS     = 17  # Q â€“ % Session Found - ID Available  (formula =L/$E)
COL_PCT_SESS_NA  = 18  # R â€“ % Session Found - ID NOT Available  (formula =M/$E)
COL_PCT_NONE     = 19  # S â€“ % Session Not Found  (formula =N/$E)
LAST_COL         = COL_PCT_NONE  # 19

# â”€â”€ Fixed row positions â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
ROW_TITLE_META   = 1
ROW_META_START   = 2   # first key/value pair
ROW_SECTION_HDR  = 11  # colored group header row
ROW_COL_HDR      = 12  # column-label header row
DATA_START_ROW   = 13  # first data row

# â”€â”€ Colors (hex, no leading #) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
C_DARK_BLUE = "1F3864"   # title / day-breakdown section
C_KV_BG     = "D6E4F0"   # metadata key background
C_WHITE     = "FFFFFF"
C_BLACK     = "000000"
C_GREEN     = "92D050"   # UNIQUE SET section
C_AMBER     = "FFC000"   # ROWS AGAINST section
C_PURPLE    = "7030A0"   # % Data of TOTAL section
C_ORANGE    = "E36209"   # TOTAL row
C_ALT       = "F2F2F2"   # alternating row background


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SECTION A â€“ DuckDB / Lake helpers
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def get_conn() -> duckdb.DuckDBPyConnection:
    mem_avail = psutil.virtual_memory().available / (1024 ** 3)
    cpus      = os.cpu_count() or 4
    mem_gb    = max(4, min(int(mem_avail * 0.70), 48))
    threads   = max(1, min(cpus // 4, 4))
    con = duckdb.connect()
    con.execute(f"SET threads={threads}")
    con.execute(f"SET memory_limit='{mem_gb}GB'")
    con.execute("SET preserve_insertion_order=false")
    return con


def sql_path(path: Path) -> str:
    return path.as_posix().replace("'", "''")


def scoped_roots(lake_root: Path, year: str | None, month: str | None) -> list[Path]:
    if not year:
        return [lake_root]

    suffix = [f"year={year}"]
    if month:
        suffix.append(f"month={month}")

    roots: list[Path] = []
    legacy = lake_root.joinpath(*suffix)
    if legacy.exists():
        roots.append(legacy)
    for source_root in sorted(lake_root.glob("source=*")):
        candidate = source_root.joinpath(*suffix)
        if candidate.exists():
            roots.append(candidate)
    return roots


def scoped_parquet_globs(lake_root: Path, year: str | None, month: str | None) -> list[str]:
    roots = scoped_roots(lake_root, year, month)
    if not roots:
        return []
    return [f"{sql_path(root)}/**/*.parquet" for root in roots]


def scoped_partition_exists(lake_root: Path, year: str | None, month: str | None) -> bool:
    return any(root.exists() and any(root.rglob("*.parquet")) for root in scoped_roots(lake_root, year, month))


def ist_timestamp_sql(epoch_expr: str) -> str:
    return (
        "epoch_ms(CAST(FLOOR(("
        f"TRY_CAST({epoch_expr} AS DOUBLE) + {IST_OFFSET_SECONDS}"
        ") * 1000) AS BIGINT))"
    )


def lake_reader(lake_root: Path, year: str | None, month: str | None) -> str:
    globs = scoped_parquet_globs(lake_root, year, month)
    if len(globs) == 1:
        return (
            f"read_parquet('{globs[0]}', "
            f"hive_partitioning=true, union_by_name=true)"
        )
    if len(globs) > 1:
        list_sql = "[" + ", ".join(f"'{glob}'" for glob in globs) + "]"
        return (
            f"read_parquet({list_sql}, "
            f"hive_partitioning=true, union_by_name=true)"
        )
    return (
        f"read_parquet('{sql_path(lake_root)}/**/*.parquet', "
        f"hive_partitioning=true, union_by_name=true)"
    )


def qs_extract(qs_col: str, param: str) -> str:
    return f"NULLIF(regexp_extract({qs_col}, '(?:^|&){param}=([^&]+)', 1), '')"


def pa_schema(lake_root: Path, year: str | None, month: str | None) -> set[str]:
    cols: set[str] = set()
    skip = {"year", "month", "day"}
    preferred = {"reqTimeSec", "queryStr", "cliIP", "UA"}
    for scan_root in scoped_roots(lake_root, year, month) or [lake_root]:
        for pf in scan_root.rglob("*.parquet"):
            try:
                s = pq.read_schema(pf)
                cols.update(n for n in s.names if n not in skip)
            except Exception as exc:
                log.warning(f"Could not read parquet schema for {pf}: {exc}")
            if preferred <= cols:
                break
        if preferred <= cols:
            break
    return cols


def pa_row_counts_by_day(lake_root: Path, year: str | None, month: str | None) -> dict[str, int]:
    counts: dict[str, int] = {}
    for scan_root in scoped_roots(lake_root, year, month) or [lake_root]:
        for day_dir in sorted(scan_root.rglob("day=*")):
            if not day_dir.is_dir():
                continue
            try:
                parts = {p.split("=")[0]: p.split("=")[1] for p in day_dir.parts if "=" in p}
                key = f"{parts.get('year','0')}-{parts.get('month','0')}-{parts.get('day','0')}"
                total = sum(pq.read_metadata(pf).num_rows for pf in day_dir.glob("*.parquet"))
                counts[key] = counts.get(key, 0) + total
            except Exception:
                pass
    return counts


def pa_total_file_count(root: Path) -> int:
    return sum(1 for _ in root.rglob("*.parquet"))


def pa_total_file_count_scoped(lake_root: Path, year: str | None, month: str | None) -> int:
    return sum(pa_total_file_count(root) for root in scoped_roots(lake_root, year, month) or [lake_root])


def safe_query(
    con,
    sql: str,
    label: str = "",
    *,
    raise_on_error: bool = False,
) -> pd.DataFrame:
    try:
        return con.execute(sql).df()
    except Exception as exc:
        message = f"Query failed [{label}]: {exc}"
        if raise_on_error:
            raise RuntimeError(message) from exc
        log.warning(message)
        return pd.DataFrame()


def partition_date_to_ist_str(partition_date) -> str:
    """Format an IST lake partition date as 'DD/MM/YY'."""
    if isinstance(partition_date, pd.Timestamp):
        dt = partition_date.to_pydatetime()
    elif isinstance(partition_date, datetime):
        dt = partition_date
    else:
        dt = datetime.combine(partition_date, datetime.min.time())
    return dt.strftime("%d/%m/%y")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SECTION B â€“ Prompt helpers
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def prompt_path(question: str) -> Path:
    while True:
        raw = input(question).strip()
        if not raw:
            print("  Path cannot be empty.")
            continue
        p = Path(raw).expanduser().resolve()
        if p.is_dir():
            return p
        print(f"  âŒ Directory not found: {p}")


def prompt_year_month(
    lake_root: Path,
    year_pref: str | None = None,
    month_pref: str | None = None,
) -> tuple[str | None, str | None]:
    if year_pref is not None or month_pref is not None:
        year = year_pref.strip() if year_pref else None
        month = month_pref.strip() if month_pref else None
        if year and (not year.isdigit() or len(year) != 4):
            raise SystemExit("Month filter year must be YYYY.")
        if month and (not month.isdigit() or len(month) != 2 or not (1 <= int(month) <= 12)):
            raise SystemExit("Month filter month must be MM (01-12).")
        if month and not year:
            raise SystemExit("Month cannot be provided without year.")
        if year and month:
            if not scoped_partition_exists(lake_root, year, month):
                print(f"  âš ï¸  Partition not found for year={year}, month={month}. Continuing with full lake.")
                return None, None
        return year, month

    print("\nðŸ“…  Filter by month? (y/n): ", end="")
    try:
        if input().strip().lower() != "y":
            return None, None
        available = sorted(
            d.name.replace("year=", "")
            for d in lake_root.glob("**/year=*")
            if d.is_dir()
        )
        available = sorted(set(available))
        if available:
            print(f"  Available years: {', '.join(available)}")
        while True:
            y = input("  Year (e.g. 2026): ").strip()
            if y.isdigit() and 2000 <= int(y) <= 2100:
                break
            print("  âŒ Invalid year")
        while True:
            m = input("  Month 1â€“12: ").strip()
            if m.isdigit() and 1 <= int(m) <= 12:
                break
            print("  âŒ Invalid month")
        year_str  = y
        month_str = f"{int(m):02d}"
        if not scoped_partition_exists(lake_root, year_str, month_str):
            print(f"  âš ï¸  Partition not found for year={year_str}, month={month_str}. Continue? (y/n): ", end="")
            if input().strip().lower() != "y":
                return None, None
        return year_str, month_str
    except KeyboardInterrupt:
        print("\n  Cancelled.")
        return None, None


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SECTION C â€“ Append helpers  (read existing Overview data)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def get_existing_dates(output_path: Path) -> set[str]:
    """Return the set of 'DD/MM/YY' IST date strings already in the Overview."""
    if not output_path.exists():
        return set()
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
        if OVERVIEW_SHEET not in wb.sheetnames:
            wb.close()
            return set()
        ws = wb[OVERVIEW_SHEET]
        dates: set[str] = set()
        for row in ws.iter_rows(
            min_row=DATA_START_ROW, max_col=COL_DATE, min_col=COL_DATE, values_only=True
        ):
            val = row[0]
            if val and isinstance(val, str) and "/" in val:
                dates.add(val.strip())
        wb.close()
        return dates
    except Exception as exc:
        log.warning(f"Could not read existing dates: {exc}")
        return set()


def get_existing_row_counts(output_path: Path) -> dict[str, int]:
    """Return existing Overview row counts by 'DD/MM/YY' IST date."""
    if not output_path.exists():
        return {}
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
        if OVERVIEW_SHEET not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb[OVERVIEW_SHEET]
        counts: dict[str, int] = {}
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            date_val = row[COL_DATE - 1]
            if not date_val or not isinstance(date_val, str) or "/" not in str(date_val):
                break
            try:
                counts[str(date_val).strip()] = int(row[COL_ROWS - 1] or 0)
            except Exception:
                counts[str(date_val).strip()] = 0
        wb.close()
        return counts
    except Exception as exc:
        log.warning(f"Could not read existing row counts: {exc}")
        return {}


def load_existing_rows(output_path: Path) -> list[dict]:
    """Load data rows from an existing Overview sheet as a list of dicts."""
    if not output_path.exists():
        return []
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
        if OVERVIEW_SHEET not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[OVERVIEW_SHEET]
        rows: list[dict] = []
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            date_val = row[COL_DATE - 1]        # 0-indexed
            if not date_val or not isinstance(date_val, str) or "/" not in str(date_val):
                break  # hit TOTAL row or empty area
            def _int(v):
                try:
                    return int(v) if v is not None else 0
                except Exception:
                    return 0
            def _flt(v):
                try:
                    return float(v) if v is not None else 0.0
                except Exception:
                    return 0.0
            rows.append({
                "ist_date":   str(date_val).strip(),
                "bytes_gib":  _flt(row[COL_BYTES      - 1]),
                "total_rows": _int(row[COL_ROWS        - 1]),
                "ip_rows":    _int(row[COL_IP_ROWS     - 1]),
                "dist_ip":    _int(row[COL_DIST_IP     - 1]),
                "dist_ip_ua": _int(row[COL_DIST_IP_UA  - 1]),
                "dist_dev":   _int(row[COL_DIST_DEV    - 1]),
                "dist_sess":  _int(row[COL_DIST_SESS   - 1]),
                "sess_avail": _int(row[COL_SESS_AVAIL  - 1]),
                "sess_na":    _int(row[COL_SESS_NA     - 1]),
                "sess_none":  _int(row[COL_SESS_NONE   - 1]),
            })
        wb.close()
        return rows
    except Exception as exc:
        log.warning(f"Could not load existing rows: {exc}")
        return []


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SECTION D â€“ DuckDB query
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def query_new_dates(
    con,
    lake_root: Path,
    year: str | None,
    month: str | None,
    cols: set[str],
    existing_row_counts: dict[str, int],
    day_counts: dict[str, int],
) -> list[dict]:
    """Query DuckDB for day-by-day stats, refreshing dates whose parquet row count changed."""
    ip_col = "cliIP"
    ua_col = "UA"
    qs_col = "queryStr"
    ts_col = "reqTimeSec"

    if ts_col not in cols:
        log.warning("reqTimeSec not found â€” cannot build day breakdown.")
        return []

    has_ip = ip_col in cols
    has_ua = ua_col in cols
    has_qs = qs_col in cols
    has_tb = "totalBytes" in cols

    ip_rows_expr   = f"COUNT({ip_col})"                             if has_ip else "0"
    d_ip_expr      = f"COUNT(DISTINCT {ip_col})"                    if has_ip else "0"
    d_ipua_expr    = f"COUNT(DISTINCT ({ip_col}, {ua_col}))"        if (has_ip and has_ua) else "0"
    bytes_expr     = f"SUM(CAST(totalBytes AS BIGINT))"             if has_tb else "NULL"
    d_dev_expr     = f"COUNT(DISTINCT {qs_extract(qs_col,'device_id')})"   if has_qs else "0"
    d_sess_expr    = f"COUNT(DISTINCT {qs_extract(qs_col,'session_id')})"  if has_qs else "0"

    if has_qs:
        sess_extract  = qs_extract(qs_col, "session_id")
        sess_present  = f"SUM(CASE WHEN {qs_col} LIKE '%session_id=%' THEN 1 ELSE 0 END)"
        sess_blank    = (
            f"SUM(CASE WHEN {qs_col} LIKE '%session_id=%' "
            f"AND {sess_extract} IS NULL THEN 1 ELSE 0 END)"
        )
        sess_avail    = f"({sess_present} - {sess_blank})"
        sess_na       = sess_blank
        sess_none     = (
            f"SUM(CASE WHEN {qs_col} IS NULL "
            f"OR {qs_col} NOT LIKE '%session_id=%' THEN 1 ELSE 0 END)"
        )
    else:
        sess_avail = sess_na = sess_none = "0"

    reader = lake_reader(lake_root, year, month)
    df = safe_query(con, f"""
        SELECT
            make_date(CAST(year AS INT), CAST(month AS INT), CAST(day AS INT)) AS partition_date,
            COUNT(*)        AS total_rows,
            {ip_rows_expr}  AS ip_rows,
            {d_ip_expr}     AS distinct_ip,
            {d_ipua_expr}   AS distinct_ipua,
            {bytes_expr}    AS total_bytes,
            {d_dev_expr}    AS distinct_dev,
            {d_sess_expr}   AS distinct_sess,
            {sess_avail}    AS sess_avail,
            {sess_na}       AS sess_na,
            {sess_none}     AS sess_none
        FROM {reader}
        GROUP BY year, month, day
        ORDER BY partition_date
    """, "day breakdown", raise_on_error=True)

    if df.empty:
        return []

    new_rows: list[dict] = []
    for _, row in df.iterrows():
        ist_str = partition_date_to_ist_str(row["partition_date"])
        date_key = pd.to_datetime(row["partition_date"]).strftime("%Y-%m-%d")
        pa_count  = day_counts.get(date_key, int(row["total_rows"]))
        existing_count = existing_row_counts.get(ist_str)
        if existing_count == pa_count:
            log.info(f"  Skipping {ist_str} (already loaded, {pa_count:,} rows unchanged)")
            continue
        if existing_count is not None:
            log.info(f"  Refreshing {ist_str} (row count changed {existing_count:,} -> {pa_count:,})")
        bytes_gib = (
            float(row["total_bytes"]) / (1024 ** 3)
            if has_tb and pd.notna(row["total_bytes"]) else 0.0
        )
        new_rows.append({
            "ist_date":   ist_str,
            "bytes_gib":  round(bytes_gib, 2),
            "total_rows": pa_count,
            "ip_rows":    int(row["ip_rows"]),
            "dist_ip":    int(row["distinct_ip"]),
            "dist_ip_ua": int(row["distinct_ipua"]),
            "dist_dev":   int(row["distinct_dev"]),
            "dist_sess":  int(row["distinct_sess"]),
            "sess_avail": int(row["sess_avail"]),
            "sess_na":    int(row["sess_na"]),
            "sess_none":  int(row["sess_none"]),
        })
    return new_rows


def source_dates_from_lake(lake_root: Path, year: str | None, month: str | None) -> dict[str, set[str]]:
    source_dates: dict[str, set[str]] = {}
    roots = scoped_roots(lake_root, year, month) or [lake_root]
    for root in roots:
        for day_dir in root.rglob("day=*"):
            if not day_dir.is_dir() or not any(day_dir.glob("*.parquet")):
                continue
            parts = {p.split("=", 1)[0]: p.split("=", 1)[1] for p in day_dir.parts if "=" in p}
            if {"year", "month", "day"} - set(parts):
                continue
            source = parts.get("source", "stream").lower()
            try:
                date_key = f"{int(parts['year']):04d}-{int(parts['month']):02d}-{int(parts['day']):02d}"
            except ValueError:
                continue
            source_dates.setdefault(date_key, set()).add(source)
    return source_dates


def date_filter_sql(dates: set[str]) -> str:
    pieces = []
    for value in sorted(dates):
        try:
            yyyy, mm, dd = value.split("-", 2)
            pieces.append(f"(year = {int(yyyy)} AND month = '{int(mm):02d}' AND day = '{int(dd):02d}')")
        except ValueError:
            continue
    return " OR ".join(pieces) if pieces else "1=0"


def query_source_split_rows(
    con,
    lake_root: Path,
    year: str | None,
    month: str | None,
    cols: set[str],
    split_dates: set[str],
) -> list[dict]:
    if not split_dates or "reqTimeSec" not in cols:
        return []

    ip_col = "cliIP"
    ua_col = "UA"
    qs_col = "queryStr"
    has_ip = ip_col in cols
    has_ua = ua_col in cols
    has_qs = qs_col in cols
    has_tb = "totalBytes" in cols

    ip_rows_expr   = f"COUNT({ip_col})" if has_ip else "0"
    d_ip_expr      = f"COUNT(DISTINCT {ip_col})" if has_ip else "0"
    d_ipua_expr    = f"COUNT(DISTINCT ({ip_col}, {ua_col}))" if (has_ip and has_ua) else "0"
    bytes_expr     = f"SUM(CAST(totalBytes AS BIGINT))" if has_tb else "NULL"
    d_dev_expr     = f"COUNT(DISTINCT {qs_extract(qs_col,'device_id')})" if has_qs else "0"
    d_sess_expr    = f"COUNT(DISTINCT {qs_extract(qs_col,'session_id')})" if has_qs else "0"

    if has_qs:
        sess_extract = qs_extract(qs_col, "session_id")
        sess_present = f"SUM(CASE WHEN {qs_col} LIKE '%session_id=%' THEN 1 ELSE 0 END)"
        sess_blank = (
            f"SUM(CASE WHEN {qs_col} LIKE '%session_id=%' "
            f"AND {sess_extract} IS NULL THEN 1 ELSE 0 END)"
        )
        sess_avail = f"({sess_present} - {sess_blank})"
        sess_na = sess_blank
        sess_none = (
            f"SUM(CASE WHEN {qs_col} IS NULL "
            f"OR {qs_col} NOT LIKE '%session_id=%' THEN 1 ELSE 0 END)"
        )
    else:
        sess_avail = sess_na = sess_none = "0"

    reader = lake_reader(lake_root, year, month)
    df = safe_query(con, f"""
        SELECT
            make_date(CAST(year AS INT), CAST(month AS INT), CAST(day AS INT)) AS partition_date,
            lower(COALESCE(CAST(source AS VARCHAR), 'stream')) AS source,
            COUNT(*)        AS total_rows,
            {ip_rows_expr}  AS ip_rows,
            {d_ip_expr}     AS distinct_ip,
            {d_ipua_expr}   AS distinct_ipua,
            {bytes_expr}    AS total_bytes,
            {d_dev_expr}    AS distinct_dev,
            {d_sess_expr}   AS distinct_sess,
            {sess_avail}    AS sess_avail,
            {sess_na}       AS sess_na,
            {sess_none}     AS sess_none
        FROM {reader}
        WHERE {date_filter_sql(split_dates)}
        GROUP BY year, month, day, source
        ORDER BY partition_date, source
    """, "source day breakdown", raise_on_error=True)

    rows: list[dict] = []
    for _, row in df.iterrows():
        ip = int(row["ip_rows"])
        dist_ip = int(row["distinct_ip"])
        dist_ipua = int(row["distinct_ipua"])
        sa = int(row["sess_avail"])
        sna = int(row["sess_na"])
        sno = int(row["sess_none"])
        bytes_gib = (
            float(row["total_bytes"]) / (1024 ** 3)
            if has_tb and pd.notna(row["total_bytes"]) else 0.0
        )
        rows.append({
            "source":       str(row["source"] or "stream").lower(),
            "date":         pd.to_datetime(row["partition_date"]).strftime("%Y-%m-%d"),
            "bytes":        round(bytes_gib, 2),
            "rows":         int(row["total_rows"]),
            "ip_rows":      ip,
            "dist_ip":      dist_ip,
            "dist_ipua":    dist_ipua,
            "dist_dev":     int(row["distinct_dev"]),
            "dist_sess":    int(row["distinct_sess"]),
            "dist_ip_r2":   dist_ip,
            "dist_ipua_r2": dist_ipua,
            "sess_avail":   sa,
            "sess_na":      sna,
            "sess_none":    sno,
            "pct_ip":       round(dist_ip / ip, 6) if ip else 0,
            "pct_ipua":     round(dist_ipua / ip, 6) if ip else 0,
            "pct_sess":     round(sa / ip, 6) if ip else 0,
            "pct_sessna":   round(sna / ip, 6) if ip else 0,
            "pct_none":     round(sno / ip, 6) if ip else 0,
        })
    return rows


def overview_rows_to_source_seed(rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    for row in rows:
        if "date" in row:
            out.append(dict(row))
            continue
        ip_rows = int(row.get("ip_rows", 0) or 0)
        dist_ip = int(row.get("dist_ip", 0) or 0)
        dist_ipua = int(row.get("dist_ip_ua", row.get("dist_ipua", 0)) or 0)
        sess_avail = int(row.get("sess_avail", 0) or 0)
        sess_na = int(row.get("sess_na", 0) or 0)
        sess_none = int(row.get("sess_none", 0) or 0)
        out.append({
            "date": datetime.strptime(row["ist_date"], "%d/%m/%y").strftime("%Y-%m-%d"),
            "bytes": row.get("bytes_gib", row.get("bytes", 0)),
            "rows": row.get("total_rows", row.get("rows", 0)),
            "ip_rows": ip_rows,
            "dist_ip": dist_ip,
            "dist_ipua": dist_ipua,
            "dist_dev": row.get("dist_dev", 0),
            "dist_sess": row.get("dist_sess", 0),
            "dist_ip_r2": dist_ip,
            "dist_ipua_r2": dist_ipua,
            "sess_avail": sess_avail,
            "sess_na": sess_na,
            "sess_none": sess_none,
            "pct_ip": round(dist_ip / ip_rows, 6) if ip_rows else 0,
            "pct_ipua": round(dist_ipua / ip_rows, 6) if ip_rows else 0,
            "pct_sess": round(sess_avail / ip_rows, 6) if ip_rows else 0,
            "pct_sessna": round(sess_na / ip_rows, 6) if ip_rows else 0,
            "pct_none": round(sess_none / ip_rows, 6) if ip_rows else 0,
        })
    return out


def _overview_output_root(out_dir: Path) -> Path:
    candidates = [
        out_dir.parent,
        out_dir,
        Path(__file__).resolve().parents[2] / "output",
    ]
    for candidate in candidates:
        if (candidate / "watch_hours" / "profile" / "daily_volume.parquet").exists():
            return candidate
    return out_dir.parent


def _read_mart(path: Path, required: bool = False) -> pd.DataFrame:
    if not path.exists():
        if required:
            log.warning(f"Overview mart fallback missing required file: {path}")
        return pd.DataFrame()
    try:
        return pd.read_parquet(path)
    except Exception as exc:
        if required:
            log.warning(f"Overview mart fallback could not read required file {path}: {exc}")
        else:
            log.warning(f"Overview mart fallback skipped optional file {path}: {exc}")
        return pd.DataFrame()


def _filter_mart_dates(
    df: pd.DataFrame,
    date_col: str,
    year: str | None,
    month: str | None,
) -> pd.DataFrame:
    if df.empty or date_col not in df.columns:
        return pd.DataFrame()
    out = df.copy()
    out[date_col] = out[date_col].astype(str)
    if year:
        out = out[out[date_col].str.slice(0, 4) == year]
    if month:
        out = out[out[date_col].str.slice(5, 7) == month]
    return out


def _int_value(value, default: int = 0) -> int:
    try:
        if pd.isna(value):
            return default
        return int(value)
    except Exception:
        return default


def _float_value(value, default: float = 0.0) -> float:
    try:
        if pd.isna(value):
            return default
        return float(value)
    except Exception:
        return default


def _pct(part: int, whole: int) -> float:
    return round(part / whole, 6) if whole else 0.0


def _identity_lookup(identity: pd.DataFrame) -> tuple[dict[tuple[str, str], dict], dict[str, dict]]:
    empty: tuple[dict[tuple[str, str], dict], dict[str, dict]] = ({}, {})
    if identity.empty or {"log_date", "source"} - set(identity.columns):
        return empty

    work = identity.copy()
    work["log_date"] = work["log_date"].astype(str)
    work["source"] = work["source"].astype(str).str.lower()
    value_cols = [
        col
        for col in [
            "total_devices",
            "total_sessions",
            "total_ipua_sessions",
            "device_identity_rows",
            "session_identity_rows",
            "ipua_identity_rows",
        ]
        if col in work.columns
    ]
    if not value_cols:
        return empty

    by_source = (
        work.groupby(["log_date", "source"], as_index=False)[value_cols]
        .sum(numeric_only=True)
    )
    by_date = (
        work.groupby("log_date", as_index=False)[value_cols]
        .sum(numeric_only=True)
    )
    source_map = {
        (str(row["log_date"]), str(row["source"])): row.to_dict()
        for _, row in by_source.iterrows()
    }
    date_map = {
        str(row["log_date"]): row.to_dict()
        for _, row in by_date.iterrows()
    }
    return source_map, date_map


def _source_byte_lookup(latency_daily: pd.DataFrame) -> dict[tuple[str, str], float]:
    if latency_daily.empty or {"log_date", "source", "rows", "avg_total_bytes"} - set(latency_daily.columns):
        return {}
    work = latency_daily.copy()
    work["log_date"] = work["log_date"].astype(str)
    work["source"] = work["source"].astype(str).str.lower()
    work["byte_estimate"] = (
        pd.to_numeric(work["rows"], errors="coerce").fillna(0)
        * pd.to_numeric(work["avg_total_bytes"], errors="coerce").fillna(0)
    )
    grouped = work.groupby(["log_date", "source"], as_index=False)["byte_estimate"].sum()
    return {
        (str(row["log_date"]), str(row["source"])): float(row["byte_estimate"])
        for _, row in grouped.iterrows()
    }


def build_overview_from_marts(
    out_dir: Path,
    year: str | None,
    month: str | None,
) -> tuple[list[dict], list[dict]]:
    """Build Overview rows from compact marts when a direct lake scan fails."""
    output_root = _overview_output_root(out_dir)
    daily = _filter_mart_dates(
        _read_mart(output_root / "watch_hours" / "profile" / "daily_volume.parquet", required=True),
        "log_date",
        year,
        month,
    )
    source_daily = _filter_mart_dates(
        _read_mart(output_root / "watch_hours" / "daily_tables" / "daily_volume.parquet", required=True),
        "log_date",
        year,
        month,
    )
    if daily.empty or source_daily.empty:
        return [], []

    identity = _filter_mart_dates(
        _read_mart(output_root / "identity" / "identity_daily.parquet"),
        "log_date",
        year,
        month,
    )
    latency_daily = _filter_mart_dates(
        _read_mart(output_root / "latency" / "profile" / "daily.parquet"),
        "log_date",
        year,
        month,
    )
    identity_by_source, identity_by_date = _identity_lookup(identity)
    bytes_by_source = _source_byte_lookup(latency_daily)

    overview_rows: list[dict] = []
    daily = daily.sort_values("log_date")
    for _, row in daily.iterrows():
        date_key = str(row["log_date"])
        total_rows = _int_value(row.get("rows"))
        identity_row = identity_by_date.get(date_key, {})
        dist_ip = _int_value(row.get("approx_unique_ips"))
        dist_ipua = _int_value(identity_row.get("total_ipua_sessions"))
        dist_dev = _int_value(identity_row.get("total_devices"))
        dist_sess = _int_value(identity_row.get("total_sessions"))
        sess_avail = min(total_rows, _int_value(identity_row.get("session_identity_rows")))
        sess_none = max(0, total_rows - sess_avail)
        overview_rows.append({
            "ist_date": datetime.strptime(date_key, "%Y-%m-%d").strftime("%d/%m/%y"),
            "bytes_gib": round(_float_value(row.get("total_bytes")) / (1024 ** 3), 2),
            "total_rows": total_rows,
            "ip_rows": total_rows,
            "dist_ip": dist_ip,
            "dist_ip_ua": dist_ipua,
            "dist_dev": dist_dev,
            "dist_sess": dist_sess,
            "sess_avail": sess_avail,
            "sess_na": 0,
            "sess_none": sess_none,
        })

    source_rows: list[dict] = []
    source_daily = source_daily.sort_values(["log_date", "source"])
    for _, row in source_daily.iterrows():
        date_key = str(row["log_date"])
        source = str(row.get("source") or "stream").lower()
        total_rows = _int_value(row.get("rows"))
        identity_row = identity_by_source.get((date_key, source), {})
        dist_ip = _int_value(row.get("approx_unique_ips"))
        dist_ipua = _int_value(identity_row.get("total_ipua_sessions"))
        dist_dev = _int_value(identity_row.get("total_devices"))
        dist_sess = _int_value(identity_row.get("total_sessions"))
        sess_avail = min(total_rows, _int_value(identity_row.get("session_identity_rows")))
        sess_none = max(0, total_rows - sess_avail)
        source_rows.append({
            "source": source,
            "date": date_key,
            "bytes": round(bytes_by_source.get((date_key, source), 0.0) / (1024 ** 3), 2),
            "rows": total_rows,
            "ip_rows": total_rows,
            "dist_ip": dist_ip,
            "dist_ipua": dist_ipua,
            "dist_dev": dist_dev,
            "dist_sess": dist_sess,
            "dist_ip_r2": dist_ip,
            "dist_ipua_r2": dist_ipua,
            "sess_avail": sess_avail,
            "sess_na": 0,
            "sess_none": sess_none,
            "pct_ip": _pct(dist_ip, total_rows),
            "pct_ipua": _pct(dist_ipua, total_rows),
            "pct_sess": _pct(sess_avail, total_rows),
            "pct_sessna": 0.0,
            "pct_none": _pct(sess_none, total_rows),
        })

    return overview_rows, source_rows


def mart_date_ranges(data_rows: list[dict]) -> tuple[str, str]:
    if not data_rows:
        return "N/A", "N/A"
    dates = sorted(datetime.strptime(row["ist_date"], "%d/%m/%y") for row in data_rows)
    start = dates[0]
    end = dates[-1]
    return (
        f"{start.strftime('%Y-%m-%d')} -> {end.strftime('%Y-%m-%d')}",
        f"{start.strftime('%Y-%m-%d')} 00:00:00 -> {end.strftime('%Y-%m-%d')} 23:59:59",
    )


def write_source_rows_csv(output_path: Path, source_rows: list[dict]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(source_rows).to_csv(output_path, index=False)
    log.info(f"Wrote source daily rows: {output_path} ({len(source_rows):,} rows)")


def write_source_daily_csv(
    output_path: Path,
    all_rows: list[dict],
    lake_root: Path,
    year: str | None,
    month: str | None,
    cols: set[str],
) -> None:
    _, mart_source_rows = build_overview_from_marts(output_path.parent, year, month)
    if mart_source_rows:
        # Source-aware dashboard filtering must come from the full processed mart.
        # The active lake may only contain recent days when historical lake
        # partitions are archived to another drive, so lake-only source splitting
        # would silently drop older FAST rows.
        write_source_rows_csv(output_path, mart_source_rows)
        log.info("Source daily rows came from compact watch-hours marts.")
        return

    source_dates = source_dates_from_lake(lake_root, year, month)
    split_dates = {
        date_key
        for date_key, sources in source_dates.items()
        if "fast" in sources or len(sources) > 1
    }
    con = get_conn()
    try:
        split_rows = query_source_split_rows(con, lake_root, year, month, cols, split_dates)
    finally:
        con.close()

    split_by_date = {row["date"] for row in split_rows}
    source_rows: list[dict] = []
    for row in all_rows:
        if row["date"] in split_by_date:
            continue
        item = dict(row)
        item["source"] = "stream"
        source_rows.append(item)
    source_rows.extend(split_rows)
    source_rows.sort(key=lambda r: (r["date"], r.get("source", "stream")))

    write_source_rows_csv(output_path, source_rows)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SECTION E â€“ openpyxl style helpers
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(color: str = C_BLACK, bold: bool = False, size: int = 10) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)


def _align(h: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write(
    ws,
    row: int,
    col: int,
    value=None,
    fill_hex: str = None,
    font_color: str = C_BLACK,
    bold: bool = False,
    h_align: str = "left",
    num_fmt: str = None,
    wrap: bool = False,
):
    c = ws.cell(row=row, column=col, value=value)
    if fill_hex:
        c.fill = _fill(fill_hex)
    c.font      = _font(font_color, bold)
    c.alignment = _align(h_align, wrap)
    c.border    = _thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SECTION F â€“ Excel writer
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def write_overview_excel(
    output_path:      Path,
    data_rows:        list[dict],
    lake_root:        Path,
    year_filter:      str | None,
    month_filter:     str | None,
    total_rows_pa:    int,
    total_files:      int,
    date_range_ist:   str,
    time_range_ist:   str,
) -> None:
    """Write (or rewrite) the Overview Excel file with exact target formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = OVERVIEW_SHEET

    # â”€â”€ Column widths â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws.column_dimensions["A"].width = 30     # % Drop label
    ws.column_dimensions["B"].width = 14     # Date
    ws.column_dimensions["C"].width = 14     # Bytes
    ws.column_dimensions[get_column_letter(COL_ROWS)].width       = 14
    ws.column_dimensions[get_column_letter(COL_IP_ROWS)].width    = 13
    ws.column_dimensions[get_column_letter(COL_DIST_IP)].width    = 13
    ws.column_dimensions[get_column_letter(COL_DIST_IP_UA)].width = 16
    ws.column_dimensions[get_column_letter(COL_DIST_DEV)].width   = 14
    ws.column_dimensions[get_column_letter(COL_DIST_SESS)].width  = 14
    ws.column_dimensions[get_column_letter(COL_UNIQ_IP)].width    = 13
    ws.column_dimensions[get_column_letter(COL_UNIQ_IP_UA)].width = 16
    ws.column_dimensions[get_column_letter(COL_SESS_AVAIL)].width = 14
    ws.column_dimensions[get_column_letter(COL_SESS_NA)].width    = 16
    ws.column_dimensions[get_column_letter(COL_SESS_NONE)].width  = 14
    ws.column_dimensions[get_column_letter(COL_PCT_IP)].width     = 13
    ws.column_dimensions[get_column_letter(COL_PCT_IP_UA)].width  = 16
    ws.column_dimensions[get_column_letter(COL_PCT_SESS)].width   = 14
    ws.column_dimensions[get_column_letter(COL_PCT_SESS_NA)].width= 16
    ws.column_dimensions[get_column_letter(COL_PCT_NONE)].width   = 14

    # Row heights
    ws.row_dimensions[ROW_SECTION_HDR].height = 22
    ws.row_dimensions[ROW_COL_HDR].height     = 42

    # â”€â”€ Row 1: "ðŸ“Š Report Metadata" title â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws.merge_cells(start_row=ROW_TITLE_META, start_column=2,
                   end_row=ROW_TITLE_META, end_column=3)
    _write(ws, ROW_TITLE_META, 2, "ðŸ“Š  Report Metadata",
           fill_hex=C_DARK_BLUE, font_color=C_WHITE, bold=True)

    # â”€â”€ Rows 2â€“9: key/value metadata â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ist_now = datetime.now(timezone.utc) + IST_OFFSET
    kv_pairs = [
        ("Report generated",      ist_now.strftime("%Y-%m-%d %H:%M IST")),
        ("Filtered to",           f"year={year_filter}, month={month_filter}"
                                  if year_filter else "full lake"),
        ("Lake folder",           str(lake_root)),
        ("DuckDB version",        duckdb.__version__),
        ("Total rows (metadata)", f"{total_rows_pa:,}"),
        ("Total parquet files",   f"{total_files:,}"),
        ("Data date range (IST)", date_range_ist),
        ("Data time range (IST)", time_range_ist),
    ]
    for i, (key, val) in enumerate(kv_pairs):
        r = ROW_META_START + i
        _write(ws, r, 2, key, fill_hex=C_KV_BG, font_color=C_BLACK, bold=True)
        _write(ws, r, 3, val, fill_hex=C_WHITE,  font_color=C_BLACK, bold=False)

    # â”€â”€ Row 11: colored section-group headers (merged) â”€â”€â”€â”€â”€â”€â”€â”€
    section_groups = [
        (2,  5,  "ðŸ“…  Day-by-Day Breakdown (IST dates)", C_DARK_BLUE),
        (6,  9,  "UNIQUE SET",                           C_GREEN),
        (10, 14, "ROWS AGAINST - UNIQUE SET",            C_AMBER),
        (15, 19, "% Data of TOTAL",                      C_PURPLE),
    ]
    for start_c, end_c, label, color in section_groups:
        ws.merge_cells(start_row=ROW_SECTION_HDR, start_column=start_c,
                   end_row=ROW_SECTION_HDR, end_column=end_c)
    # Label on the first cell only
    _write(ws, ROW_SECTION_HDR, start_c, label,
           fill_hex=color, font_color=C_WHITE, bold=True, h_align="center")
    # Other cells in the merged range: fill and border only, no value
    for c in range(start_c + 1, end_c + 1):
        cell = ws.cell(row=ROW_SECTION_HDR, column=c)
        cell.fill = _fill(color)
        cell.border = _thin_border()           
        cell.alignment = _align("center")

    # â”€â”€ Row 12: column labels (match section colors) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    col_labels = [
        (COL_DATE,       "Date (IST)",                      C_DARK_BLUE),
        (COL_BYTES,      "Total Bytes (GiB)",                C_DARK_BLUE),
        (COL_ROWS,       "Total Rows*",                      C_DARK_BLUE),
        (COL_IP_ROWS,    "cliIP rows",                       C_DARK_BLUE),
        (COL_DIST_IP,    "Distinct cliIP",                   C_GREEN),
        (COL_DIST_IP_UA, "Distinct cliIP & UA",              C_GREEN),
        (COL_DIST_DEV,   "Distinct device_id",               C_GREEN),
        (COL_DIST_SESS,  "Distinct session_id",              C_GREEN),
        (COL_UNIQ_IP,    "Distinct cliIP",                   C_AMBER),
        (COL_UNIQ_IP_UA, "Distinct cliIP & UA",              C_AMBER),
        (COL_SESS_AVAIL, "Session Found - ID Available",     C_AMBER),
        (COL_SESS_NA,    "Session Found - ID NOT Available", C_AMBER),
        (COL_SESS_NONE,  "Session Not Found",                C_AMBER),
        (COL_PCT_IP,     "Distinct cliIP",                   C_PURPLE),
        (COL_PCT_IP_UA,  "Distinct cliIP & UA",              C_PURPLE),
        (COL_PCT_SESS,   "Session Found - ID Available",     C_PURPLE),
        (COL_PCT_SESS_NA,"Session Found - ID NOT Available", C_PURPLE),
        (COL_PCT_NONE,   "Session Not Found",                C_PURPLE),
    ]
    for col_idx, label, color in col_labels:
        _write(ws, ROW_COL_HDR, col_idx, label,
               fill_hex=color, font_color=C_WHITE, bold=True,
               h_align="center", wrap=True)

    # â”€â”€ Rows 13+: data rows â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    e_ltr = get_column_letter(COL_IP_ROWS)   # E
    j_ltr = get_column_letter(COL_UNIQ_IP)   # J
    k_ltr = get_column_letter(COL_UNIQ_IP_UA) # K
    l_ltr = get_column_letter(COL_SESS_AVAIL) # L
    m_ltr = get_column_letter(COL_SESS_NA)    # M
    n_ltr = get_column_letter(COL_SESS_NONE)  # N

    r = DATA_START_ROW
    for i, d in enumerate(data_rows):
        bg = C_ALT if i % 2 == 1 else C_WHITE

        _write(ws, r, COL_DATE,       d["ist_date"],   bg, C_BLACK, h_align="center")
        _write(ws, r, COL_BYTES,      d["bytes_gib"],  bg, C_BLACK, h_align="right", num_fmt="#,##0.00")
        _write(ws, r, COL_ROWS,       d["total_rows"], bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_IP_ROWS,    d["ip_rows"],    bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_IP,    d["dist_ip"],    bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_IP_UA, d["dist_ip_ua"],bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_DEV,   d["dist_dev"],   bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_SESS,  d["dist_sess"],  bg, C_BLACK, h_align="right", num_fmt="#,##0")
        # UNIQUE SET â€“ mirrors cliIP rows (col E)
        _write(ws, r, COL_UNIQ_IP,    f"={e_ltr}{r}",  bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_UNIQ_IP_UA, f"={e_ltr}{r}",  bg, C_BLACK, h_align="right", num_fmt="#,##0")
        # ROWS AGAINST
        _write(ws, r, COL_SESS_AVAIL, d["sess_avail"], bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_SESS_NA,    d["sess_na"],     bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_SESS_NONE,  d["sess_none"],   bg, C_BLACK, h_align="right", num_fmt="#,##0")
        # % Data of TOTAL â€“ formula /{E locked}
        cur_e = f"${e_ltr}{r}"
        _write(ws, r, COL_PCT_IP,     f"={j_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_IP_UA,  f"={k_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_SESS,   f"={l_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_SESS_NA,f"={m_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_NONE,   f"={n_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        r += 1

    data_end_row = r - 1   # last data row (1-indexed)
    total_row    = r        # TOTAL is immediately after last data row

    # â”€â”€ TOTAL row (orange) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    _write(ws, total_row, COL_DATE, "TOTAL*",
           fill_hex=C_ORANGE, font_color=C_WHITE, bold=True, h_align="left")
    sum_cols = [
        COL_BYTES, COL_ROWS, COL_IP_ROWS,
        COL_DIST_IP, COL_DIST_IP_UA, COL_DIST_DEV, COL_DIST_SESS,
        COL_UNIQ_IP, COL_UNIQ_IP_UA,
        COL_SESS_AVAIL, COL_SESS_NA, COL_SESS_NONE,
    ]
    for col_idx in sum_cols:
        ltr = get_column_letter(col_idx)
        nfmt = "#,##0.00" if col_idx == COL_BYTES else "#,##0"
        _write(ws, total_row, col_idx,
               f"=SUM({ltr}{DATA_START_ROW}:{ltr}{data_end_row})",
               fill_hex=C_ORANGE, font_color=C_WHITE, bold=True,
               h_align="right", num_fmt=nfmt)

    # â”€â”€ % Drop/Gain row (2 rows below TOTAL) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    drop_row = total_row + 2
    ws.merge_cells(start_row=drop_row, start_column=1,
                   end_row=drop_row, end_column=2)
    _write(ws, drop_row, 1, "% Drop/Gain w.r.t. last 5 days",
           font_color=C_BLACK, bold=True, h_align="left")

    # Need at least 6 data rows to build a meaningful drop formula
    # Compares last-but-one day (most likely "complete") to avg of 4 days before it
    n_rows = data_end_row - DATA_START_ROW + 1
    if n_rows >= 6:
        last_full = data_end_row - 1
        avg_start = data_end_row - 5
        avg_end   = data_end_row - 2
        drop_cols = [
            COL_BYTES, COL_ROWS, COL_IP_ROWS,
            COL_DIST_IP, COL_DIST_IP_UA, COL_DIST_DEV, COL_DIST_SESS,
            COL_UNIQ_IP, COL_UNIQ_IP_UA,
            COL_SESS_AVAIL, COL_SESS_NA, COL_SESS_NONE,
            COL_PCT_SESS, COL_PCT_SESS_NA, COL_PCT_NONE,
        ]
        for col_idx in drop_cols:
            ltr = get_column_letter(col_idx)
            _write(ws, drop_row, col_idx,
                   f"=({ltr}{last_full}/AVERAGE({ltr}{avg_start}:{ltr}{avg_end}))-1",
                   font_color=C_BLACK, h_align="right", num_fmt="0.00%")

    # â”€â”€ Footnote â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    fn_row = drop_row + 2
    c = ws.cell(row=fn_row, column=COL_DATE,
                value="* Total Rows sourced from Parquet file metadata (zero-scan). "
                      "All other columns via DuckDB aggregation.")
    c.font = _font(C_BLACK, bold=False)

    # â”€â”€ Freeze top rows and date column â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=COL_DATE + 1)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SECTION G â€“ Main run
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def run(
    lake_root: Path,
    out_dir: Path | None = None,
    year_pref: str | None = None,
    month_pref: str | None = None,
    auto_confirm: bool = False,
) -> None:
    print("\n" + "=" * 60)
    print("  Overview Generator - append-aware")
    print("=" * 60 + "\n")

    # â”€â”€ Output directory â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if out_dir is None:
        print("  Output folder (Excel will be saved / updated here):")
        out_dir = prompt_path("  Output folder path: ")
    else:
        out_dir = out_dir.expanduser().resolve()
        if not out_dir.exists():
            out_dir.mkdir(parents=True, exist_ok=True)
    output_file = out_dir / OVERVIEW_FILENAME
    is_first    = not output_file.exists()

    existing_dates = get_existing_dates(output_file)
    existing_row_counts = get_existing_row_counts(output_file)
    if is_first:
        print("\n  No existing file - will create fresh.")
    else:
        sorted_d = sorted(existing_dates)
        print(f"\n  Found existing file with {len(existing_dates)} date(s) loaded.")
        if sorted_d:
            print(f"      Earliest: {sorted_d[0]}   Latest: {sorted_d[-1]}")

    # â”€â”€ Date filter â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    year_filter, month_filter = prompt_year_month(lake_root, year_pref, month_pref)

    # â”€â”€ Confirm â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    print("\n" + "-" * 60)
    print(f"  Lake        : {lake_root}")
    print(f"  Output file : {output_file}")
    print(f"  Filter      : " + (
        f"year={year_filter}" + (f", month={month_filter}" if month_filter else "")
        if year_filter else "none (full lake)"
    ))
    print(f"  Mode        : {'CREATE (first run)' if is_first else 'APPEND/REFRESH (new or changed dates)'}")
    print("-" * 60)
    if not auto_confirm:
        if input("\n  Proceed? (y/n): ").strip().lower() != "y":
            print("  Cancelled.")
            return

    # â”€â”€ Parquet metadata (zero-scan) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    log.info("Reading Parquet metadata (zero-scan) ...")
    day_counts  = pa_row_counts_by_day(lake_root, year_filter, month_filter)
    total_rows  = sum(day_counts.values())
    total_files = pa_total_file_count_scoped(lake_root, year_filter, month_filter)
    log.info(f"  {total_rows:,} rows across {total_files:,} parquet files")

    # â”€â”€ Schema detection â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    log.info("Detecting schema ...")
    cols = pa_schema(lake_root, year_filter, month_filter)
    log.info(f"  {len(cols)} columns detected")

    # â”€â”€ DuckDB connection â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    log.info("Connecting to DuckDB ...")
    con = get_conn()

    # â”€â”€ Date range metadata â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    date_range_ist = time_range_ist = "N/A"
    ts = "reqTimeSec"
    if ts in cols:
        reader = lake_reader(lake_root, year_filter, month_filter)
        ist_ts = ist_timestamp_sql(ts)
        meta   = safe_query(con, f"""
            SELECT
                MIN(CAST({ist_ts} AS DATE)) AS min_d,
                MAX(CAST({ist_ts} AS DATE)) AS max_d,
                MIN({ist_ts})               AS min_ts,
                MAX({ist_ts})               AS max_ts
            FROM {reader}
        """, "date range")
        if not meta.empty:
            row = meta.iloc[0]
            date_range_ist = (
                f"{row['min_d'].strftime('%Y-%m-%d')} -> "
                f"{row['max_d'].strftime('%Y-%m-%d')}"
            )
            time_range_ist = (
                f"{row['min_ts'].strftime('%Y-%m-%d %H:%M:%S')} -> "
                f"{row['max_ts'].strftime('%Y-%m-%d %H:%M:%S')}"
            )

    # â”€â”€ Query new dates â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    log.info("Querying day-by-day breakdown â€¦")
    query_failed = False
    try:
        new_rows = query_new_dates(
            con, lake_root, year_filter, month_filter,
            cols, existing_row_counts, day_counts,
        )
    except RuntimeError as exc:
        query_failed = True
        new_rows = []
        log.warning(f"Day breakdown query unavailable; trying mart fallback. {exc}")
    finally:
        con.close()

    if query_failed:
        fallback_rows, fallback_source_rows = build_overview_from_marts(
            out_dir,
            year_filter,
            month_filter,
        )
        if not fallback_rows:
            raise SystemExit(
                "Overview day breakdown failed and mart fallback was unavailable. "
                "The pipeline stopped so stale Overview output is not mistaken for fresh output."
            )

        if date_range_ist == "N/A" or time_range_ist == "N/A":
            date_range_ist, time_range_ist = mart_date_ranges(fallback_rows)

        log.info(f"Writing {len(fallback_rows)} Overview rows from compact marts ...")
        write_source_rows_csv(out_dir / SOURCE_DAILY_FILENAME, fallback_source_rows)
        write_overview_excel(
            output_path    = output_file,
            data_rows      = fallback_rows,
            lake_root      = lake_root,
            year_filter    = year_filter,
            month_filter   = month_filter,
            total_rows_pa  = total_rows,
            total_files    = total_files,
            date_range_ist = date_range_ist,
            time_range_ist = time_range_ist,
        )

        print(f"\n{'=' * 60}")
        print(f"  Overview saved from compact marts: {output_file}")
        print(f"{'=' * 60}\n")
        return

    if not new_rows:
        source_csv = out_dir / SOURCE_DAILY_FILENAME
        if not source_csv.exists():
            log.info("Source daily CSV missing; creating it from current Overview rows.")
            current_rows = load_existing_rows(output_file)
            source_rows_for_csv = overview_rows_to_source_seed(current_rows)
            write_source_daily_csv(source_csv, source_rows_for_csv, lake_root, year_filter, month_filter, cols)
        print("\n  Nothing new to add - your file is already up to date.")
        return

    unchanged = sum(1 for date, count in existing_row_counts.items() if day_counts.get(datetime.strptime(date, "%d/%m/%y").strftime("%Y-%m-%d")) == count)
    print(f"\n  {len(new_rows)} date(s) to add/refresh."
          + (f" ({unchanged} existing date(s) unchanged)" if unchanged else ""))

    # â”€â”€ Merge with existing â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    if is_first:
        all_rows = new_rows
    else:
        log.info("Loading existing data rows ...")
        old_rows = load_existing_rows(output_file)
        # Remove overlap (safety) then combine
        new_dates_set = {r["ist_date"] for r in new_rows}
        old_rows = [r for r in old_rows if r["ist_date"] not in new_dates_set]
        all_rows = old_rows + new_rows
        all_rows.sort(key=lambda x: datetime.strptime(x["ist_date"], "%d/%m/%y"))

    source_rows_for_csv = overview_rows_to_source_seed(all_rows)
    write_source_daily_csv(
        out_dir / SOURCE_DAILY_FILENAME,
        source_rows_for_csv,
        lake_root,
        year_filter,
        month_filter,
        cols,
    )

    # â”€â”€ Write Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    log.info(f"Writing {len(all_rows)} total rows to Excel ...")
    write_overview_excel(
        output_path    = output_file,
        data_rows      = all_rows,
        lake_root      = lake_root,
        year_filter    = year_filter,
        month_filter   = month_filter,
        total_rows_pa  = total_rows,
        total_files    = total_files,
        date_range_ist = date_range_ist,
        time_range_ist = time_range_ist,
    )

    print(f"\n{'=' * 60}")
    print(f"  Overview saved: {output_file}")
    print(f"{'=' * 60}\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Overview report generator")
    parser.add_argument("lake_root", nargs="?", default=None, help="Lake folder path (year=/month=/day=/... partitions)")
    parser.add_argument("--lake-root", dest="lake_root_opt", default=None)
    parser.add_argument("--out-dir", default=None, help="Folder to write overview_report.xlsx")
    parser.add_argument("--year", default=None, help="Filter YYYY (e.g. 2026)")
    parser.add_argument("--month", default=None, help="Filter MM (01-12)")
    parser.add_argument("--yes", action="store_true", help="Run without interactive confirmation")
    parser.add_argument(
        "--auto",
        action="store_true",
        help="Compatibility alias for --yes",
    )

    args = parser.parse_args()
    lake_root = Path(args.lake_root_opt or args.lake_root or "").expanduser() if (args.lake_root_opt or args.lake_root) else None

    if lake_root is None or not lake_root.is_dir():
        env_lake = os.getenv("VG_DASH_LAKE_ROOT")
        if env_lake:
            lake_root = Path(env_lake).expanduser()
        else:
            etl_root = Path(__file__).resolve().parents[2]
            candidate = Path(
                os.getenv(
                    "VG_ETL_BASE",
                    str(etl_root / "data"),
                )
            ).expanduser()
            lake_root = candidate / "lake"

    if not lake_root.is_dir():
        print("\nðŸ“  Lake folder â€“ must contain year=* partitions")
        lake_root = prompt_path("  Lake folder path: ")

    if args.out_dir:
        out_dir = Path(args.out_dir).expanduser().resolve()
    else:
        out_dir = Path(
            os.getenv(
                "VG_DASH_OVERVIEW_BASE",
                str(Path(__file__).resolve().parent),
            )
        ).resolve()

    run(
        lake_root=lake_root,
        out_dir=out_dir,
        year_pref=args.year,
        month_pref=args.month,
        auto_confirm=args.yes or args.auto,
    )


if __name__ == "__main__":
    main()


