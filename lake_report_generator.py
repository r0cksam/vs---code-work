#!/usr/bin/env python3
"""
lake_report_generator.py
═══════════════════════════════════════════════════════════════════════
Generates a comprehensive Excel report from a Hive-partitioned parquet
data lake (year= / month= / day= layout created by the lake pipeline).

Usage:
    python lake_report_generator.py
      ─ or ─
    python lake_report_generator.py "Z:\\05 Veto Logs\\lake"

Sheets generated:
  1. Overview       — metadata + day-by-day row & coverage breakdown
  2. Distinct Counts — field-level distinct count summary
  3. UA             — distinct user-agent strings with counts
  4. ASN            — distinct ASN values
  5. City           — distinct cities
  6. State          — distinct states / regions
  7. Country        — distinct countries
  8. reqHost        — distinct request hosts
  9. Channels       — channels parsed from queryString
 10. Platforms      — platforms parsed from queryString
 11. Devices        — device names parsed from queryString
 12. Categories     — category names parsed from queryString
 13. Content        — top content titles from queryString
 14. Channel×Platform — channel broken down by platform (pivot)
"""

import sys
import json
import time
from pathlib import Path
from datetime import datetime, timezone

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ═══════════════════════════════════════════════════════════════
# CONFIG — edit these if not passing a CLI argument
# ═══════════════════════════════════════════════════════════════

LAKE_FOLDER   = Path(r"Z:\05 Veto Logs\lake")   # override via CLI arg
OUTPUT_FOLDER = LAKE_FOLDER.parent               # where to save the .xlsx
THREADS       = 8
MEMORY        = "16GB"
TOP_N         = 5000        # max distinct values per sheet
QS_COL        = "queryStr"  # query-string column name in your lake
TS_COL        = "reqTimeSec"

# Column names in your lake (set to None / "" if absent)
FIELD_MAP = {
    "ua":       "UA",
    "asn":      "asn",
    "city":     "city",
    "state":    "state",
    "country":  "country",
    "reqhost":  "reqHost",
    "qs":       "queryStr",
    "ip":       "cliIP",
    "path":     "reqPath",
    "ts":       "reqTimeSec",
}

# ═══════════════════════════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════════════════════════

C_DARK_BLUE  = "1F3864"
C_MID_BLUE   = "2E75B6"
C_LIGHT_BLUE = "D6E4F0"
C_WHITE      = "FFFFFF"
C_LIGHT_GRAY = "F2F2F2"
C_ORANGE     = "E36209"
C_GREEN      = "1E7E34"

def _hdr_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _hdr_font(hex_color=C_WHITE, bold=True, size=10):
    return Font(name="Arial", bold=bold, color=hex_color, size=size)

def _body_font(bold=False, size=10):
    return Font(name="Arial", bold=bold, size=size, color="000000")

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        w = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        ws.column_dimensions[col_letter].width = max(min_w, min(w + 3, max_w))

def _write_header_row(ws, row_num, headers, bg=C_MID_BLUE, fg=C_WHITE):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.font      = _hdr_font(fg, bold=True, size=10)
        cell.fill      = _hdr_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin_border()

def _write_data_row(ws, row_num, values, zebra=False):
    bg = C_LIGHT_GRAY if zebra else C_WHITE
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=v)
        cell.font      = _body_font()
        cell.fill      = _hdr_fill(bg)
        cell.border    = _thin_border()
        if isinstance(v, (int, float)):
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="left")

def _section_title(ws, row_num, text, n_cols=2):
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font      = _hdr_font(C_WHITE, bold=True, size=11)
    cell.fill      = _hdr_fill(C_DARK_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if n_cols > 1:
        ws.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num,   end_column=n_cols
        )

def _kv_row(ws, row_num, key, value, bold_key=True):
    k = ws.cell(row=row_num, column=1, value=key)
    v = ws.cell(row=row_num, column=2, value=value)
    k.font      = Font(name="Arial", bold=bold_key, size=10)
    v.font      = _body_font(size=10)
    k.fill      = _hdr_fill(C_LIGHT_BLUE)
    v.fill      = _hdr_fill(C_WHITE)
    k.border    = _thin_border()
    v.border    = _thin_border()
    k.alignment = Alignment(horizontal="left")
    v.alignment = Alignment(horizontal="left")

def _freeze_and_filter(ws, freeze_at="A2"):
    ws.freeze_panes = freeze_at

def _add_pct_bar(ws, col_letter, data_start, data_end):
    """Add Excel data bars to a percentage column."""
    from openpyxl.formatting.rule import DataBarRule
    rule = DataBarRule(
        start_type="num", start_value=0,
        end_type="num",   end_value=100,
        color="2E75B6",
    )
    ws.conditional_formatting.add(f"{col_letter}{data_start}:{col_letter}{data_end}", rule)


# ═══════════════════════════════════════════════════════════════
# DuckDB helpers
# ═══════════════════════════════════════════════════════════════

def get_conn(lake_root: Path):
    con = duckdb.connect()
    con.execute(f"SET threads={THREADS};")
    con.execute(f"SET memory_limit='{MEMORY}';")
    con.execute("SET preserve_insertion_order=false;")
    return con


def lake_reader(lake_root: Path) -> str:
    """SQL expression to read the entire Hive lake with partition pruning."""
    return f"read_parquet('{lake_root.as_posix()}/**/*.parquet', hive_partitioning=true, union_by_name=true)"


def detect_columns(con, lake_root: Path) -> set:
    """Return the set of column names actually present in the lake."""
    try:
        df = con.execute(
            f"DESCRIBE SELECT * FROM {lake_reader(lake_root)} LIMIT 0"
        ).df()
        return set(df["column_name"].tolist())
    except Exception:
        return set()


def col(name: str, existing: set, fallback="NULL") -> str:
    """Return the column reference if it exists, else a NULL placeholder."""
    if name and name in existing:
        return name
    return f"{fallback} AS {name}" if fallback != "NULL" else "NULL"


def qs_extract(qs_col: str, param: str) -> str:
    return f"NULLIF(regexp_extract({qs_col}, '(?:^|&){param}=([^&]+)', 1), '')"


def safe_df(con, sql: str, label: str = "") -> pd.DataFrame:
    try:
        return con.execute(sql).df()
    except Exception as e:
        print(f"  ⚠️  Query failed{' (' + label + ')' if label else ''}: {e}")
        return pd.DataFrame()


# ═══════════════════════════════════════════════════════════════
# Sheet builders
# ═══════════════════════════════════════════════════════════════

def build_sheet_overview(ws, con, lake_root: Path, existing: set, generated_at: str):
    """Sheet 1 — metadata header + day-by-day coverage breakdown."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36

    reader = lake_reader(lake_root)
    fm     = FIELD_MAP
    ts     = fm["ts"] if fm["ts"] in existing else None

    # ── Global metadata ───────────────────────────────────────
    _section_title(ws, 1, "📊  Report Metadata", 2)
    _kv_row(ws, 2, "Report generated",  generated_at)
    _kv_row(ws, 3, "Lake folder",       str(lake_root))
    _kv_row(ws, 4, "DuckDB version",    duckdb.__version__)

    # Date + time range
    if ts:
        meta = safe_df(con, f"""
            SELECT
                MIN(to_timestamp(CAST({ts} AS DOUBLE)))::DATE   AS min_date,
                MAX(to_timestamp(CAST({ts} AS DOUBLE)))::DATE   AS max_date,
                MIN(to_timestamp(CAST({ts} AS DOUBLE)))         AS min_ts,
                MAX(to_timestamp(CAST({ts} AS DOUBLE)))         AS max_ts,
                COUNT(*)                                         AS total_rows
            FROM {reader}
        """, "global meta")
        if not meta.empty:
            r = meta.iloc[0]
            _kv_row(ws, 5, "Data date range",
                    f"{r['min_date']}  →  {r['max_date']}")
            _kv_row(ws, 6, "Data time range",
                    f"{str(r['min_ts'])[:19]}  →  {str(r['max_ts'])[:19]}")
            _kv_row(ws, 7, "Total rows (all time)", f"{int(r['total_rows']):,}")
    else:
        _kv_row(ws, 5, "Data date range", "reqTimeSec column not found")
        _kv_row(ws, 6, "Data time range", "—")
        _kv_row(ws, 7, "Total rows (all time)", "—")

    # Partition info
    year_dirs = sorted(
        [d.name for d in lake_root.iterdir() if d.is_dir() and d.name.startswith("year=")]
    )
    _kv_row(ws, 8,  "Years in lake", ", ".join(y.replace("year=","") for y in year_dirs) or "—")
    total_files = sum(1 for _ in lake_root.rglob("*.parquet"))
    _kv_row(ws, 9,  "Parquet files (total)", f"{total_files:,}")

    # ── Day-by-day breakdown ───────────────────────────────────
    ws.row_dimensions[11].height = 20
    _section_title(ws, 11, "📅  Day-by-Day Breakdown", 16)

    has_ua      = fm["ua"]      in existing
    has_asn     = fm["asn"]     in existing
    has_city    = fm["city"]    in existing
    has_state   = fm["state"]   in existing
    has_country = fm["country"] in existing
    has_host    = fm["reqhost"] in existing
    has_qs      = fm["qs"]      in existing

    qs = fm["qs"]

    day_headers = [
        "Date", "Total Rows",
        "UA rows",      "ASN rows",
        "City rows",    "State rows",    "Country rows",
        "reqHost rows", "QueryStr rows",
        "Rows w/ device_id", "Rows w/ session_id",
        "Rows w/ channel",   "Rows w/ category",
        "Distinct devices",  "Distinct sessions",
        "Distinct channels",
    ]
    _write_header_row(ws, 12, day_headers, bg=C_MID_BLUE)

    if ts:
        ua_sel      = f"COUNT({fm['ua']})"       if has_ua      else "0"
        asn_sel     = f"COUNT({fm['asn']})"      if has_asn     else "0"
        city_sel    = f"COUNT({fm['city']})"     if has_city    else "0"
        state_sel   = f"COUNT({fm['state']})"    if has_state   else "0"
        country_sel = f"COUNT({fm['country']})"  if has_country else "0"
        host_sel    = f"COUNT({fm['reqhost']})"  if has_host    else "0"
        qs_sel      = f"COUNT({fm['qs']})"       if has_qs      else "0"

        dev_sel   = f"SUM(CASE WHEN {qs} LIKE '%device_id=%'   THEN 1 ELSE 0 END)" if has_qs else "0"
        sess_sel  = f"SUM(CASE WHEN {qs} LIKE '%session_id=%'  THEN 1 ELSE 0 END)" if has_qs else "0"
        chan_sel  = f"SUM(CASE WHEN {qs} LIKE '%channel=%'      THEN 1 ELSE 0 END)" if has_qs else "0"
        cat_sel   = f"SUM(CASE WHEN {qs} LIKE '%category_name=%' THEN 1 ELSE 0 END)" if has_qs else "0"
        ddev_sel  = f"COUNT(DISTINCT {qs_extract(qs,'device_id')})" if has_qs else "0"
        dsess_sel = f"COUNT(DISTINCT {qs_extract(qs,'session_id')})" if has_qs else "0"
        dchan_sel = f"COUNT(DISTINCT {qs_extract(qs,'channel')})" if has_qs else "0"

        day_df = safe_df(con, f"""
            SELECT
                make_date(CAST(year AS INT), CAST(month AS INT), CAST(day AS INT)) AS date,
                COUNT(*)                AS total_rows,
                {ua_sel}                AS ua_rows,
                {asn_sel}               AS asn_rows,
                {city_sel}              AS city_rows,
                {state_sel}             AS state_rows,
                {country_sel}           AS country_rows,
                {host_sel}              AS reqhost_rows,
                {qs_sel}                AS qs_rows,
                {dev_sel}               AS device_rows,
                {sess_sel}              AS session_rows,
                {chan_sel}              AS channel_rows,
                {cat_sel}               AS category_rows,
                {ddev_sel}              AS distinct_devices,
                {dsess_sel}             AS distinct_sessions,
                {dchan_sel}             AS distinct_channels
            FROM {reader}
            GROUP BY year, month, day
            ORDER BY year, month, day
        """, "day breakdown")

        for i, row_data in enumerate(day_df.itertuples(index=False), 13):
            dt = row_data.date
            date_str = dt.strftime("%d/%m/%y") if hasattr(dt, "strftime") else str(dt)
            values = [
                date_str,
                int(row_data.total_rows),
                int(row_data.ua_rows),
                int(row_data.asn_rows),
                int(row_data.city_rows),
                int(row_data.state_rows),
                int(row_data.country_rows),
                int(row_data.reqhost_rows),
                int(row_data.qs_rows),
                int(row_data.device_rows),
                int(row_data.session_rows),
                int(row_data.channel_rows),
                int(row_data.category_rows),
                int(row_data.distinct_devices),
                int(row_data.distinct_sessions),
                int(row_data.distinct_channels),
            ]
            _write_data_row(ws, i, values, zebra=(i % 2 == 0))

        # Totals row
        n_data = len(day_df)
        if n_data:
            total_row_num = 13 + n_data
            total_vals = ["TOTAL", f"=SUM(B13:B{total_row_num - 1})"] + [
                f"=SUM({get_column_letter(c)}13:{get_column_letter(c)}{total_row_num - 1})"
                for c in range(3, 17)
            ]
            for col_idx, v in enumerate(total_vals, 1):
                cell = ws.cell(row=total_row_num, column=col_idx, value=v)
                cell.font   = Font(name="Arial", bold=True, size=10, color=C_WHITE)
                cell.fill   = _hdr_fill(C_ORANGE)
                cell.border = _thin_border()
                cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left")

    ws.freeze_panes = "B13"
    _auto_width(ws, min_w=10)
    for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]:
        ws.column_dimensions[col_letter].width = 14


def build_sheet_distinct_counts(ws, con, lake_root: Path, existing: set):
    """Sheet 2 — distinct count per field."""
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, "📊  Distinct Value Counts per Field", 4)

    headers = ["Field", "Source Column", "Distinct Values", "Notes"]
    _write_header_row(ws, 2, headers)

    reader = lake_reader(lake_root)
    fm     = FIELD_MAP
    qs     = fm["qs"]

    fields = [
        ("User Agent (UA)",      fm["ua"],      fm["ua"] in existing, ""),
        ("ASN",                  fm["asn"],     fm["asn"] in existing, ""),
        ("City",                 fm["city"],    fm["city"] in existing, ""),
        ("State / Region",       fm["state"],   fm["state"] in existing, ""),
        ("Country",              fm["country"], fm["country"] in existing, ""),
        ("Request Host",         fm["reqhost"], fm["reqhost"] in existing, ""),
        ("Client IP",            fm["ip"],      fm["ip"] in existing, ""),
        ("Request Path",         fm["path"],    fm["path"] in existing, ""),
        ("Query String channel", "(queryStr)",  fm["qs"] in existing, "parsed via regex"),
        ("Query String platform","(queryStr)",  fm["qs"] in existing, "parsed via regex"),
        ("Query String device",  "(queryStr)",  fm["qs"] in existing, "parsed via regex"),
        ("Query String category","(queryStr)",  fm["qs"] in existing, "parsed via regex"),
        ("Query String content", "(queryStr)",  fm["qs"] in existing, "parsed via regex"),
        ("Query String device_id","(queryStr)", fm["qs"] in existing, "parsed via regex"),
        ("Query String session_id","(queryStr)",fm["qs"] in existing, "parsed via regex"),
    ]

    selects = []
    if fm["ua"]      in existing: selects.append(f"COUNT(DISTINCT {fm['ua']})      AS ua")
    if fm["asn"]     in existing: selects.append(f"COUNT(DISTINCT {fm['asn']})     AS asn")
    if fm["city"]    in existing: selects.append(f"COUNT(DISTINCT {fm['city']})    AS city")
    if fm["state"]   in existing: selects.append(f"COUNT(DISTINCT {fm['state']})   AS state")
    if fm["country"] in existing: selects.append(f"COUNT(DISTINCT {fm['country']}) AS country")
    if fm["reqhost"] in existing: selects.append(f"COUNT(DISTINCT {fm['reqhost']}) AS reqhost")
    if fm["ip"]      in existing: selects.append(f"COUNT(DISTINCT {fm['ip']})      AS ip")
    if fm["path"]    in existing: selects.append(f"COUNT(DISTINCT {fm['path']})    AS path")
    if fm["qs"]      in existing:
        selects += [
            f"COUNT(DISTINCT {qs_extract(qs,'channel')})        AS qs_channel",
            f"COUNT(DISTINCT {qs_extract(qs,'platform')})       AS qs_platform",
            f"COUNT(DISTINCT {qs_extract(qs,'device')})         AS qs_device",
            f"COUNT(DISTINCT {qs_extract(qs,'category_name')})  AS qs_category",
            f"COUNT(DISTINCT {qs_extract(qs,'content_title')})  AS qs_content",
            f"COUNT(DISTINCT {qs_extract(qs,'device_id')})      AS qs_device_id",
            f"COUNT(DISTINCT {qs_extract(qs,'session_id')})     AS qs_session_id",
        ]

    counts = {}
    if selects:
        res = safe_df(con, f"SELECT {', '.join(selects)} FROM {reader}", "distinct counts")
        if not res.empty:
            counts = res.iloc[0].to_dict()

    label_col_map = {
        "User Agent (UA)":        counts.get("ua"),
        "ASN":                    counts.get("asn"),
        "City":                   counts.get("city"),
        "State / Region":         counts.get("state"),
        "Country":                counts.get("country"),
        "Request Host":           counts.get("reqhost"),
        "Client IP":              counts.get("ip"),
        "Request Path":           counts.get("path"),
        "Query String channel":   counts.get("qs_channel"),
        "Query String platform":  counts.get("qs_platform"),
        "Query String device":    counts.get("qs_device"),
        "Query String category":  counts.get("qs_category"),
        "Query String content":   counts.get("qs_content"),
        "Query String device_id": counts.get("qs_device_id"),
        "Query String session_id":counts.get("qs_session_id"),
    }

    for i, (label, src_col, present, notes) in enumerate(fields, 3):
        count_val = label_col_map.get(label)
        if count_val is None:
            count_str = "—  (column absent)"
        else:
            count_str = int(count_val)
        _write_data_row(ws, i, [label, src_col, count_str, notes], zebra=(i % 2 == 0))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 22
    ws.freeze_panes = "A3"


def build_value_sheet(ws, con, lake_root: Path, col_expr: str, label: str,
                      value_col_name: str = "Value", top_n: int = TOP_N):
    """Generic builder for a single-column distinct-value sheet."""
    ws.sheet_view.showGridLines = False
    reader = lake_reader(lake_root)

    df = safe_df(con, f"""
        SELECT
            COALESCE(CAST({col_expr} AS VARCHAR), '(null)') AS value,
            COUNT(*) AS count
        FROM {reader}
        WHERE {col_expr} IS NOT NULL
          AND TRIM(CAST({col_expr} AS VARCHAR)) <> ''
        GROUP BY 1
        ORDER BY 2 DESC
        LIMIT {top_n}
    """, label)

    _section_title(ws, 1, f"📋  {label} — Top {top_n:,} distinct values", 4)

    if df.empty:
        ws.cell(row=2, column=1, value="No data found (column may be absent from lake).")
        return

    total = int(df["count"].sum())
    df["% of total"] = (df["count"] / total * 100).round(2)
    df["count"] = df["count"].astype(int)

    headers = [value_col_name, "Row Count", "% of Total"]
    _write_header_row(ws, 2, headers)

    for i, row in enumerate(df.itertuples(index=False), 3):
        _write_data_row(ws, i, [row.value, int(row.count), round(row[2], 2)], zebra=(i % 2 == 0))

    # Totals
    last = 2 + len(df)
    tot_row = last + 1
    for ci, v in enumerate([f"TOTAL (shown {min(top_n, len(df)):,} values)",
                             f"=SUM(B3:B{last})", ""], 1):
        cell = ws.cell(row=tot_row, column=ci, value=v)
        cell.font   = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        cell.fill   = _hdr_fill(C_ORANGE)
        cell.border = _thin_border()

    _add_pct_bar(ws, "C", 3, last)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.freeze_panes = "A3"


def build_sheet_channel_platform(ws, con, lake_root: Path, existing: set):
    """Pivot: channel × platform with request counts."""
    ws.sheet_view.showGridLines = False
    reader = lake_reader(lake_root)
    qs = FIELD_MAP["qs"]
    if qs not in existing:
        ws.cell(row=1, column=1, value="queryStr column not found in lake.")
        return

    _section_title(ws, 1, "📡  Channel × Platform breakdown", 10)

    df = safe_df(con, f"""
        SELECT
            COALESCE({qs_extract(qs,'channel')},
                     {qs_extract(qs,'channel_name')},
                     'Unknown') AS channel,
            COALESCE({qs_extract(qs,'platform')}, 'Unknown') AS platform,
            COUNT(*)                                          AS requests,
            COUNT(DISTINCT {qs_extract(qs,'device_id')})     AS devices,
            COUNT(DISTINCT {qs_extract(qs,'session_id')})    AS sessions
        FROM {reader}
        WHERE {qs} IS NOT NULL
          AND {qs} LIKE '%channel=%'
        GROUP BY 1, 2
        ORDER BY 3 DESC
        LIMIT {TOP_N}
    """, "channel × platform")

    if df.empty:
        ws.cell(row=2, column=1, value="No channel data found.")
        return

    headers = ["Channel", "Platform", "Requests", "Unique Devices", "Unique Sessions"]
    _write_header_row(ws, 2, headers)

    for i, row in enumerate(df.itertuples(index=False), 3):
        _write_data_row(ws, i, [
            row.channel, row.platform,
            int(row.requests), int(row.devices), int(row.sessions)
        ], zebra=(i % 2 == 0))

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "A3"


# ═══════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════

def main():
    lake_root = Path(sys.argv[1]) if len(sys.argv) > 1 else LAKE_FOLDER
    lake_root = lake_root.resolve()

    if not lake_root.exists():
        print(f"❌  Lake folder not found: {lake_root}")
        sys.exit(1)

    print(f"\n{'═'*60}")
    print(f"  Lake Report Generator")
    print(f"  Lake   : {lake_root}")
    print(f"{'═'*60}\n")

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ts_slug      = datetime.now().strftime("%Y%m%d_%H%M")
    out_path     = OUTPUT_FOLDER / f"lake_report_{ts_slug}.xlsx"

    print("  Connecting to DuckDB …")
    con = get_conn(lake_root)

    print("  Detecting columns …")
    existing = detect_columns(con, lake_root)
    print(f"  Found {len(existing)} columns: {', '.join(sorted(existing))}\n")

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    fm  = FIELD_MAP
    qs  = fm["qs"]
    has_qs = qs in existing

    # ── Sheet 1: Overview ─────────────────────────────────────
    print("  Building Sheet 1: Overview …")
    ws1 = wb.create_sheet("Overview")
    build_sheet_overview(ws1, con, lake_root, existing, generated_at)

    # ── Sheet 2: Distinct Counts ──────────────────────────────
    print("  Building Sheet 2: Distinct Counts …")
    ws2 = wb.create_sheet("Distinct Counts")
    build_sheet_distinct_counts(ws2, con, lake_root, existing)

    # ── Sheets 3–8: Per-field distinct values ─────────────────
    field_sheets = [
        ("UA",        fm["ua"],      "User Agent"),
        ("ASN",       fm["asn"],     "ASN"),
        ("City",      fm["city"],    "City"),
        ("State",     fm["state"],   "State / Region"),
        ("Country",   fm["country"], "Country"),
        ("reqHost",   fm["reqhost"], "Request Host"),
    ]
    for sheet_name, col_name, label in field_sheets:
        print(f"  Building Sheet: {sheet_name} …")
        ws = wb.create_sheet(sheet_name)
        if col_name in existing:
            build_value_sheet(ws, con, lake_root, col_name, label, sheet_name)
        else:
            _section_title(ws, 1, f"📋  {label}", 2)
            ws.cell(row=2, column=1, value=f"Column '{col_name}' not found in lake.")

    # ── Query-string derived sheets ───────────────────────────
    qs_sheets = [
        ("Channels",   qs_extract(qs,"channel"),       "Channel (from queryString)",   "Channel"),
        ("Platforms",  qs_extract(qs,"platform"),      "Platform (from queryString)",  "Platform"),
        ("Devices",    qs_extract(qs,"device"),         "Device Name (from queryString)","Device"),
        ("Categories", qs_extract(qs,"category_name"), "Category (from queryString)",  "Category"),
        ("Content",    qs_extract(qs,"content_title"),  "Content Title (from queryString)","Content Title"),
    ]
    for sheet_name, col_expr, label, value_col in qs_sheets:
        print(f"  Building Sheet: {sheet_name} …")
        ws = wb.create_sheet(sheet_name)
        if has_qs:
            build_value_sheet(ws, con, lake_root, col_expr, label, value_col)
        else:
            ws.cell(row=1, column=1, value=f"queryStr column not found in lake.")

    # ── Channel × Platform pivot ──────────────────────────────
    print("  Building Sheet: Channel×Platform …")
    ws_cp = wb.create_sheet("Channel×Platform")
    build_sheet_channel_platform(ws_cp, con, lake_root, existing)

    # ── Save ─────────────────────────────────────────────────
    print(f"\n  Saving → {out_path}")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    con.close()

    print(f"\n{'═'*60}")
    print(f"  ✅  Report saved:")
    print(f"      {out_path}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"      • {s}")
    print(f"{'═'*60}\n")


if __name__ == "__main__":
    main()