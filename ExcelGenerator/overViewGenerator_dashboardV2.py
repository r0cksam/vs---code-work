"""
generate_dashboard.py  —  Fixed version
────────────────────────────────────────
Key fix: pct_* and totals are now computed in Python from raw values.
No longer relies on Excel formula cells (which openpyxl can't always cache).

Usage:  python generate_dashboard.py
"""

import sys, json, urllib.request
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","--quiet"])
    import openpyxl

# ╔══════════════════════════════════════════════════════════╗
# ║  SET THESE PATHS                                         ║
# ╚══════════════════════════════════════════════════════════╝
XLSX_PATH = Path(r"\\YOUR_SERVER\share\overview_report.xlsx")
HTML_OUT  = Path(r"\\YOUR_SERVER\share\dashboard\overview_dashboard.html")
# ══════════════════════════════════════════════════════════

if len(sys.argv) == 3:
    XLSX_PATH, HTML_OUT = Path(sys.argv[1]), Path(sys.argv[2])

# Only the RAW (non-formula) columns — avoids openpyxl formula caching issues
COL = dict(
    DATE=2, BYTES=3, ROWS=4, IP_ROWS=5,
    DIST_IP=6, DIST_IPUA=7, DIST_DEV=8, DIST_SESS=9,
    # Skip J(10) K(11) — formula =E, not useful
    SESS_AVAIL=12, SESS_NA=13, SESS_NONE=14,
    # Skip O-S (15-19) — % formulas, we compute in Python below
)
DATA_START_ROW = 13

def sn(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

CHARTJS_CACHE = Path(__file__).parent / ".chartjs_cache.js"
CHARTJS_URL   = "https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"

def get_chartjs():
    if CHARTJS_CACHE.exists():
        return CHARTJS_CACHE.read_text(encoding="utf-8")
    print("  Downloading Chart.js once...", end=" ", flush=True)
    try:
        req = urllib.request.Request(CHARTJS_URL, headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=20) as r:
            js = r.read().decode("utf-8")
        CHARTJS_CACHE.write_text(js, encoding="utf-8")
        print(f"done ({len(js)//1024}KB cached)")
        return js
    except Exception as e:
        print(f"failed: {e}\n  Charts will be disabled. Run with internet once to cache.")
        return None

print(f"Reading: {XLSX_PATH}")
if not XLSX_PATH.exists():
    print(f"  ERROR: File not found: {XLSX_PATH}"); sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb.active

# Metadata rows 2-9
meta = {}
for r in range(2, 10):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    k = str(row[1]).strip() if row[1] else ''
    v = str(row[2]).strip() if row[2] else ''
    if k and k not in ('None','nan'): meta[k] = v

# Data rows — compute pct in Python, don't read formula cols
data_rows = []
for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
    ds = str(row[COL['DATE']-1]).strip() if row[COL['DATE']-1] else ''
    if not ds or 'TOTAL' in ds.upper() or ds in ('None','nan'): continue
    if '/' not in ds: continue

    ip  = int(sn(row[COL['IP_ROWS']-1]))
    sa  = int(sn(row[COL['SESS_AVAIL']-1]))
    sna = int(sn(row[COL['SESS_NA']-1]))
    sno = int(sn(row[COL['SESS_NONE']-1]))
    pct = lambda n: round(n / ip, 6) if ip > 0 else 0

    data_rows.append({
        "date": ds,
        "bytes":      round(sn(row[COL['BYTES']-1]), 2),
        "rows":       int(sn(row[COL['ROWS']-1])),
        "ip_rows":    ip,
        "dist_ip":    int(sn(row[COL['DIST_IP']-1])),
        "dist_ipua":  int(sn(row[COL['DIST_IPUA']-1])),
        "dist_dev":   int(sn(row[COL['DIST_DEV']-1])),
        "dist_sess":  int(sn(row[COL['DIST_SESS']-1])),
        "sess_avail": sa,
        "sess_na":    sna,
        "sess_none":  sno,
        # Computed in Python — no formula dependency
        "pct_sess":   pct(sa),
        "pct_sessna": pct(sna),
        "pct_none":   pct(sno),
    })
wb.close()

# Totals computed in Python — no SUM formula dependency
keys_to_sum = ["rows","ip_rows","dist_ip","dist_ipua","dist_dev",
               "dist_sess","sess_avail","sess_na","sess_none"]
tot = {k: sum(r[k] for r in data_rows) for k in keys_to_sum}
tot["bytes"] = round(sum(r["bytes"] for r in data_rows), 2)

print(f"  {len(data_rows)} rows | total rows: {tot['rows']:,} | total GiB: {tot['bytes']:,.2f}")

data_blob = json.dumps({
    "meta": meta, "rows": data_rows, "total": tot,
    "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
    "source": XLSX_PATH.name,
}, ensure_ascii=False)

chartjs = get_chartjs()
chartjs_tag = f"<script>{chartjs}</script>" if chartjs else "<script>window.Chart=null;</script>"

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Overview Dashboard</title>
""" + chartjs_tag + """
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#1a1a2e;min-height:100vh}
.topbar{background:#1F3864;color:#fff;padding:14px 28px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}
.topbar h1{font-size:16px;font-weight:600}.topbar .sub{font-size:11.5px;opacity:.6}
#main{padding:20px 26px;max-width:1500px;margin:0 auto}
.meta-strip{background:#fff;border-radius:10px;border:1px solid #e2e8f0;padding:12px 18px;margin-bottom:16px;display:flex;flex-wrap:wrap;gap:8px 26px;font-size:12px;color:#555}
.meta-strip b{color:#1a1a2e;font-weight:600}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:12px;margin-bottom:20px}
.card{background:#fff;border-radius:10px;border:1px solid #e2e8f0;padding:15px 17px}
.clbl{font-size:11px;color:#888;margin-bottom:5px;display:flex;align-items:center;gap:5px}
.cdot{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.cval{font-size:24px;font-weight:700;line-height:1}.csub{font-size:10px;color:#bbb;margin-top:3px}
.charts{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:20px}
@media(max-width:820px){.charts{grid-template-columns:1fr}}
.cbox{background:#fff;border-radius:10px;border:1px solid #e2e8f0;padding:14px 17px}
.cbox h3{font-size:12px;font-weight:600;margin-bottom:10px}
.table-wrap{background:#fff;border-radius:10px;border:1px solid #e2e8f0;overflow:hidden;margin-bottom:18px}
.thead{padding:12px 17px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid #e2e8f0}
.thead h3{font-size:12.5px;font-weight:600}
.tbl-c{overflow-x:auto;max-height:600px;overflow-y:auto}
table{width:100%;border-collapse:collapse;font-size:11.5px}
thead{position:sticky;top:0;z-index:5}
thead tr:first-child th{padding:5px 10px;font-size:10px;font-weight:700;color:#fff;text-align:center;white-space:nowrap}
thead tr:last-child th{padding:7px 10px;text-align:right;font-size:10px;font-weight:600;color:#444;border-bottom:1px solid #e2e8f0;white-space:nowrap;background:#f8fafc}
thead tr:last-child th:first-child{text-align:left}
td{padding:6px 10px;text-align:right;border-bottom:.5px solid #f0f0f0;white-space:nowrap}
td:first-child{text-align:left;font-weight:500;position:sticky;left:0;background:inherit;z-index:1;min-width:80px}
tr:nth-child(even) td{background:#fafafa}
tr:hover td{background:#eef4ff!important}
tr.total-row td{background:#1F3864!important;color:#fff!important;font-weight:700}
tr.drop-row td{background:#f5f5ff}
.pct{color:#185FA5;font-weight:500}.up{color:#2d7a22;font-weight:600}.dn{color:#b71c1c;font-weight:600}
.footer{text-align:center;font-size:11px;color:#ccc;padding:8px 0 20px}
th.b{background:#1F3864} th.g{background:#2e6b10} th.a{background:#9a6000} th.p{background:#4a41a8}
</style>
</head>
<body>
<div class="topbar">
  <h1>&#128202;&nbsp; Overview Report Dashboard</h1>
  <span class="sub" id="genLbl"></span>
</div>
<div id="main">
  <div class="meta-strip" id="metaStrip"></div>
  <div class="cards" id="cards"></div>
  <div class="charts">
    <div class="cbox"><h3>Daily row volume</h3><canvas id="cRows" height="185"></canvas></div>
    <div class="cbox"><h3>Unique visitors (distinct cliIP)</h3><canvas id="cIP" height="185"></canvas></div>
    <div class="cbox"><h3>Session breakdown (daily)</h3><canvas id="cSess" height="185"></canvas></div>
    <div class="cbox"><h3>Daily GiB transferred</h3><canvas id="cBytes" height="185"></canvas></div>
  </div>
  <div class="table-wrap">
    <div class="thead"><h3>Day-by-Day Breakdown</h3><span style="font-size:11px;color:#aaa" id="drLbl"></span></div>
    <div class="tbl-c"><table id="tbl"></table></div>
  </div>
  <div class="footer" id="footer"></div>
</div>
<script>
const D = """ + data_blob + """;
const {meta,rows,total:tot,generated,source} = D;
document.getElementById('genLbl').textContent='Last updated '+generated+'   \u2022   '+source;
document.getElementById('footer').textContent='Re-run generate_dashboard.py after updating '+source+' to refresh';
document.getElementById('metaStrip').innerHTML=Object.entries(meta).map(([k,v])=>`<span><b>${k}:</b> ${v}</span>`).join('');
const fmt=(n,d=0)=>n==null||isNaN(n)?'\u2014':Number(n).toLocaleString('en-IN',{minimumFractionDigits:d,maximumFractionDigits:d});
const fmtP=n=>{if(n==null||isNaN(n))return'\u2014';const p=Math.abs(n)<1?n*100:n;return p.toFixed(2)+'%';};
const ndays=rows.length;
[{l:'Total rows',v:fmt(tot.rows),clr:'#1F3864',s:'from Parquet metadata'},
 {l:'Total GiB',v:fmt(tot.bytes,2),clr:'#0F6E56',s:'data transferred'},
 {l:'Distinct cliIPs',v:fmt(tot.dist_ip),clr:'#4a41a8',s:'unique visitors (cumul.)'},
 {l:'Distinct sessions',v:fmt(tot.dist_sess),clr:'#9a6000',s:'session_id values (cumul.)'},
 {l:'Days loaded',v:ndays,clr:'#2e6b10',s:'date records in file'},
 {l:'Avg rows / day',v:fmt(Math.round(tot.rows/ndays)),clr:'#b71c1c',s:'daily average'},
].forEach(c=>document.getElementById('cards').insertAdjacentHTML('beforeend',
  `<div class="card"><div class="clbl"><span class="cdot" style="background:${c.clr}"></span>${c.l}</div><div class="cval" style="color:${c.clr}">${c.v}</div><div class="csub">${c.s}</div></div>`));
if(rows.length>=2)document.getElementById('drLbl').textContent=rows[0].date+' \u2192 '+rows[rows.length-1].date;
const lbl=rows.map(r=>r.date);
const CD={tension:.35,fill:true,pointRadius:2,pointHoverRadius:5,borderWidth:2};
const mkOpt=yfmt=>({responsive:true,plugins:{legend:{position:'bottom',labels:{font:{size:10},boxWidth:10,padding:8}}},
  scales:{x:{ticks:{font:{size:9},maxTicksLimit:15,maxRotation:45},grid:{color:'#f0f0f0'}},
          y:{ticks:{font:{size:9},callback:v=>yfmt==='gib'?v.toFixed(0)+' G':v>=1e6?(v/1e6).toFixed(1)+'M':v>=1e3?(v/1e3).toFixed(0)+'K':v},grid:{color:'#f0f0f0'}}}});
if(typeof Chart!=='undefined'){
  new Chart(document.getElementById('cRows'),{type:'line',data:{labels:lbl,datasets:[{label:'Total rows',data:rows.map(r=>r.rows),borderColor:'#185FA5',backgroundColor:'rgba(24,95,165,.07)',...CD}]},options:mkOpt()});
  new Chart(document.getElementById('cIP'),{type:'line',data:{labels:lbl,datasets:[{label:'Distinct cliIP',data:rows.map(r=>r.dist_ip),borderColor:'#4a41a8',backgroundColor:'rgba(74,65,168,.07)',...CD},{label:'Distinct device',data:rows.map(r=>r.dist_dev),borderColor:'#0F6E56',backgroundColor:'rgba(15,110,86,.06)',...CD}]},options:mkOpt()});
  new Chart(document.getElementById('cSess'),{type:'line',data:{labels:lbl,datasets:[{label:'Sess+ID avail',data:rows.map(r=>r.sess_avail),borderColor:'#2e6b10',backgroundColor:'rgba(46,107,16,.07)',...CD},{label:'Sess-ID missing',data:rows.map(r=>r.sess_na),borderColor:'#9a6000',backgroundColor:'rgba(154,96,0,.06)',...CD},{label:'Sess not found',data:rows.map(r=>r.sess_none),borderColor:'#b71c1c',backgroundColor:'rgba(183,28,28,.06)',...CD}]},options:mkOpt()});
  new Chart(document.getElementById('cBytes'),{type:'line',data:{labels:lbl,datasets:[{label:'GiB',data:rows.map(r=>r.bytes),borderColor:'#1D9E75',backgroundColor:'rgba(29,158,117,.08)',...CD}]},options:mkOpt('gib')});
}
const subH=['Date (IST)','Bytes (GiB)','Total Rows','cliIP rows','Distinct cliIP','Dist cliIP+UA','Distinct device','Distinct session','Sess+ID avail','Sess-ID missing','Sess not found','% Sess+ID','% Sess-ID','% Sess none'];
let h=`<thead><tr><th class="b" style="background:#f8fafc;color:#666"></th><th colspan="4" class="b">Day-by-Day Breakdown</th><th colspan="4" class="g">Unique Set</th><th colspan="3" class="a">Rows Against</th><th colspan="3" class="p">% of cliIP rows</th></tr><tr>${subH.map(s=>`<th>${s}</th>`).join('')}</tr></thead><tbody>`;
rows.forEach(d=>{h+=`<tr><td>${d.date}</td><td>${fmt(d.bytes,2)}</td><td>${fmt(d.rows)}</td><td>${fmt(d.ip_rows)}</td><td>${fmt(d.dist_ip)}</td><td>${fmt(d.dist_ipua)}</td><td>${fmt(d.dist_dev)}</td><td>${fmt(d.dist_sess)}</td><td>${fmt(d.sess_avail)}</td><td>${fmt(d.sess_na)}</td><td>${fmt(d.sess_none)}</td><td class="pct">${fmtP(d.pct_sess)}</td><td class="pct">${fmtP(d.pct_sessna)}</td><td class="pct">${fmtP(d.pct_none)}</td></tr>`;});
if(rows.length>=6){const last=rows[rows.length-2],prev=rows.slice(-6,-2);const avg=k=>prev.reduce((s,r)=>s+r[k],0)/prev.length;const drop=k=>{const d=last[k]/avg(k)-1;if(!isFinite(d))return'\u2014';return`<span class="${d>=0?'up':'dn'}">${d>=0?'+':''}${(d*100).toFixed(2)}%</span>`;};h+=`<tr class="drop-row"><td style="font-weight:600;font-size:10.5px;color:#4a41a8">% vs prev<br><small style="font-weight:400;color:#bbb">4-day avg</small></td>${['bytes','rows','ip_rows','dist_ip','dist_ipua','dist_dev','dist_sess','sess_avail','sess_na','sess_none','pct_sess','pct_sessna','pct_none'].map(k=>`<td>${drop(k)}</td>`).join('')}</tr>`;}
h+=`<tr class="total-row"><td>TOTAL</td><td>${fmt(tot.bytes,2)}</td><td>${fmt(tot.rows)}</td><td>${fmt(tot.ip_rows)}</td><td>${fmt(tot.dist_ip)}</td><td>${fmt(tot.dist_ipua)}</td><td>${fmt(tot.dist_dev)}</td><td>${fmt(tot.dist_sess)}</td><td>${fmt(tot.sess_avail)}</td><td>${fmt(tot.sess_na)}</td><td>${fmt(tot.sess_none)}</td><td colspan="3" style="text-align:center;opacity:.6;font-size:9.5px">— per-row only —</td></tr>`;
h+='</tbody>';document.getElementById('tbl').innerHTML=h;
</script>
</body>
</html>"""

HTML_OUT.parent.mkdir(parents=True, exist_ok=True)
HTML_OUT.write_text(HTML, encoding='utf-8')
size_kb = HTML_OUT.stat().st_size // 1024
print(f"\nDone -> {HTML_OUT}")
print(f"  {size_kb} KB | {len(data_rows)} days | {datetime.now().strftime('%H:%M')}")