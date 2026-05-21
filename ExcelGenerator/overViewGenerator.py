#!/usr/bin/env python3
"""
overview_generator.py
═══════════════════════════════════════════════════════════════════════
Standalone, append-aware Overview sheet generator.
Reproduces the exact Overview format with colored section headers.

First run  → creates  <output_dir>/overview_report.xlsx
Later runs → queries only new dates, merges with existing data,
             rewrites the file so formatting stays consistent.

Run:
    python overview_generator.py
    python overview_generator.py <lake_folder>

Requires:
    duckdb, pandas, pyarrow, openpyxl, psutil
"""

from __future__ import annotations

import logging
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional

import psutil
import duckdb
import pandas as pd
import pyarrow.parquet as pq
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────

logging.basicConfig(level=logging.INFO, format="  %(levelname)-7s %(message)s")
log = logging.getLogger("overview_generator")


# ──────────────────────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────────────────────

IST_OFFSET        = timedelta(hours=5, minutes=30)
OVERVIEW_FILENAME = "overview_report.xlsx"
OVERVIEW_SHEET    = "Overview"

# ── Column indices (1-indexed for openpyxl) ──────────────────
#   A (1) = reserved for % Drop label
#   B (2) = Date (IST)  …  S (19) = % Session Not Found
COL_DATE         = 2   # B
COL_BYTES        = 3   # C – Total Bytes (GiB)
COL_ROWS         = 4   # D – Total Rows* (from Parquet metadata)
COL_IP_ROWS      = 5   # E – cliIP rows
COL_DIST_IP      = 6   # F – Distinct cliIP
COL_DIST_IP_UA   = 7   # G – Distinct cliIP & UA
COL_DIST_DEV     = 8   # H – Distinct device_id
COL_DIST_SESS    = 9   # I – Distinct session_id
COL_UNIQ_IP      = 10  # J – UNIQUE SET: Distinct cliIP  (formula =E)
COL_UNIQ_IP_UA   = 11  # K – UNIQUE SET: Distinct cliIP & UA  (formula =E)
COL_SESS_AVAIL   = 12  # L – ROWS AGAINST: Session Found - ID Available
COL_SESS_NA      = 13  # M – ROWS AGAINST: Session Found - ID NOT Available
COL_SESS_NONE    = 14  # N – ROWS AGAINST: Session Not Found
COL_PCT_IP       = 15  # O – % Distinct cliIP  (formula =J/$E)
COL_PCT_IP_UA    = 16  # P – % Distinct cliIP & UA  (formula =K/$E)
COL_PCT_SESS     = 17  # Q – % Session Found - ID Available  (formula =L/$E)
COL_PCT_SESS_NA  = 18  # R – % Session Found - ID NOT Available  (formula =M/$E)
COL_PCT_NONE     = 19  # S – % Session Not Found  (formula =N/$E)
LAST_COL         = COL_PCT_NONE  # 19

# ── Fixed row positions ───────────────────────────────────────
ROW_TITLE_META   = 1
ROW_META_START   = 2   # first key/value pair
ROW_SECTION_HDR  = 11  # colored group header row
ROW_COL_HDR      = 12  # column-label header row
DATA_START_ROW   = 13  # first data row

# ── Colors (hex, no leading #) ───────────────────────────────
C_DARK_BLUE = "1F3864"   # title / day-breakdown section
C_KV_BG     = "D6E4F0"   # metadata key background
C_WHITE     = "FFFFFF"
C_BLACK     = "000000"
C_GREEN     = "92D050"   # UNIQUE SET section
C_AMBER     = "FFC000"   # ROWS AGAINST section
C_PURPLE    = "7030A0"   # % Data of TOTAL section
C_ORANGE    = "E36209"   # TOTAL row
C_ALT       = "F2F2F2"   # alternating row background


# ══════════════════════════════════════════════════════════════
# SECTION A – DuckDB / Lake helpers
# ══════════════════════════════════════════════════════════════

def get_conn() -> duckdb.DuckDBPyConnection:
    mem_avail = psutil.virtual_memory().available / (1024 ** 3)
    cpus      = os.cpu_count() or 4
    mem_gb    = max(4, min(int(mem_avail * 0.70), 48))
    threads   = max(1, min(cpus // 4, 4))
    con = duckdb.connect()
    con.execute(f"SET threads={threads}")
    con.execute(f"SET memory_limit='{mem_gb}GB'")
    con.execute("SET preserve_insertion_order=false")
    return con


def lake_reader(lake_root: Path, year: str | None, month: str | None) -> str:
    if year and month:
        path = lake_root / f"year={year}" / f"month={month}"
        if path.exists():
            return f"read_parquet('{path.as_posix()}/**/*.parquet', union_by_name=true)"
    return (
        f"read_parquet('{lake_root.as_posix()}/**/*.parquet', "
        f"hive_partitioning=true, union_by_name=true)"
    )


def qs_extract(qs_col: str, param: str) -> str:
    return f"NULLIF(regexp_extract({qs_col}, '(?:^|&){param}=([^&]+)', 1), '')"


def pa_schema(lake_root: Path, year: str | None, month: str | None) -> set[str]:
    scan_root = lake_root
    if year:
        scan_root = lake_root / f"year={year}"
        if month:
            scan_root = scan_root / f"month={month}"
    cols: set[str] = set()
    skip = {"year", "month", "day"}
    for pf in scan_root.rglob("*.parquet"):
        try:
            s = pq.read_schema(pf)
            cols.update(n for n in s.names if n not in skip)
        except Exception:
            pass
        if len(cols) > 8:
            break
    return cols


def pa_row_counts_by_day(lake_root: Path, year: str | None, month: str | None) -> dict[str, int]:
    scan_root = lake_root
    if year:
        scan_root = lake_root / f"year={year}"
        if month:
            scan_root = scan_root / f"month={month}"
    counts: dict[str, int] = {}
    for day_dir in sorted(scan_root.rglob("day=*")):
        if not day_dir.is_dir():
            continue
        try:
            parts = {p.split("=")[0]: p.split("=")[1] for p in day_dir.parts if "=" in p}
            key = f"{parts.get('year','0')}-{parts.get('month','0')}-{parts.get('day','0')}"
            total = sum(pq.read_metadata(pf).num_rows for pf in day_dir.glob("*.parquet"))
            counts[key] = counts.get(key, 0) + total
        except Exception:
            pass
    return counts


def pa_total_file_count(root: Path) -> int:
    return sum(1 for _ in root.rglob("*.parquet"))


def safe_query(con, sql: str, label: str = "") -> pd.DataFrame:
    try:
        return con.execute(sql).df()
    except Exception as exc:
        log.warning(f"Query failed [{label}]: {exc}")
        return pd.DataFrame()


def utc_date_to_ist_str(utc_date) -> str:
    """Convert UTC date (date/Timestamp/datetime) → 'DD/MM/YY' IST string."""
    if isinstance(utc_date, pd.Timestamp):
        dt = utc_date.to_pydatetime()
    elif isinstance(utc_date, datetime):
        dt = utc_date
    else:
        dt = datetime.combine(utc_date, datetime.min.time())
    return (dt + IST_OFFSET).strftime("%d/%m/%y")


# ══════════════════════════════════════════════════════════════
# SECTION B – Prompt helpers
# ══════════════════════════════════════════════════════════════

def prompt_path(question: str) -> Path:
    while True:
        raw = input(question).strip()
        if not raw:
            print("  Path cannot be empty.")
            continue
        p = Path(raw).expanduser().resolve()
        if p.is_dir():
            return p
        print(f"  ❌ Directory not found: {p}")


def prompt_year_month(lake_root: Path) -> tuple[str | None, str | None]:
    print("\n📅  Filter by month? (y/n): ", end="")
    try:
        if input().strip().lower() != "y":
            return None, None
        available = sorted(
            d.name.replace("year=", "")
            for d in lake_root.iterdir()
            if d.is_dir() and d.name.startswith("year=")
        )
        if available:
            print(f"  Available years: {', '.join(available)}")
        while True:
            y = input("  Year (e.g. 2026): ").strip()
            if y.isdigit() and 2000 <= int(y) <= 2100:
                break
            print("  ❌ Invalid year")
        while True:
            m = input("  Month 1–12: ").strip()
            if m.isdigit() and 1 <= int(m) <= 12:
                break
            print("  ❌ Invalid month")
        year_str  = y
        month_str = f"{int(m):02d}"
        partition = lake_root / f"year={year_str}" / f"month={month_str}"
        if not partition.exists():
            print(f"  ⚠️  Partition not found: {partition}. Continue? (y/n): ", end="")
            if input().strip().lower() != "y":
                return None, None
        return year_str, month_str
    except KeyboardInterrupt:
        print("\n  Cancelled.")
        return None, None


# ══════════════════════════════════════════════════════════════
# SECTION C – Append helpers  (read existing Overview data)
# ══════════════════════════════════════════════════════════════

def get_existing_dates(output_path: Path) -> set[str]:
    """Return the set of 'DD/MM/YY' IST date strings already in the Overview."""
    if not output_path.exists():
        return set()
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
        if OVERVIEW_SHEET not in wb.sheetnames:
            wb.close()
            return set()
        ws = wb[OVERVIEW_SHEET]
        dates: set[str] = set()
        for row in ws.iter_rows(
            min_row=DATA_START_ROW, max_col=COL_DATE, min_col=COL_DATE, values_only=True
        ):
            val = row[0]
            if val and isinstance(val, str) and "/" in val:
                dates.add(val.strip())
        wb.close()
        return dates
    except Exception as exc:
        log.warning(f"Could not read existing dates: {exc}")
        return set()


def load_existing_rows(output_path: Path) -> list[dict]:
    """Load data rows from an existing Overview sheet as a list of dicts."""
    if not output_path.exists():
        return []
    try:
        wb = load_workbook(output_path, read_only=True, data_only=True)
        if OVERVIEW_SHEET not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[OVERVIEW_SHEET]
        rows: list[dict] = []
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            date_val = row[COL_DATE - 1]        # 0-indexed
            if not date_val or not isinstance(date_val, str) or "/" not in str(date_val):
                break  # hit TOTAL row or empty area
            def _int(v):
                try:
                    return int(v) if v is not None else 0
                except Exception:
                    return 0
            def _flt(v):
                try:
                    return float(v) if v is not None else 0.0
                except Exception:
                    return 0.0
            rows.append({
                "ist_date":   str(date_val).strip(),
                "bytes_gib":  _flt(row[COL_BYTES      - 1]),
                "total_rows": _int(row[COL_ROWS        - 1]),
                "ip_rows":    _int(row[COL_IP_ROWS     - 1]),
                "dist_ip":    _int(row[COL_DIST_IP     - 1]),
                "dist_ip_ua": _int(row[COL_DIST_IP_UA  - 1]),
                "dist_dev":   _int(row[COL_DIST_DEV    - 1]),
                "dist_sess":  _int(row[COL_DIST_SESS   - 1]),
                "sess_avail": _int(row[COL_SESS_AVAIL  - 1]),
                "sess_na":    _int(row[COL_SESS_NA     - 1]),
                "sess_none":  _int(row[COL_SESS_NONE   - 1]),
            })
        wb.close()
        return rows
    except Exception as exc:
        log.warning(f"Could not load existing rows: {exc}")
        return []


# ══════════════════════════════════════════════════════════════
# SECTION D – DuckDB query
# ══════════════════════════════════════════════════════════════

def query_new_dates(
    con,
    lake_root: Path,
    year: str | None,
    month: str | None,
    cols: set[str],
    existing_dates: set[str],
    day_counts: dict[str, int],
) -> list[dict]:
    """Query DuckDB for day-by-day stats, skip dates already in existing_dates."""
    ip_col = "cliIP"
    ua_col = "UA"
    qs_col = "queryStr"
    ts_col = "reqTimeSec"

    if ts_col not in cols:
        log.warning("reqTimeSec not found — cannot build day breakdown.")
        return []

    has_ip = ip_col in cols
    has_ua = ua_col in cols
    has_qs = qs_col in cols
    has_tb = "totalBytes" in cols

    ip_rows_expr   = f"COUNT({ip_col})"                             if has_ip else "0"
    d_ip_expr      = f"COUNT(DISTINCT {ip_col})"                    if has_ip else "0"
    d_ipua_expr    = f"COUNT(DISTINCT ({ip_col}, {ua_col}))"        if (has_ip and has_ua) else "0"
    bytes_expr     = f"SUM(CAST(totalBytes AS BIGINT))"             if has_tb else "NULL"
    d_dev_expr     = f"COUNT(DISTINCT {qs_extract(qs_col,'device_id')})"   if has_qs else "0"
    d_sess_expr    = f"COUNT(DISTINCT {qs_extract(qs_col,'session_id')})"  if has_qs else "0"

    if has_qs:
        sess_extract  = qs_extract(qs_col, "session_id")
        sess_present  = f"SUM(CASE WHEN {qs_col} LIKE '%session_id=%' THEN 1 ELSE 0 END)"
        sess_blank    = (
            f"SUM(CASE WHEN {qs_col} LIKE '%session_id=%' "
            f"AND {sess_extract} IS NULL THEN 1 ELSE 0 END)"
        )
        sess_avail    = f"({sess_present} - {sess_blank})"
        sess_na       = sess_blank
        sess_none     = (
            f"SUM(CASE WHEN {qs_col} IS NULL "
            f"OR {qs_col} NOT LIKE '%session_id=%' THEN 1 ELSE 0 END)"
        )
    else:
        sess_avail = sess_na = sess_none = "0"

    reader = lake_reader(lake_root, year, month)
    df = safe_query(con, f"""
        SELECT
            make_date(CAST(year AS INT), CAST(month AS INT), CAST(day AS INT)) AS utc_date,
            COUNT(*)        AS total_rows,
            {ip_rows_expr}  AS ip_rows,
            {d_ip_expr}     AS distinct_ip,
            {d_ipua_expr}   AS distinct_ipua,
            {bytes_expr}    AS total_bytes,
            {d_dev_expr}    AS distinct_dev,
            {d_sess_expr}   AS distinct_sess,
            {sess_avail}    AS sess_avail,
            {sess_na}       AS sess_na,
            {sess_none}     AS sess_none
        FROM {reader}
        GROUP BY year, month, day
        ORDER BY utc_date
    """, "day breakdown")

    if df.empty:
        return []

    new_rows: list[dict] = []
    for _, row in df.iterrows():
        ist_str = utc_date_to_ist_str(row["utc_date"])
        if ist_str in existing_dates:
            log.info(f"  Skipping {ist_str} (already loaded)")
            continue
        utc_key   = str(row["utc_date"])
        pa_count  = day_counts.get(utc_key, int(row["total_rows"]))
        bytes_gib = (
            float(row["total_bytes"]) / (1024 ** 3)
            if has_tb and pd.notna(row["total_bytes"]) else 0.0
        )
        new_rows.append({
            "ist_date":   ist_str,
            "bytes_gib":  round(bytes_gib, 2),
            "total_rows": pa_count,
            "ip_rows":    int(row["ip_rows"]),
            "dist_ip":    int(row["distinct_ip"]),
            "dist_ip_ua": int(row["distinct_ipua"]),
            "dist_dev":   int(row["distinct_dev"]),
            "dist_sess":  int(row["distinct_sess"]),
            "sess_avail": int(row["sess_avail"]),
            "sess_na":    int(row["sess_na"]),
            "sess_none":  int(row["sess_none"]),
        })
    return new_rows


# ══════════════════════════════════════════════════════════════
# SECTION E – openpyxl style helpers
# ══════════════════════════════════════════════════════════════

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(color: str = C_BLACK, bold: bool = False, size: int = 10) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)


def _align(h: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write(
    ws,
    row: int,
    col: int,
    value=None,
    fill_hex: str = None,
    font_color: str = C_BLACK,
    bold: bool = False,
    h_align: str = "left",
    num_fmt: str = None,
    wrap: bool = False,
):
    c = ws.cell(row=row, column=col, value=value)
    if fill_hex:
        c.fill = _fill(fill_hex)
    c.font      = _font(font_color, bold)
    c.alignment = _align(h_align, wrap)
    c.border    = _thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


# ══════════════════════════════════════════════════════════════
# SECTION F – Excel writer
# ══════════════════════════════════════════════════════════════

def write_overview_excel(
    output_path:      Path,
    data_rows:        list[dict],
    lake_root:        Path,
    year_filter:      str | None,
    month_filter:     str | None,
    total_rows_pa:    int,
    total_files:      int,
    date_range_ist:   str,
    time_range_ist:   str,
) -> None:
    """Write (or rewrite) the Overview Excel file with exact target formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = OVERVIEW_SHEET

    # ── Column widths ─────────────────────────────────────────
    ws.column_dimensions["A"].width = 30     # % Drop label
    ws.column_dimensions["B"].width = 14     # Date
    ws.column_dimensions["C"].width = 14     # Bytes
    ws.column_dimensions[get_column_letter(COL_ROWS)].width       = 14
    ws.column_dimensions[get_column_letter(COL_IP_ROWS)].width    = 13
    ws.column_dimensions[get_column_letter(COL_DIST_IP)].width    = 13
    ws.column_dimensions[get_column_letter(COL_DIST_IP_UA)].width = 16
    ws.column_dimensions[get_column_letter(COL_DIST_DEV)].width   = 14
    ws.column_dimensions[get_column_letter(COL_DIST_SESS)].width  = 14
    ws.column_dimensions[get_column_letter(COL_UNIQ_IP)].width    = 13
    ws.column_dimensions[get_column_letter(COL_UNIQ_IP_UA)].width = 16
    ws.column_dimensions[get_column_letter(COL_SESS_AVAIL)].width = 14
    ws.column_dimensions[get_column_letter(COL_SESS_NA)].width    = 16
    ws.column_dimensions[get_column_letter(COL_SESS_NONE)].width  = 14
    ws.column_dimensions[get_column_letter(COL_PCT_IP)].width     = 13
    ws.column_dimensions[get_column_letter(COL_PCT_IP_UA)].width  = 16
    ws.column_dimensions[get_column_letter(COL_PCT_SESS)].width   = 14
    ws.column_dimensions[get_column_letter(COL_PCT_SESS_NA)].width= 16
    ws.column_dimensions[get_column_letter(COL_PCT_NONE)].width   = 14

    # Row heights
    ws.row_dimensions[ROW_SECTION_HDR].height = 22
    ws.row_dimensions[ROW_COL_HDR].height     = 42

    # ── Row 1: "📊 Report Metadata" title ──────────────────────
    ws.merge_cells(start_row=ROW_TITLE_META, start_column=2,
                   end_row=ROW_TITLE_META, end_column=3)
    _write(ws, ROW_TITLE_META, 2, "📊  Report Metadata",
           fill_hex=C_DARK_BLUE, font_color=C_WHITE, bold=True)

    # ── Rows 2–9: key/value metadata ──────────────────────────
    ist_now = datetime.now(timezone.utc) + IST_OFFSET
    kv_pairs = [
        ("Report generated",      ist_now.strftime("%Y-%m-%d %H:%M IST")),
        ("Filtered to",           f"year={year_filter}, month={month_filter}"
                                  if year_filter else "full lake"),
        ("Lake folder",           str(lake_root)),
        ("DuckDB version",        duckdb.__version__),
        ("Total rows (metadata)", f"{total_rows_pa:,}"),
        ("Total parquet files",   f"{total_files:,}"),
        ("Data date range (IST)", date_range_ist),
        ("Data time range (IST)", time_range_ist),
    ]
    for i, (key, val) in enumerate(kv_pairs):
        r = ROW_META_START + i
        _write(ws, r, 2, key, fill_hex=C_KV_BG, font_color=C_BLACK, bold=True)
        _write(ws, r, 3, val, fill_hex=C_WHITE,  font_color=C_BLACK, bold=False)

    # ── Row 11: colored section-group headers (merged) ────────
    section_groups = [
        (2,  5,  "📅  Day-by-Day Breakdown (IST dates)", C_DARK_BLUE),
        (6,  9,  "UNIQUE SET",                           C_GREEN),
        (10, 14, "ROWS AGAINST - UNIQUE SET",            C_AMBER),
        (15, 19, "% Data of TOTAL",                      C_PURPLE),
    ]
    for start_c, end_c, label, color in section_groups:
        ws.merge_cells(start_row=ROW_SECTION_HDR, start_column=start_c,
                   end_row=ROW_SECTION_HDR, end_column=end_c)
    # Label on the first cell only
    _write(ws, ROW_SECTION_HDR, start_c, label,
           fill_hex=color, font_color=C_WHITE, bold=True, h_align="center")
    # Other cells in the merged range: fill and border only, no value
    for c in range(start_c + 1, end_c + 1):
        cell = ws.cell(row=ROW_SECTION_HDR, column=c)
        cell.fill = _fill(color)
        cell.border = _thin_border()           
        cell.alignment = _align("center")

    # ── Row 12: column labels (match section colors) ──────────
    col_labels = [
        (COL_DATE,       "Date (IST)",                      C_DARK_BLUE),
        (COL_BYTES,      "Total Bytes (GiB)",                C_DARK_BLUE),
        (COL_ROWS,       "Total Rows*",                      C_DARK_BLUE),
        (COL_IP_ROWS,    "cliIP rows",                       C_DARK_BLUE),
        (COL_DIST_IP,    "Distinct cliIP",                   C_GREEN),
        (COL_DIST_IP_UA, "Distinct cliIP & UA",              C_GREEN),
        (COL_DIST_DEV,   "Distinct device_id",               C_GREEN),
        (COL_DIST_SESS,  "Distinct session_id",              C_GREEN),
        (COL_UNIQ_IP,    "Distinct cliIP",                   C_AMBER),
        (COL_UNIQ_IP_UA, "Distinct cliIP & UA",              C_AMBER),
        (COL_SESS_AVAIL, "Session Found - ID Available",     C_AMBER),
        (COL_SESS_NA,    "Session Found - ID NOT Available", C_AMBER),
        (COL_SESS_NONE,  "Session Not Found",                C_AMBER),
        (COL_PCT_IP,     "Distinct cliIP",                   C_PURPLE),
        (COL_PCT_IP_UA,  "Distinct cliIP & UA",              C_PURPLE),
        (COL_PCT_SESS,   "Session Found - ID Available",     C_PURPLE),
        (COL_PCT_SESS_NA,"Session Found - ID NOT Available", C_PURPLE),
        (COL_PCT_NONE,   "Session Not Found",                C_PURPLE),
    ]
    for col_idx, label, color in col_labels:
        _write(ws, ROW_COL_HDR, col_idx, label,
               fill_hex=color, font_color=C_WHITE, bold=True,
               h_align="center", wrap=True)

    # ── Rows 13+: data rows ────────────────────────────────────
    e_ltr = get_column_letter(COL_IP_ROWS)   # E
    j_ltr = get_column_letter(COL_UNIQ_IP)   # J
    k_ltr = get_column_letter(COL_UNIQ_IP_UA) # K
    l_ltr = get_column_letter(COL_SESS_AVAIL) # L
    m_ltr = get_column_letter(COL_SESS_NA)    # M
    n_ltr = get_column_letter(COL_SESS_NONE)  # N

    r = DATA_START_ROW
    for i, d in enumerate(data_rows):
        bg = C_ALT if i % 2 == 1 else C_WHITE

        _write(ws, r, COL_DATE,       d["ist_date"],   bg, C_BLACK, h_align="center")
        _write(ws, r, COL_BYTES,      d["bytes_gib"],  bg, C_BLACK, h_align="right", num_fmt="#,##0.00")
        _write(ws, r, COL_ROWS,       d["total_rows"], bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_IP_ROWS,    d["ip_rows"],    bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_IP,    d["dist_ip"],    bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_IP_UA, d["dist_ip_ua"],bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_DEV,   d["dist_dev"],   bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_DIST_SESS,  d["dist_sess"],  bg, C_BLACK, h_align="right", num_fmt="#,##0")
        # UNIQUE SET – mirrors cliIP rows (col E)
        _write(ws, r, COL_UNIQ_IP,    f"={e_ltr}{r}",  bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_UNIQ_IP_UA, f"={e_ltr}{r}",  bg, C_BLACK, h_align="right", num_fmt="#,##0")
        # ROWS AGAINST
        _write(ws, r, COL_SESS_AVAIL, d["sess_avail"], bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_SESS_NA,    d["sess_na"],     bg, C_BLACK, h_align="right", num_fmt="#,##0")
        _write(ws, r, COL_SESS_NONE,  d["sess_none"],   bg, C_BLACK, h_align="right", num_fmt="#,##0")
        # % Data of TOTAL – formula /{E locked}
        cur_e = f"${e_ltr}{r}"
        _write(ws, r, COL_PCT_IP,     f"={j_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_IP_UA,  f"={k_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_SESS,   f"={l_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_SESS_NA,f"={m_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        _write(ws, r, COL_PCT_NONE,   f"={n_ltr}{r}/{cur_e}", bg, C_BLACK, h_align="right", num_fmt="0.00%")
        r += 1

    data_end_row = r - 1   # last data row (1-indexed)
    total_row    = r        # TOTAL is immediately after last data row

    # ── TOTAL row (orange) ─────────────────────────────────────
    _write(ws, total_row, COL_DATE, "TOTAL*",
           fill_hex=C_ORANGE, font_color=C_WHITE, bold=True, h_align="left")
    sum_cols = [
        COL_BYTES, COL_ROWS, COL_IP_ROWS,
        COL_DIST_IP, COL_DIST_IP_UA, COL_DIST_DEV, COL_DIST_SESS,
        COL_UNIQ_IP, COL_UNIQ_IP_UA,
        COL_SESS_AVAIL, COL_SESS_NA, COL_SESS_NONE,
    ]
    for col_idx in sum_cols:
        ltr = get_column_letter(col_idx)
        nfmt = "#,##0.00" if col_idx == COL_BYTES else "#,##0"
        _write(ws, total_row, col_idx,
               f"=SUM({ltr}{DATA_START_ROW}:{ltr}{data_end_row})",
               fill_hex=C_ORANGE, font_color=C_WHITE, bold=True,
               h_align="right", num_fmt=nfmt)

    # ── % Drop/Gain row (2 rows below TOTAL) ──────────────────
    drop_row = total_row + 2
    ws.merge_cells(start_row=drop_row, start_column=1,
                   end_row=drop_row, end_column=2)
    _write(ws, drop_row, 1, "% Drop/Gain w.r.t. last 5 days",
           font_color=C_BLACK, bold=True, h_align="left")

    # Need at least 6 data rows to build a meaningful drop formula
    # Compares last-but-one day (most likely "complete") to avg of 4 days before it
    n_rows = data_end_row - DATA_START_ROW + 1
    if n_rows >= 6:
        last_full = data_end_row - 1
        avg_start = data_end_row - 5
        avg_end   = data_end_row - 2
        drop_cols = [
            COL_BYTES, COL_ROWS, COL_IP_ROWS,
            COL_DIST_IP, COL_DIST_IP_UA, COL_DIST_DEV, COL_DIST_SESS,
            COL_UNIQ_IP, COL_UNIQ_IP_UA,
            COL_SESS_AVAIL, COL_SESS_NA, COL_SESS_NONE,
            COL_PCT_SESS, COL_PCT_SESS_NA, COL_PCT_NONE,
        ]
        for col_idx in drop_cols:
            ltr = get_column_letter(col_idx)
            _write(ws, drop_row, col_idx,
                   f"=({ltr}{last_full}/AVERAGE({ltr}{avg_start}:{ltr}{avg_end}))-1",
                   font_color=C_BLACK, h_align="right", num_fmt="0.00%")

    # ── Footnote ───────────────────────────────────────────────
    fn_row = drop_row + 2
    c = ws.cell(row=fn_row, column=COL_DATE,
                value="* Total Rows sourced from Parquet file metadata (zero-scan). "
                      "All other columns via DuckDB aggregation.")
    c.font = _font(C_BLACK, bold=False)

    # ── Freeze top rows and date column ───────────────────────
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=COL_DATE + 1)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


# ══════════════════════════════════════════════════════════════
# SECTION G – Main run
# ══════════════════════════════════════════════════════════════

def run(lake_root: Path) -> None:
    print("\n" + "═" * 60)
    print("  Overview Generator  –  append-aware")
    print("═" * 60 + "\n")

    # ── Output directory ──────────────────────────────────────
    print("💾  Output folder (Excel will be saved / updated here):")
    out_dir     = prompt_path("  Output folder path: ")
    output_file = out_dir / OVERVIEW_FILENAME
    is_first    = not output_file.exists()

    existing_dates = get_existing_dates(output_file)
    if is_first:
        print("\n  📋  No existing file — will create fresh.")
    else:
        sorted_d = sorted(existing_dates)
        print(f"\n  📋  Found existing file with {len(existing_dates)} date(s) loaded.")
        if sorted_d:
            print(f"      Earliest: {sorted_d[0]}   Latest: {sorted_d[-1]}")

    # ── Date filter ───────────────────────────────────────────
    year_filter, month_filter = prompt_year_month(lake_root)

    # ── Confirm ───────────────────────────────────────────────
    print("\n" + "─" * 60)
    print(f"  Lake        : {lake_root}")
    print(f"  Output file : {output_file}")
    print(f"  Filter      : " + (
        f"year={year_filter}" + (f", month={month_filter}" if month_filter else "")
        if year_filter else "none (full lake)"
    ))
    print(f"  Mode        : {'CREATE (first run)' if is_first else 'APPEND (new dates only)'}")
    print("─" * 60)
    if input("\n  Proceed? (y/n): ").strip().lower() != "y":
        print("  Cancelled.")
        return

    # ── Parquet metadata (zero-scan) ──────────────────────────
    log.info("Reading Parquet metadata (zero-scan) …")
    day_counts  = pa_row_counts_by_day(lake_root, year_filter, month_filter)
    total_rows  = sum(day_counts.values())
    scan_root   = lake_root / f"year={year_filter}" if year_filter else lake_root
    total_files = pa_total_file_count(scan_root)
    log.info(f"  {total_rows:,} rows across {total_files:,} parquet files")

    # ── Schema detection ──────────────────────────────────────
    log.info("Detecting schema …")
    cols = pa_schema(lake_root, year_filter, month_filter)
    log.info(f"  {len(cols)} columns detected")

    # ── DuckDB connection ─────────────────────────────────────
    log.info("Connecting to DuckDB …")
    con = get_conn()

    # ── Date range metadata ───────────────────────────────────
    date_range_ist = time_range_ist = "N/A"
    ts = "reqTimeSec"
    if ts in cols:
        reader = lake_reader(lake_root, year_filter, month_filter)
        meta   = safe_query(con, f"""
            SELECT
                MIN(to_timestamp(CAST({ts} AS DOUBLE)))::DATE AS min_d,
                MAX(to_timestamp(CAST({ts} AS DOUBLE)))::DATE AS max_d,
                MIN(to_timestamp(CAST({ts} AS DOUBLE)))       AS min_ts,
                MAX(to_timestamp(CAST({ts} AS DOUBLE)))       AS max_ts
            FROM {reader}
        """, "date range")
        if not meta.empty:
            row = meta.iloc[0]
            date_range_ist = (
                f"{(row['min_d']  + IST_OFFSET).strftime('%Y-%m-%d')} → "
                f"{(row['max_d']  + IST_OFFSET).strftime('%Y-%m-%d')}"
            )
            time_range_ist = (
                f"{(row['min_ts'] + IST_OFFSET).strftime('%Y-%m-%d %H:%M:%S')} → "
                f"{(row['max_ts'] + IST_OFFSET).strftime('%Y-%m-%d %H:%M:%S')}"
            )

    # ── Query new dates ───────────────────────────────────────
    log.info("Querying day-by-day breakdown …")
    new_rows = query_new_dates(
        con, lake_root, year_filter, month_filter,
        cols, existing_dates, day_counts,
    )
    con.close()

    if not new_rows:
        print("\n  ✅  Nothing new to add — your file is already up to date.")
        return

    skipped = len(existing_dates)
    print(f"\n  ➕  {len(new_rows)} new date(s) to add."
          + (f" ({skipped} already existed, skipped)" if skipped else ""))

    # ── Merge with existing ───────────────────────────────────
    if is_first:
        all_rows = new_rows
    else:
        log.info("Loading existing data rows …")
        old_rows = load_existing_rows(output_file)
        # Remove overlap (safety) then combine
        new_dates_set = {r["ist_date"] for r in new_rows}
        old_rows = [r for r in old_rows if r["ist_date"] not in new_dates_set]
        all_rows = old_rows + new_rows
        all_rows.sort(key=lambda x: datetime.strptime(x["ist_date"], "%d/%m/%y"))

    # ── Write Excel ───────────────────────────────────────────
    log.info(f"Writing {len(all_rows)} total rows to Excel …")
    write_overview_excel(
        output_path    = output_file,
        data_rows      = all_rows,
        lake_root      = lake_root,
        year_filter    = year_filter,
        month_filter   = month_filter,
        total_rows_pa  = total_rows,
        total_files    = total_files,
        date_range_ist = date_range_ist,
        time_range_ist = time_range_ist,
    )

    print(f"\n{'═' * 60}")
    print(f"  ✅  Overview saved: {output_file}")
    print(f"{'═' * 60}\n")


def main() -> None:
    print("\n" + "═" * 60)
    print("  Overview Report Generator")
    print("═" * 60)

    if len(sys.argv) > 1:
        lake_root = Path(sys.argv[1]).resolve()
        if not lake_root.is_dir():
            log.error(f"Lake folder not found: {lake_root}")
            sys.exit(1)
        log.info(f"Lake folder (from CLI): {lake_root}")
    else:
        print("\n📁  Lake folder – must contain year=* partitions")
        lake_root = prompt_path("  Lake folder path: ")
        year_dirs = [d for d in lake_root.iterdir() if d.is_dir() and d.name.startswith("year=")]
        if not year_dirs:
            print(f"  ⚠️  No year= partitions found in {lake_root}. Continue? (y/n): ", end="")
            if input().strip().lower() != "y":
                sys.exit(0)

    run(lake_root)


if __name__ == "__main__":
    main()