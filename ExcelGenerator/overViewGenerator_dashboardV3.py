"""
generate_dashboard.py  —  Enterprise Edition (Averages Restored)
──────────────────────────────────────────────────────────────────────────────────
Key Updates:
- RESTORED: Dynamic dashed average lines on all charts.
- MAINTAINED: Fix for bracket syntax error on all percentage indices.
- MAINTAINED: Bound-safe column fetcher prevents IndexError on truncated Excel rows.
- MAINTAINED: Python-side date parsing normalizes regional text dates.
- MAINTAINED: Date range filter (7D/30D/ALL) with full chart & table syncing.

Usage:  python generate_dashboard.py
"""

import sys, json, urllib.request
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","--quiet"])
    import openpyxl

# ╔══════════════════════════════════════════════════════════╗
# ║  SET THESE PATHS                                         ║
# ╚══════════════════════════════════════════════════════════╝
XLSX_PATH = Path(r"Y:\Veto Logs Backup\Dashboards\OverView\overview_report.xlsx")
HTML_OUT  = Path(r"Y:\Veto Logs Backup\Dashboards\OverView\overview_dashboard.html")
# ══════════════════════════════════════════════════════════

if len(sys.argv) == 3:
    XLSX_PATH, HTML_OUT = Path(sys.argv[1]), Path(sys.argv[2])

COL = {
    'DATE': 2, 'BYTES': 3, 'ROWS': 4, 'IP_ROWS': 5,
    'DIST_IP': 6, 'DIST_IPUA': 7, 'DIST_DEV': 8, 'DIST_SESS': 9,
    'DIST_IP_R2': 10, 'DIST_IPUA_R2': 11, 
    'SESS_AVAIL': 12, 'SESS_NA': 13, 'SESS_NONE': 14,
    'PCT_IP': 15, 'PCT_IPUA': 16, 'PCT_SESS': 17, 'PCT_SESSNA': 18, 'PCT_NONE': 19
}
DATA_START_ROW = 13

def sn(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

def get_val(row, col_num):
    """Safely fetch cell value by 1-based index to prevent unexpected IndexError on short rows"""
    idx = col_num - 1
    return row[idx] if idx < len(row) else None

CHARTJS_CACHE = Path(__file__).parent / ".chartjs_cache.js"
CHARTJS_URL   = "https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"

def get_chartjs():
    if CHARTJS_CACHE.exists():
        return CHARTJS_CACHE.read_text(encoding="utf-8")
    print("  Downloading Chart.js once...", end=" ", flush=True)
    try:
        req = urllib.request.Request(CHARTJS_URL, headers={'User-Agent':'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=20) as r:
            js = r.read().decode("utf-8")
        CHARTJS_CACHE.write_text(js, encoding="utf-8")
        print(f"done ({len(js)//1024}KB cached)")
        return js
    except Exception as e:
        print(f"failed: {e}\n  Charts will be disabled.")
        return None

print(f"Reading: {XLSX_PATH}")
if not XLSX_PATH.exists():
    print(f"  ERROR: File not found: {XLSX_PATH}"); sys.exit(1)

wb = openpyxl.load_workbook(XLSX_PATH, read_only=False, data_only=True)
ws = wb.active

meta = {}
data_time_range = ""
exclude_keywords = ["report generated", "filtered to", "lake folder", "duck db", "duckdb", "parquet files", "date range", "data time range", "total rows"]
for row in ws.iter_rows(min_row=2, max_row=9, values_only=True):
    k = str(row[1]).strip() if row[1] else ''
    v = str(row[2]).strip() if row[2] else ''
    if k and v and ("data time range" in k.lower() or "date range" in k.lower()):
        data_time_range = v
    if k and k not in ('None','nan') and not any(keyword in k.lower() for keyword in exclude_keywords):
        meta[k] = v

data_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
    if ws.row_dimensions[row_idx].hidden:
        continue
    raw_d = get_val(row, COL['DATE'])
    if isinstance(raw_d, datetime):
        ds = raw_d.strftime("%Y-%m-%d")
    elif raw_d:
        ds = str(raw_d).strip()
        if not ds or 'TOTAL' in ds.upper() or ds in ('None','nan'): continue
        
        # Normalize regional variations (e.g., 14/05/26 to standard 2026-05-14) on Python side
        parsed_dt = None
        for fmt_str in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d-%m-%Y"):
            try:
                parsed_dt = datetime.strptime(ds, fmt_str)
                break
            except ValueError:
                continue
        if parsed_dt:
            ds = parsed_dt.strftime("%Y-%m-%d")
        else:
            if '/' not in ds and '-' not in ds: continue
    else:
        continue
        
    if not ds or 'TOTAL' in ds.upper(): continue

    ip  = int(sn(get_val(row, COL['IP_ROWS'])))
    sa  = int(sn(get_val(row, COL['SESS_AVAIL'])))
    sna = int(sn(get_val(row, COL['SESS_NA'])))
    sno = int(sn(get_val(row, COL['SESS_NONE'])))
    
    dist_ip = int(sn(get_val(row, COL['DIST_IP'])))
    dist_ipua = int(sn(get_val(row, COL['DIST_IPUA'])))
    
    dip_r2 = get_val(row, COL['DIST_IP_R2'])
    dipua_r2 = get_val(row, COL['DIST_IPUA_R2'])
    dist_ip_r2 = int(sn(dip_r2)) if dip_r2 not in (None, '', 'None', 'nan') else dist_ip
    dist_ipua_r2 = int(sn(dipua_r2)) if dipua_r2 not in (None, '', 'None', 'nan') else dist_ipua

    p_ip = get_val(row, COL['PCT_IP'])
    p_ipua = get_val(row, COL['PCT_IPUA'])
    p_sess = get_val(row, COL['PCT_SESS'])
    p_sessna = get_val(row, COL['PCT_SESSNA'])
    p_none = get_val(row, COL['PCT_NONE'])
    
    data_rows.append({
        "date":          ds,
        "bytes":         round(sn(get_val(row, COL['BYTES'])), 2),
        "rows":          int(sn(get_val(row, COL['ROWS']))),
        "ip_rows":       ip,
        "dist_ip":       dist_ip,
        "dist_ipua":     dist_ipua,
        "dist_dev":      int(sn(get_val(row, COL['DIST_DEV']))),
        "dist_sess":     int(sn(get_val(row, COL['DIST_SESS']))),
        "dist_ip_r2":    dist_ip_r2,
        "dist_ipua_r2":  dist_ipua_r2,
        "sess_avail":    sa,
        "sess_na":       sna,
        "sess_none":     sno,
        "pct_ip":        sn(p_ip) if (p_ip not in (None, '', 'None', 'nan')) else (round(dist_ip / ip, 6) if ip > 0 else 0),
        "pct_ipua":      sn(p_ipua) if (p_ipua not in (None, '', 'None', 'nan')) else (round(dist_ipua / ip, 6) if ip > 0 else 0),
        "pct_sess":      sn(p_sess) if (p_sess not in (None, '', 'None', 'nan')) else (round(sa / ip, 6) if ip > 0 else 0),
        "pct_sessna":    sn(p_sessna) if (p_sessna not in (None, '', 'None', 'nan')) else (round(sna / ip, 6) if ip > 0 else 0),
        "pct_none":      sn(p_none) if (p_none not in (None, '', 'None', 'nan')) else (round(sno / ip, 6) if ip > 0 else 0),
    })
wb.close()

# The latest date in the workbook is often an in-progress day. Match the Excel
# summary logic by excluding that newest partial date from the dashboard.
if data_rows:
    latest_date = max(r["date"] for r in data_rows)
    data_rows = [r for r in data_rows if r["date"] != latest_date]

data_blob = json.dumps({
    "meta": meta, "rows": data_rows,
    "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
    "source": XLSX_PATH.name,
    "data_range": data_time_range,
}, ensure_ascii=False, separators=(',', ':'))

chartjs = get_chartjs()
chartjs_tag = f"<script>{chartjs}</script>" if chartjs else "<script>window.Chart=null;</script>"

HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Overview Dashboard</title>
""" + chartjs_tag + """
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{--ink:#172033;--muted:#64748b;--panel:#fff;--line:#d8e1ee;--brand:#173b5c;--brand2:#1f6f8b;--accent:#c9a227;--bg:#eef3f8}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:var(--bg);color:var(--ink);min-height:100vh}
.topbar{background:var(--brand);color:#fff;padding:18px 30px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:14px;border-bottom:4px solid var(--accent);box-shadow:0 8px 22px rgba(15,23,42,.16)}
.title-block{display:flex;flex-direction:column;gap:3px}
.topbar h1{font-size:20px;font-weight:700;letter-spacing:0}.topbar .eyebrow{font-size:11px;text-transform:uppercase;letter-spacing:.08em;color:#cbd5e1;font-weight:700}.header-meta{display:flex;flex-direction:column;align-items:flex-end;gap:6px}.topbar .sub{font-size:12px;opacity:.82;background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.16);border-radius:999px;padding:6px 10px}.range-pill{font-size:12px;font-weight:700;background:rgba(201,162,39,.18);border:1px solid rgba(201,162,39,.45);border-radius:999px;padding:6px 10px;color:#fff}
#main{padding:22px 28px;max-width:1760px;margin:0 auto}
.meta-strip{background:var(--panel);border-radius:8px;border:1px solid var(--line);padding:12px 16px;margin-bottom:16px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px 18px;font-size:12px;color:#475569;box-shadow:0 1px 2px rgba(15,23,42,.05)}
.meta-items{display:flex;flex-wrap:wrap;gap:8px 26px;align-items:center}
.meta-strip b{color:var(--ink);font-weight:700}
.control-row{display:flex;align-items:stretch;justify-content:space-between;gap:14px;margin-bottom:18px;flex-wrap:wrap}
.cards{display:grid;grid-template-columns:repeat(4,minmax(170px,1fr));gap:12px;flex:1 1 760px}
.card{background:var(--panel);border-radius:8px;border:1px solid var(--line);padding:14px 16px;transition:all 0.3s;box-shadow:0 1px 2px rgba(15,23,42,.05);border-top:3px solid currentColor}
.clbl{font-size:11px;color:var(--muted);margin-bottom:8px;display:flex;align-items:center;gap:6px;font-weight:700;text-transform:uppercase;letter-spacing:.04em}
.cdot{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.cval{font-size:26px;font-weight:800;line-height:1.05}.csub{font-size:10.5px;color:#94a3b8;margin-top:6px}
.charts{display:grid;grid-template-columns:repeat(2, minmax(350px, 1fr));gap:16px;margin-bottom:18px}
.cbox{background:var(--panel);border-radius:8px;border:1px solid var(--line);padding:15px 17px;box-shadow:0 1px 2px rgba(15,23,42,.05)}
.cbox h3{font-size:13px;font-weight:700;margin-bottom:10px;display:flex;justify-content:space-between;color:#24324a}
.table-wrap{background:var(--panel);border-radius:8px;border:1px solid var(--line);overflow:hidden;margin-bottom:18px;box-shadow:0 2px 8px rgba(15,23,42,.06)}
.thead{padding:13px 17px;display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid var(--line);flex-wrap:wrap;gap:15px;background:#f8fafc}
.thead h3{font-size:14px;font-weight:700;color:#24324a}
.filter-group{display:flex;align-items:center;gap:8px;background:#fff;padding:12px 14px;border-radius:8px;border:1px solid var(--line);border-top:3px solid var(--accent);align-self:stretch;box-shadow:0 1px 2px rgba(15,23,42,.05)}
.filter-group input[type="date"]{border:1px solid #cbd5e1;padding:4px 6px;border-radius:4px;font-size:11px;outline:none;font-family:inherit;color:#1F3864;font-weight:500;cursor:pointer}
.filter-group span{font-size:11px;color:#888;font-weight:500}
.btn-preset{background:#fff;border:1px solid #cbd5e1;padding:4px 10px;border-radius:4px;font-size:11px;font-weight:600;cursor:pointer;color:#555;transition:all 0.2s}
.btn-preset:hover{background:#edf2f7;color:#1F3864}
.btn-preset.active{background:#1F3864;color:#fff;border-color:#1F3864}
.tbl-c{overflow-x:auto;max-height:650px;overflow-y:auto;background:#fff}
table{width:100%;border-collapse:separate;border-spacing:0;font-size:10px;table-layout:fixed}
thead{position:sticky;top:0;z-index:5}
thead tr:first-child th{padding:5px 5px;font-size:9px;font-weight:700;color:#fff;text-align:center;white-space:normal;border-right:1px solid rgba(255,255,255,.25);border-bottom:1px solid #b8c4d4}
thead tr:last-child th{padding:6px 4px;text-align:right;font-size:9px;font-weight:600;color:#334155;border-right:1px solid #cbd5e1;border-bottom:1px solid #b8c4d4;white-space:normal;line-height:1.15;background:#f8fafc;cursor:pointer;user-select:none;overflow-wrap:anywhere}
thead tr:last-child th:hover{background:#edf2f7}
thead tr:last-child th:first-child{text-align:left}
td{padding:6px 4px;text-align:right;border-right:1px solid #dbe3ef;border-bottom:1px solid #dbe3ef;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
th:first-child,td:first-child{border-left:1px solid #dbe3ef}
td:first-child{text-align:left;font-weight:500;position:sticky;left:0;background:#fff;z-index:1;width:72px}
tr:nth-child(even) td:first-child{background:#fafafa}
tr:hover td{background:#eef4ff!important}
tr.total-row td{background:#1F3864!important;color:#fff!important;font-weight:700}
tr.drop-row td{background:#fff7ed!important;color:#7c2d12!important;font-weight:700}
tr.drop-row td:first-child{text-align:left}
.chg-pos{color:#166534!important}.chg-neg{color:#b91c1c!important}.chg-flat{color:#57534e!important}
.pct{color:#185FA5;font-weight:500}
.footer{text-align:center;font-size:11px;color:#94a3b8;padding:8px 0 20px}
th.b{background:#1F3864} th.g{background:#2e6b10} th.m{background:#7c3aed} th.a{background:#9a6000} th.p{background:#4a41a8}
@media (min-width:1500px){.charts{grid-template-columns:repeat(3,minmax(0,1fr))}.tbl-c{max-height:720px}}
@media (min-width:1900px){#main{max-width:1880px}.charts{grid-template-columns:repeat(5,minmax(0,1fr))}.cbox canvas{height:190px!important}.tbl-c{max-height:760px}body{font-size:14px}}
@media (max-width:1050px){#main{padding:16px}.cards{grid-template-columns:repeat(2,minmax(0,1fr))}.charts{grid-template-columns:1fr}.topbar{padding:16px 18px}.topbar h1{font-size:18px}.header-meta{align-items:flex-start}.meta-strip{align-items:flex-start}.filter-group{width:100%;overflow-x:auto}.control-row{align-items:stretch}}
@media print{body{background:#fff}.topbar,.meta-strip,.card,.cbox,.table-wrap{box-shadow:none}.filter-group{display:none}.tbl-c{max-height:none;overflow:visible}}
</style>
</head>
<body>
<div class="topbar">
  <div class="title-block">
    <h1>Overview Dashboard Veto</h1>
  </div>
  <div class="header-meta">
    <span class="range-pill" id="rangeLbl"></span>
    <span class="sub" id="genLbl"></span>
  </div>
</div>
<div id="main">
  <div class="meta-strip" id="metaWrap">
    <div class="meta-items" id="metaStrip"></div>
  </div>
  <div class="control-row">
    <div class="cards" id="cards"></div>
    <div class="filter-group">
       <button class="btn-preset active" id="btnAll" onclick="setPreset('all')">ALL</button>
       <button class="btn-preset" id="btn30" onclick="setPreset(30)">30D</button>
       <button class="btn-preset" id="btn7" onclick="setPreset(7)">7D</button>
       <span style="margin:0 4px">|</span>
       <input type="date" id="dateStart" onchange="handleCustomDate()" />
       <span>to</span>
       <input type="date" id="dateEnd" onchange="handleCustomDate()" />
    </div>
  </div>
  
  <div class="charts">
    <div class="cbox"><h3><span>Daily Row Volume</span><span id="cH1" style="color:#aaa;font-weight:400"></span></h3><canvas id="cRows" height="170"></canvas></div>
    <div class="cbox"><h3><span>Data Density Profile (GiB)</span><span id="cH2" style="color:#aaa;font-weight:400"></span></h3><canvas id="cBytes" height="170"></canvas></div>
    <div class="cbox"><h3><span>Unique Visitors & Devices</span><span id="cH3" style="color:#aaa;font-weight:400"></span></h3><canvas id="cIP" height="170"></canvas></div>
    <div class="cbox"><h3><span>Session Available & Missing</span><span id="cH4" style="color:#aaa;font-weight:400"></span></h3><canvas id="cSessFound" height="170"></canvas></div>
    <div class="cbox"><h3><span>Session Not Found</span><span id="cH5" style="color:#aaa;font-weight:400"></span></h3><canvas id="cSessNone" height="170"></canvas></div>
  </div>

  <div class="table-wrap">
    <div class="thead">
      <h3>Day-by-Day Breakdown (Full Schema)</h3>
    </div>
    <div class="tbl-c"><table id="tbl"></table></div>
  </div>
</div>
<script>
const {meta,rows,generated,source,data_range} = """ + data_blob + """;

// Parse strictly using localized component metrics to remove browser engine discrepancies
rows.forEach(r => { 
    const [y, m, d] = r.date.split('-').map(Number);
    r.ts = new Date(y, m - 1, d).getTime(); 
});
rows.sort((a,b) => a.ts - b.ts);

document.getElementById('rangeLbl').style.display = 'none';
document.getElementById('genLbl').textContent=(data_range ? 'Available data range: '+data_range+' | ' : '')+'Last updated '+generated+' | Source: '+source;
const visibleMeta = Object.entries(meta).filter(([k]) => {
  const key = k.toLowerCase().replace(/\\s+/g, '');
  return !key.includes('duckdb') && !key.includes('totalrows') && !key.includes('datetimerange');
});
if(visibleMeta.length > 0) document.getElementById('metaStrip').innerHTML=visibleMeta.map(([k,v])=>`<span><b>${k}:</b> ${v}</span>`).join('');
else document.getElementById('metaWrap').style.display = 'none';

const fmt=(n,d=0)=>n==null||isNaN(n)?'\u2014':Number(n).toLocaleString('en-IN',{minimumFractionDigits:d,maximumFractionDigits:d});
const fmtP=n=>{ if(n==null||isNaN(n))return'\u2014'; const p=Math.abs(n)<=1?n*100:n; return p.toFixed(2)+'%'; };

let viewRows = [...rows];
let sortKey = null;
let sortAsc = true;
const chartInstances = {};

const CD={tension:.35,fill:true,pointRadius:1.5,pointHoverRadius:4,borderWidth:1.5,animation:{duration:400}};
const mkOpt=yfmt=>({responsive:true,plugins:{legend:{position:'bottom',labels:{font:{size:9.5},boxWidth:8,padding:6}}},
  scales:{x:{ticks:{font:{size:8.5},maxTicksLimit:12},grid:{color:'#f5f5f5'}},
          y:{ticks:{font:{size:8.5},callback:v=>yfmt==='gib'?v.toFixed(0)+' G':v>=1e6?(v/1e6).toFixed(1)+'M':v>=1e3?(v/1e3).toFixed(0)+'K':v},grid:{color:'#f5f5f5'}}}});

function initCharts() {
  if(typeof Chart==='undefined') return;
  const render = (id, datasets, yfmt) => { chartInstances[id] = new Chart(document.getElementById(id), { type:'line', data:{labels:[], datasets:datasets}, options:mkOpt(yfmt) }); };
  
  // Notice the x2 pairing structure. Dataset 0 is actual data, Dataset 1 is the dashed average line.
  render('cRows', [
      {label:'Total rows', borderColor:'#185FA5', backgroundColor:'rgba(24,95,165,.04)', ...CD},
      {label:'Avg', borderColor:'rgba(24,95,165,.6)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]);
  
  render('cIP', [
      {label:'Distinct cliIP', borderColor:'#4a41a8', backgroundColor:'transparent', ...CD},
      {label:'Avg', borderColor:'rgba(74,65,168,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false},
      {label:'Distinct device', borderColor:'#0F6E56', backgroundColor:'transparent', ...CD},
      {label:'Avg', borderColor:'rgba(15,110,86,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]);
  
  render('cSessFound', [
      {label:'Sess Avail', borderColor:'#2e6b10', backgroundColor:'transparent', ...CD},
      {label:'Avg', borderColor:'rgba(46,107,16,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false},
      {label:'Sess Missing', borderColor:'#9a6000', backgroundColor:'transparent', ...CD},
      {label:'Avg', borderColor:'rgba(154,96,0,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]);

  render('cSessNone', [
      {label:'Sess Not Found', borderColor:'#b71c1c', backgroundColor:'transparent', ...CD},
      {label:'Avg', borderColor:'rgba(183,28,28,.5)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ]);
  
  render('cBytes', [
      {label:'GiB Data', borderColor:'#1D9E75', backgroundColor:'rgba(29,158,117,.04)', ...CD},
      {label:'Avg', borderColor:'rgba(29,158,117,.6)', borderWidth:1.5, borderDash:[5,5], pointRadius:0, fill:false}
  ], 'gib');
}

function updateKPIs() {
    const ndays = viewRows.length || 1;
    const tBytes = viewRows.reduce((s, r) => s + r.bytes, 0);
    const tIP = viewRows.reduce((s, r) => s + r.dist_ip, 0);
    const tSess = viewRows.reduce((s, r) => s + r.dist_sess, 0);
    const tRows = viewRows.reduce((s, r) => s + r.rows, 0);

    const kpiHTML = [
        {l:'Avg Volume / Day',v:fmt(tBytes/ndays,2)+' G',clr:'#0F6E56',s:'mean daily transfer for selected range'},
        {l:'Avg Distinct IPs / Day',v:fmt(Math.round(tIP/ndays)),clr:'#4a41a8',s:'mean daily distinct IP count'},
        {l:'Avg Sessions / Day',v:fmt(Math.round(tSess/ndays)),clr:'#9a6000',s:'mean daily tracked sessions'},
        {l:'Avg Rows / Day',v:fmt(Math.round(tRows/ndays)),clr:'#b71c1c',s:'mean daily logs for range'}
    ].map(c => `<div class="card"><div class="clbl"><span class="cdot" style="background:${c.clr}"></span>${c.l}</div><div class="cval" style="color:${c.clr}">${c.v}</div><div class="csub">${c.s}</div></div>`).join('');
    document.getElementById('cards').innerHTML = kpiHTML;
}

function updateCharts() {
  if(!chartInstances['cRows']) return;
  const lbl = viewRows.map(r=>r.date);
  const len = viewRows.length || 1;
  const vTot = k => viewRows.reduce((s,r) => s+r[k], 0);

  const sync = (id, dataMaps) => {
    const ch = chartInstances[id]; 
    ch.data.labels = lbl;
    dataMaps.forEach((k, i) => { 
        ch.data.datasets[i * 2].data = viewRows.map(r => r[k]); // The actual data
        ch.data.datasets[(i * 2) + 1].data = Array(len).fill(vTot(k)/len); // The flat average line
    });
    ch.update();
  };
  
  sync('cRows', ['rows']); 
  sync('cIP', ['dist_ip', 'dist_dev']);
  sync('cSessFound', ['sess_avail', 'sess_na']); 
  sync('cSessNone', ['sess_none']); 
  sync('cBytes', ['bytes']);
  document.querySelectorAll('.cbox h3 span+span').forEach(e => e.textContent = viewRows.length + ' days visible');
}

const subH=['Date', 'GiB', 'Rows*', 'cliIP rows', 'Dist IP', 'Dist IP+UA', 'Devices', 'Sessions', 'Dist IP', 'Dist IP+UA', 'Sess Avail', 'Sess Missing', 'Sess Not Found', 'Dist IP', 'Dist IP+UA', 'Sess Avail', 'Sess Missing', 'Sess Not Found'];
const fields=['date','bytes','rows','ip_rows','dist_ip','dist_ipua','dist_dev','dist_sess','dist_ip_r2','dist_ipua_r2','sess_avail','sess_na','sess_none','pct_ip','pct_ipua','pct_sess','pct_sessna','pct_none'];
const numericFields = fields.filter(f => f !== 'date');

function getScaleStats() {
  const stats = {};
  numericFields.forEach(k => {
    const vals = viewRows.map(r => Number(r[k])).filter(v => !isNaN(v));
    stats[k] = vals.length ? {min: Math.min(...vals), max: Math.max(...vals)} : {min: 0, max: 0};
  });
  return stats;
}

function blend(a, b, t) {
  const ah = a.match(/\\w\\w/g).map(x => parseInt(x, 16));
  const bh = b.match(/\\w\\w/g).map(x => parseInt(x, 16));
  const rgb = ah.map((v, i) => Math.round(v + (bh[i] - v) * t));
  return '#' + rgb.map(v => v.toString(16).padStart(2, '0')).join('');
}

function colorScale(k, v, stats) {
  const n = Number(v);
  const s = stats[k];
  if(!s || isNaN(n) || s.max === s.min) return '';
  // Excel-style 3-color scale per column: min = red, midpoint = yellow, max = green.
  const t = Math.max(0, Math.min(1, (n - s.min) / (s.max - s.min)));
  const color = t < .5
    ? blend('f8696b', 'ffeb84', t * 2)
    : blend('ffeb84', '63be7b', (t - .5) * 2);
  return ` style="background:${color}"`;
}

function dataCell(d, k, body, cls='', stats=null) {
  return `<td${cls ? ` class="${cls}"` : ''}${stats ? colorScale(k, d[k], stats) : ''}>${body}</td>`;
}

function fmtChg(n) {
  if(n == null || isNaN(n)) return '\u2014';
  const cls = n > 0 ? 'chg-pos' : n < 0 ? 'chg-neg' : 'chg-flat';
  const sign = n > 0 ? '+' : '';
  return `<span class="${cls}">${sign}${(n * 100).toFixed(2)}%</span>`;
}

function getDropGain() {
  const byDate = [...viewRows].sort((a, b) => a.ts - b.ts);
  if(byDate.length < 5) return {};
  const target = byDate[byDate.length - 1];
  const baseRows = byDate.slice(byDate.length - 5, byDate.length - 1);
  const calc = k => {
    const avg = baseRows.reduce((s, r) => s + Number(r[k] || 0), 0) / baseRows.length;
    return avg ? (Number(target[k] || 0) / avg) - 1 : null;
  };
  return {
    bytes: calc('bytes'), rows: calc('rows'), ip_rows: calc('ip_rows'),
    dist_ip: calc('dist_ip'), dist_ipua: calc('dist_ipua'), dist_dev: calc('dist_dev'), dist_sess: calc('dist_sess'),
    dist_ip_r2: calc('dist_ip_r2'), dist_ipua_r2: calc('dist_ipua_r2'),
    sess_avail: calc('sess_avail'), sess_na: calc('sess_na'), sess_none: calc('sess_none'),
    pct_sess: calc('pct_sess'), pct_sessna: calc('pct_sessna'), pct_none: calc('pct_none')
  };
}

function renderTable() {
  const scaleStats = getScaleStats();
  let h = `<thead>
    <tr><th class="b" style="background:#f8fafc;"></th><th colspan="3" class="b">General Volumes</th><th colspan="4" class="g">Absolute Counts</th><th colspan="2" class="m">Unique Subsets</th><th colspan="3" class="a">Session Identifiers</th><th colspan="5" class="p">% Data of TOTAL Segment</th></tr>
    <tr>${subH.map((title, idx) => `<th onclick="handleSort('${fields[idx]}')">${title}${sortKey === fields[idx] ? (sortAsc ? ' \u25B2' : ' \u25BC') : ''}</th>`).join('')}</tr>
  </thead><tbody>`;

  h += viewRows.map(d => `<tr>
    <td>${d.date}</td>${dataCell(d,'bytes',fmt(d.bytes,2),'',scaleStats)}${dataCell(d,'rows',fmt(d.rows),'',scaleStats)}${dataCell(d,'ip_rows',fmt(d.ip_rows),'',scaleStats)}
    ${dataCell(d,'dist_ip',fmt(d.dist_ip),'',scaleStats)}${dataCell(d,'dist_ipua',fmt(d.dist_ipua),'',scaleStats)}${dataCell(d,'dist_dev',fmt(d.dist_dev),'',scaleStats)}${dataCell(d,'dist_sess',fmt(d.dist_sess),'',scaleStats)}
    ${dataCell(d,'dist_ip_r2',fmt(d.dist_ip_r2),'',scaleStats)}${dataCell(d,'dist_ipua_r2',fmt(d.dist_ipua_r2),'',scaleStats)}
    ${dataCell(d,'sess_avail',fmt(d.sess_avail),'',scaleStats)}${dataCell(d,'sess_na',fmt(d.sess_na),'',scaleStats)}${dataCell(d,'sess_none',fmt(d.sess_none),'',scaleStats)}
    ${dataCell(d,'pct_ip',fmtP(d.pct_ip),'pct',scaleStats)}${dataCell(d,'pct_ipua',fmtP(d.pct_ipua),'pct',scaleStats)}${dataCell(d,'pct_sess',fmtP(d.pct_sess),'pct',scaleStats)}${dataCell(d,'pct_sessna',fmtP(d.pct_sessna),'pct',scaleStats)}${dataCell(d,'pct_none',fmtP(d.pct_none),'pct',scaleStats)}
  </tr>`).join('');

  const tKeys = ["bytes","rows","ip_rows","dist_ip","dist_ipua","dist_dev","dist_sess","dist_ip_r2","dist_ipua_r2","sess_avail","sess_na","sess_none"];
  const tVals = {}; tKeys.forEach(k => tVals[k] = viewRows.reduce((sum, r) => sum + r[k], 0));
  const tPct = {
    pct_ip: tVals.ip_rows > 0 ? tVals.dist_ip / tVals.ip_rows : 0,
    pct_ipua: tVals.ip_rows > 0 ? tVals.dist_ipua / tVals.ip_rows : 0,
    pct_sess: tVals.ip_rows > 0 ? tVals.sess_avail / tVals.ip_rows : 0,
    pct_sessna: tVals.ip_rows > 0 ? tVals.sess_na / tVals.ip_rows : 0,
    pct_none: tVals.ip_rows > 0 ? tVals.sess_none / tVals.ip_rows : 0
  };
  const dropGain = getDropGain();

  h += `<tr class="total-row">
    <td>TOTAL</td><td>${fmt(tVals.bytes,2)}</td><td>${fmt(tVals.rows)}</td><td>${fmt(tVals.ip_rows)}</td>
    <td>${fmt(tVals.dist_ip)}</td><td>${fmt(tVals.dist_ipua)}</td><td>${fmt(tVals.dist_dev)}</td><td>${fmt(tVals.dist_sess)}</td>
    <td>${fmt(tVals.dist_ip_r2)}</td><td>${fmt(tVals.dist_ipua_r2)}</td>
    <td>${fmt(tVals.sess_avail)}</td><td>${fmt(tVals.sess_na)}</td><td>${fmt(tVals.sess_none)}</td>
    <td class="pct">${fmtP(tPct.pct_ip)}</td><td class="pct">${fmtP(tPct.pct_ipua)}</td><td class="pct">${fmtP(tPct.pct_sess)}</td><td class="pct">${fmtP(tPct.pct_sessna)}</td><td class="pct">${fmtP(tPct.pct_none)}</td>
  </tr>
  <tr class="drop-row">
    <td>% Drop/Gain w.r.t. last 5 days</td><td>${fmtChg(dropGain.bytes)}</td><td>${fmtChg(dropGain.rows)}</td><td>${fmtChg(dropGain.ip_rows)}</td>
    <td>${fmtChg(dropGain.dist_ip)}</td><td>${fmtChg(dropGain.dist_ipua)}</td><td>${fmtChg(dropGain.dist_dev)}</td><td>${fmtChg(dropGain.dist_sess)}</td>
    <td>${fmtChg(dropGain.dist_ip_r2)}</td><td>${fmtChg(dropGain.dist_ipua_r2)}</td>
    <td>${fmtChg(dropGain.sess_avail)}</td><td>${fmtChg(dropGain.sess_na)}</td><td>${fmtChg(dropGain.sess_none)}</td>
    <td>\u2014</td><td>\u2014</td><td>${fmtChg(dropGain.pct_sess)}</td><td>${fmtChg(dropGain.pct_sessna)}</td><td>${fmtChg(dropGain.pct_none)}</td>
  </tr></tbody>`;
  document.getElementById('tbl').innerHTML = h;
}

function handleSort(key) {
  if (sortKey === key) sortAsc = !sortAsc;
  else { sortKey = key; sortAsc = true; }
  executeSortLogic(); renderTable();
}

function executeSortLogic() {
  viewRows.sort((a, b) => {
    let va = a[sortKey], vb = b[sortKey];
    if(sortKey === 'date') return sortAsc ? a.ts - b.ts : b.ts - a.ts;
    if (typeof va === 'string') return sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
    return sortAsc ? va - vb : vb - va;
  });
}

function syncUIState(presetMode) {
    document.querySelectorAll('.btn-preset').forEach(b => b.classList.remove('active'));
    if(presetMode === 'all') document.getElementById('btnAll').classList.add('active');
    else if(presetMode === 30) document.getElementById('btn30').classList.add('active');
    else if(presetMode === 7) document.getElementById('btn7').classList.add('active');
}

function setPreset(days) {
    syncUIState(days);
    if(days === 'all' || rows.length === 0) {
        viewRows = [...rows];
        document.getElementById('dateStart').value = '';
        document.getElementById('dateEnd').value = '';
    } else {
        const maxTs = rows[rows.length - 1].ts;
        const targetTs = maxTs - (days * 86400000);
        viewRows = rows.filter(r => r.ts >= targetTs);
        
        const dtFmt = (ts) => { 
            const d = new Date(ts); 
            return `${d.getFullYear()}-${String(d.getMonth()+1).padStart(2,'0')}-${String(d.getDate()).padStart(2,'0')}`; 
        };
        document.getElementById('dateStart').value = dtFmt(viewRows[0].ts);
        document.getElementById('dateEnd').value = dtFmt(viewRows[viewRows.length-1].ts);
    }
    if(sortKey) executeSortLogic();
    renderTable(); updateCharts(); updateKPIs();
}

function handleCustomDate() {
    syncUIState(null);
    const startStr = document.getElementById('dateStart').value;
    const endStr = document.getElementById('dateEnd').value;
    
    let sTs = -Infinity;
    if (startStr) {
        const [y, m, d] = startStr.split('-').map(Number);
        sTs = new Date(y, m - 1, d).getTime();
    }
    let eTs = Infinity;
    if (endStr) {
        const [y, m, d] = endStr.split('-').map(Number);
        eTs = new Date(y, m - 1, d + 1).getTime() - 1; 
    }
    
    viewRows = rows.filter(r => r.ts >= sTs && r.ts <= eTs);
    if(sortKey) executeSortLogic();
    renderTable(); updateCharts(); updateKPIs();
}

initCharts();
updateKPIs();
updateCharts();
renderTable();
</script>
</body>
</html>"""

HTML_OUT.parent.mkdir(parents=True, exist_ok=True)
HTML_OUT.write_text(HTML, encoding='utf-8')
print(f"\n[Success] Dashboard output to: {HTML_OUT}")
